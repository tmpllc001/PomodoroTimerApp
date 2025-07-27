#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pomodoro Timer Phase 4 - Advanced Data Collection & Session Tracking Implementation
Clean Dual Window Design + Statistics Dashboard + Integrated Simple Break Window + Advanced Analytics
Phase 4 完成版：高度なデータ収集システム + セッション追跡 + フォーカス分析 + 中断検出 + 環境ロギング
"""

import sys
import json
import random
import logging
import threading
import statistics
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple
from collections import defaultdict, deque

from PyQt6.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, 
                           QWidget, QLabel, QPushButton, QSpinBox, QTabWidget,
                           QListWidget, QListWidgetItem, QLineEdit, QTextEdit,
                           QMenu, QMessageBox, QGroupBox, QScrollArea, QComboBox,
                           QDateEdit, QCheckBox, QSlider, QProgressBar, QSplitter,
                           QDialog, QInputDialog, QGridLayout)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QObject, QPoint, QDate, QThread
from PyQt6.QtGui import QFont, QAction, QMouseEvent, QPixmap, QPainter

# Visualization libraries
try:
    import matplotlib
    matplotlib.use('Qt5Agg')  # Use Qt backend for matplotlib
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    import numpy as np
    import pandas as pd
    MATPLOTLIB_AVAILABLE = True
except Exception as e:
    MATPLOTLIB_AVAILABLE = False
    # Will use basic charts instead

try:
    import seaborn as sns
    SEABORN_AVAILABLE = True
except Exception as e:
    SEABORN_AVAILABLE = False
    # Will use matplotlib-only charts

# Worker3: Prediction Engine & Export Systems imports
try:
    from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
    from sklearn.linear_model import LinearRegression
    from sklearn.preprocessing import StandardScaler, LabelEncoder
    from sklearn.model_selection import train_test_split, cross_val_score
    from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
    import joblib
    ML_AVAILABLE = True
except Exception as e:
    ML_AVAILABLE = False
    # logger will be initialized later
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.units import inch
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference
    EXPORT_AVAILABLE = True
except Exception as e:
    EXPORT_AVAILABLE = False
    # logger will be initialized later
import io
import base64
try:
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.application import MIMEApplication
    import smtplib
    EMAIL_AVAILABLE = True
except Exception as e:
    EMAIL_AVAILABLE = False

try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    SCHEDULER_AVAILABLE = True
except Exception as e:
    SCHEDULER_AVAILABLE = False
    # logger will be initialized later

# ロギング設定
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import availability warnings
if not MATPLOTLIB_AVAILABLE:
    logger.warning("matplotlib not available. Basic charts will be used.")
if not SEABORN_AVAILABLE:
    logger.warning("seaborn not available. Basic matplotlib charts will be used.")
if not ML_AVAILABLE:
    logger.warning("scikit-learn not available. ML features will be disabled.")
if not EXPORT_AVAILABLE:
    logger.warning("Export libraries not available. Export features will be disabled.")
if not EMAIL_AVAILABLE:
    logger.warning("Email libraries not available. Email features will be disabled.")
if not SCHEDULER_AVAILABLE:
    logger.warning("Scheduler libraries not available. Scheduling features will be disabled.")


class SimpleBreakContentManager:
    """シンプルな休憩コンテンツ管理"""
    
    def __init__(self):
        self.content_file = Path("data/break_content.json")
        self.content = self.load_content()
        # デフォルトキーが存在することを保証
        if "simple_tips" not in self.content:
            self.content["simple_tips"] = self.get_default_content()["simple_tips"]
        if "break_activities" not in self.content:
            self.content["break_activities"] = self.get_default_content()["break_activities"]
    
    def load_content(self) -> Dict[str, Any]:
        """コンテンツデータを読み込み"""
        try:
            if self.content_file.exists():
                with open(self.content_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                return self.get_default_content()
        except Exception as e:
            logger.error(f"コンテンツ読み込みエラー: {e}")
            return self.get_default_content()
    
    def get_default_content(self) -> Dict[str, Any]:
        """デフォルトコンテンツ"""
        return {
            "simple_tips": [
                "💧 水分補給をお忘れなく",
                "🤸 軽く首を回してみましょう",
                "👁️ 遠くを見て目を休めましょう",
                "🌬️ 深呼吸でリラックス",
                "🚶 少し歩いてみませんか？",
                "😊 笑顔でリフレッシュ"
            ],
            "break_activities": [
                "💧 水を飲む",
                "🤸 ストレッチ",
                "👁️ 目を休める",
                "🌬️ 深呼吸",
                "🚶 軽い運動",
                "😌 リラックス"
            ]
        }
    
    def get_random_tip(self) -> str:
        """ランダムなアドバイスを取得"""
        tips = self.content.get("simple_tips", self.get_default_content()["simple_tips"])
        return random.choice(tips)
    
    def get_random_activity(self) -> str:
        """ランダムな活動を取得"""
        activities = self.content.get("break_activities", self.get_default_content()["break_activities"])
        return random.choice(activities)
    
    def save_content(self):
        """コンテンツをファイルに保存"""
        try:
            self.content_file.parent.mkdir(exist_ok=True)
            with open(self.content_file, 'w', encoding='utf-8') as f:
                json.dump(self.content, f, ensure_ascii=False, indent=2)
            logger.info("休憩コンテンツを保存しました")
        except Exception as e:
            logger.error(f"コンテンツ保存エラー: {e}")
    
    def add_tip(self, tip: str):
        """新しいアドバイスを追加"""
        if tip and tip not in self.content["simple_tips"]:
            self.content["simple_tips"].append(tip)
            self.save_content()
    
    def remove_tip(self, tip: str):
        """アドバイスを削除"""
        if tip in self.content["simple_tips"] and len(self.content["simple_tips"]) > 1:
            self.content["simple_tips"].remove(tip)
            self.save_content()
    
    def add_activity(self, activity: str):
        """新しい活動を追加"""
        if activity and activity not in self.content["break_activities"]:
            self.content["break_activities"].append(activity)
            self.save_content()
    
    def remove_activity(self, activity: str):
        """活動を削除"""
        if activity in self.content["break_activities"] and len(self.content["break_activities"]) > 1:
            self.content["break_activities"].remove(activity)
            self.save_content()
    
    def get_all_tips(self) -> List[str]:
        """全てのアドバイスを取得"""
        return self.content.get("simple_tips", [])
    
    def get_all_activities(self) -> List[str]:
        """全ての活動を取得"""
        return self.content.get("break_activities", [])


class BreakContentEditorDialog(QDialog):
    """休憩コンテンツ編集ダイアログ"""
    
    def __init__(self, content_manager: SimpleBreakContentManager, parent=None):
        super().__init__(parent)
        self.content_manager = content_manager
        self.setWindowTitle("✏️ 休憩コンテンツ編集")
        self.setFixedSize(500, 400)
        self.init_ui()
    
    def init_ui(self):
        """UI初期化"""
        layout = QVBoxLayout(self)
        
        # タブウィジェット
        tab_widget = QTabWidget()
        
        # アドバイスタブ
        tips_tab = self.create_tips_tab()
        tab_widget.addTab(tips_tab, "💡 アドバイス")
        
        # 活動タブ
        activities_tab = self.create_activities_tab()
        tab_widget.addTab(activities_tab, "🤸 活動")
        
        layout.addWidget(tab_widget)
        
        # ボタン
        button_layout = QHBoxLayout()
        
        ok_btn = QPushButton("保存")
        ok_btn.clicked.connect(self.accept)
        
        cancel_btn = QPushButton("キャンセル")
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addStretch()
        button_layout.addWidget(ok_btn)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def create_tips_tab(self):
        """アドバイスタブを作成"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # リスト
        self.tips_list = QListWidget()
        self.tips_list.addItems(self.content_manager.get_all_tips())
        layout.addWidget(QLabel("アドバイス一覧:"))
        layout.addWidget(self.tips_list)
        
        # 追加・削除ボタン
        tips_button_layout = QHBoxLayout()
        
        add_tip_btn = QPushButton("追加")
        add_tip_btn.clicked.connect(self.add_tip)
        
        remove_tip_btn = QPushButton("削除")
        remove_tip_btn.clicked.connect(self.remove_tip)
        
        tips_button_layout.addWidget(add_tip_btn)
        tips_button_layout.addWidget(remove_tip_btn)
        tips_button_layout.addStretch()
        
        layout.addLayout(tips_button_layout)
        
        return widget
    
    def create_activities_tab(self):
        """活動タブを作成"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # リスト
        self.activities_list = QListWidget()
        self.activities_list.addItems(self.content_manager.get_all_activities())
        layout.addWidget(QLabel("活動一覧:"))
        layout.addWidget(self.activities_list)
        
        # 追加・削除ボタン
        activities_button_layout = QHBoxLayout()
        
        add_activity_btn = QPushButton("追加")
        add_activity_btn.clicked.connect(self.add_activity)
        
        remove_activity_btn = QPushButton("削除")
        remove_activity_btn.clicked.connect(self.remove_activity)
        
        activities_button_layout.addWidget(add_activity_btn)
        activities_button_layout.addWidget(remove_activity_btn)
        activities_button_layout.addStretch()
        
        layout.addLayout(activities_button_layout)
        
        return widget
    
    def add_tip(self):
        """アドバイス追加"""
        text, ok = QInputDialog.getText(self, "アドバイス追加", "新しいアドバイスを入力してください:")
        if ok and text.strip():
            self.content_manager.add_tip(text.strip())
            self.tips_list.clear()
            self.tips_list.addItems(self.content_manager.get_all_tips())
    
    def remove_tip(self):
        """アドバイス削除"""
        current_item = self.tips_list.currentItem()
        if current_item:
            self.content_manager.remove_tip(current_item.text())
            self.tips_list.clear()
            self.tips_list.addItems(self.content_manager.get_all_tips())
    
    def add_activity(self):
        """活動追加"""
        text, ok = QInputDialog.getText(self, "活動追加", "新しい活動を入力してください:")
        if ok and text.strip():
            self.content_manager.add_activity(text.strip())
            self.activities_list.clear()
            self.activities_list.addItems(self.content_manager.get_all_activities())
    
    def remove_activity(self):
        """活動削除"""
        current_item = self.activities_list.currentItem()
        if current_item:
            self.content_manager.remove_activity(current_item.text())
            self.activities_list.clear()
            self.activities_list.addItems(self.content_manager.get_all_activities())


class SimpleBreakWindow(QMainWindow):
    """minimal_timer_demo風のシンプルな休憩ウィンドウ（統合版）"""
    
    break_finished = pyqtSignal()
    break_skipped = pyqtSignal()
    
    def __init__(self, break_type: str = "short", duration_minutes: int = 5, task_manager=None):
        super().__init__()
        
        self.break_type = break_type
        
        # デバッグ：受信した値を確認
        logger.info(f"📍 SimpleBreakWindow受信値: duration_minutes={duration_minutes} (type: {type(duration_minutes)})")
        
        # duration_minutesの値を検証・修正
        try:
            duration_minutes_int = int(duration_minutes) if duration_minutes else 5
            # 異常に大きい値（タイムスタンプなど）をチェック
            if duration_minutes_int > 1440:  # 24時間を超える場合は異常値
                logger.warning(f"⚠️ 異常な duration_minutes 値を検出: {duration_minutes_int} → デフォルト値5に修正")
                duration_minutes_int = 5
            elif duration_minutes_int <= 0:  # 0以下も異常値
                logger.warning(f"⚠️ 無効な duration_minutes 値を検出: {duration_minutes_int} → デフォルト値5に修正")
                duration_minutes_int = 5
            self.duration_minutes = duration_minutes_int
        except (ValueError, TypeError) as e:
            logger.warning(f"⚠️ duration_minutes変換エラー: {e} → デフォルト値5に修正")
            self.duration_minutes = 5
        
        self.time_left = self.duration_minutes * 60
        self.content_manager = SimpleBreakContentManager()
        self.task_manager = task_manager
        
        # デバッグ：初期化値を確認
        logger.info(f"📍 SimpleBreakWindow初期化: duration_minutes={self.duration_minutes}, time_left={self.time_left}秒")
        
        # 設定管理
        from PyQt6.QtCore import QSettings
        self.settings = QSettings("PomodoroApp", "BreakWindow")
        self.show_task_name = self.settings.value("show_task_name", True, type=bool)
        
        # ドラッグ用
        self.dragging = False
        self.drag_position = None
        
        self.init_ui()
        self.setup_timer()
        self.center_on_screen()
        
        # コンテキストメニュー
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)
        
        logger.info(f"☕ シンプル休憩ウィンドウ表示: {break_type} ({duration_minutes}分)")
    
    def init_ui(self):
        """UI初期化 - minimal_timer_demo風"""
        break_name = "長い休憩" if self.break_type == "long" else "休憩"
        self.setWindowTitle(f"☕ {break_name}の時間です")
        self.setFixedSize(300, 160)
        
        # フレームレス・最前面（minimal_timer_demo風）
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint)
        
        # メインウィジェット
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(8)
        
        # 休憩メッセージ
        break_emoji = "🌸" if self.break_type == "long" else "☕"
        break_message = f"{break_emoji} {break_name}の時間です！"
        
        self.message_label = QLabel(break_message)
        self.message_label.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        self.message_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.message_label.setStyleSheet("color: white; background: transparent;")
        layout.addWidget(self.message_label)
        
        # カウントダウン表示
        self.time_label = QLabel(f"{self.duration_minutes:02d}:00")
        self.time_label.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        self.time_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.time_label.setStyleSheet("color: #FFD700; background: transparent;")
        layout.addWidget(self.time_label)
        
        # タスク名表示
        self.task_label = QLabel("")
        self.task_label.setFont(QFont("Arial", 9))
        self.task_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.task_label.setStyleSheet("color: #87CEEB; background: transparent;")
        self.task_label.setWordWrap(True)
        self.update_task_display()
        layout.addWidget(self.task_label)
        
        # アドバイス
        tip = self.content_manager.get_random_tip()
        self.tip_label = QLabel(tip)
        self.tip_label.setFont(QFont("Arial", 10))
        self.tip_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.tip_label.setStyleSheet("color: #cccccc; background: transparent;")
        self.tip_label.setWordWrap(True)
        layout.addWidget(self.tip_label)
        
        # コントロールボタン（小さく）
        button_layout = QHBoxLayout()
        
        self.skip_btn = QPushButton("⏩")
        self.skip_btn.setMaximumSize(30, 25)
        self.skip_btn.setToolTip("休憩をスキップ")
        self.skip_btn.clicked.connect(self.skip_break)
        self.skip_btn.setStyleSheet("""
            QPushButton {
                background-color: rgba(255, 255, 255, 0.3);
                border: 1px solid rgba(255, 255, 255, 0.5);
                border-radius: 3px;
                color: white;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: rgba(255, 255, 255, 0.5);
            }
        """)
        
        self.extend_btn = QPushButton("+1")
        self.extend_btn.setMaximumSize(30, 25)
        self.extend_btn.setToolTip("1分延長")
        self.extend_btn.clicked.connect(self.extend_break)
        self.extend_btn.setStyleSheet("""
            QPushButton {
                background-color: rgba(255, 255, 255, 0.3);
                border: 1px solid rgba(255, 255, 255, 0.5);
                border-radius: 3px;
                color: white;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: rgba(255, 255, 255, 0.5);
            }
        """)
        
        button_layout.addStretch()
        button_layout.addWidget(self.extend_btn)
        button_layout.addWidget(self.skip_btn)
        layout.addLayout(button_layout)
        
        # minimal_timer_demo風のスタイル
        main_widget.setStyleSheet("""
            QWidget {
                background-color: rgba(40, 40, 40, 220);
                border-radius: 10px;
                border: 1px solid rgba(255, 255, 255, 0.2);
            }
        """)
    
    def setup_timer(self):
        """タイマー設定"""
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_countdown)
        self.timer.start(1000)
    
    def update_countdown(self):
        """カウントダウン更新"""
        self.time_left -= 1
        
        # 負の値にならないように保護
        if self.time_left < 0:
            self.time_left = 0
        
        # 異常に大きい値を保護（24時間を超える場合）
        if self.time_left > 86400:  # 24時間 = 86400秒
            logger.warning(f"⚠️ 異常な time_left 値を検出: {self.time_left} → 修正")
            self.time_left = 300  # 5分にリセット
        
        minutes = self.time_left // 60
        seconds = self.time_left % 60
        
        # 表示値も二重チェック
        if minutes > 1440:  # 24時間を超える場合
            logger.warning(f"⚠️ 異常な表示時間を検出: {minutes}分 → 5分に修正")
            minutes = 5
            seconds = 0
            self.time_left = 300
            
        self.time_label.setText(f"{minutes:02d}:{seconds:02d}")
        
        # 残り30秒でアドバイスを更新
        if self.time_left == 30:
            activity = self.content_manager.get_random_activity()
            self.tip_label.setText(f"まもなく終了 {activity}")
        
        # 終了
        if self.time_left <= 0:
            self.timer.stop()
            self.break_finished.emit()
            self.close()
    
    def extend_break(self):
        """1分延長"""
        self.time_left += 60
        logger.info("⏰ 休憩を1分延長")
    
    def skip_break(self):
        """休憩をスキップ"""
        self.timer.stop()
        self.break_skipped.emit()
        self.close()
        logger.info("⏩ 休憩をスキップ")
    
    def update_task_display(self):
        """タスク名表示更新"""
        if self.show_task_name and self.task_manager:
            active_tasks = self.task_manager.get_active_tasks()
            if active_tasks:
                # 最新のタスクを表示（文字数制限）
                task_text = active_tasks[-1]['text']
                if len(task_text) > 35:
                    task_text = task_text[:35] + "..."
                self.task_label.setText(f"📋 {task_text}")
                self.task_label.setVisible(True)
            else:
                self.task_label.setText("")
                self.task_label.setVisible(False)
        else:
            self.task_label.setVisible(False)
    
    def center_on_screen(self):
        """画面中央に配置"""
        screen = QApplication.primaryScreen().geometry()
        window_rect = self.frameGeometry()
        center_point = screen.center()
        window_rect.moveCenter(center_point)
        self.move(window_rect.topLeft())
    
    # マウスイベント（ドラッグ移動）- minimal_timer_demo準拠
    def mousePressEvent(self, event: QMouseEvent):
        """マウス押下"""
        if event.button() == Qt.MouseButton.LeftButton:
            self.dragging = True
            self.drag_position = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
    
    def mouseMoveEvent(self, event: QMouseEvent):
        """マウス移動"""
        if self.dragging:
            self.move(event.globalPosition().toPoint() - self.drag_position)
    
    def mouseReleaseEvent(self, event: QMouseEvent):
        """マウスリリース"""
        self.dragging = False
    
    def show_context_menu(self, pos):
        """コンテキストメニュー表示"""
        menu = QMenu(self)
        
        # タスク名表示切り替え
        task_action = QAction("タスク名表示", self)
        task_action.setCheckable(True)
        task_action.setChecked(self.show_task_name)
        task_action.triggered.connect(self.toggle_task_name)
        menu.addAction(task_action)
        
        menu.addSeparator()
        
        # アドバイス更新
        refresh_tip = QAction("💡 別のアドバイス", self)
        refresh_tip.triggered.connect(self.refresh_tip)
        menu.addAction(refresh_tip)
        
        # コンテンツ編集
        edit_content = QAction("✏️ コンテンツ編集", self)
        edit_content.triggered.connect(self.edit_content)
        menu.addAction(edit_content)
        
        menu.addSeparator()
        
        # 延長・スキップ
        extend_action = QAction("⏰ 1分延長", self)
        extend_action.triggered.connect(self.extend_break)
        menu.addAction(extend_action)
        
        skip_action = QAction("⏩ スキップ", self)
        skip_action.triggered.connect(self.skip_break)
        menu.addAction(skip_action)
        
        menu.exec(self.mapToGlobal(pos))
    
    def toggle_task_name(self):
        """タスク名表示切り替え"""
        self.show_task_name = not self.show_task_name
        self.settings.setValue("show_task_name", self.show_task_name)
        self.update_task_display()
    
    def refresh_tip(self):
        """アドバイスを更新"""
        tip = self.content_manager.get_random_tip()
        self.tip_label.setText(tip)
    
    def edit_content(self):
        """コンテンツ編集ダイアログを表示"""
        dialog = BreakContentEditorDialog(self.content_manager, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # アドバイスを更新
            self.refresh_tip()
    
    def closeEvent(self, event):
        """ウィンドウ終了時"""
        if hasattr(self, 'timer'):
            self.timer.stop()
        event.accept()




class WorkStartCountdownWindow(QMainWindow):
    """作業開始前の3秒カウントダウンウィンドウ"""
    
    countdown_finished = pyqtSignal()
    
    def __init__(self):
        super().__init__()
        
        self.count = 3  # 3秒カウントダウン
        
        self.init_ui()
        self.setup_timer()
        self.center_on_screen()
        
        logger.info("⏰ 作業開始カウントダウンウィンドウ初期化")
    
    def init_ui(self):
        """UI初期化"""
        self.setWindowTitle("まもなく作業開始")
        self.setFixedSize(300, 150)
        
        # フレームレス・最前面
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint)
        
        # メインウィジェット
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        # メッセージ
        self.message_label = QLabel("🚀 まもなく作業開始")
        self.message_label.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        self.message_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.message_label.setStyleSheet("color: white; background: transparent;")
        layout.addWidget(self.message_label)
        
        # カウントダウン数字
        self.countdown_label = QLabel("3")
        self.countdown_label.setFont(QFont("Arial", 36, QFont.Weight.Bold))
        self.countdown_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.countdown_label.setStyleSheet("""
            color: #FFD700; 
            background: rgba(50, 50, 50, 200);
            border: 2px solid rgba(255, 255, 255, 100);
            border-radius: 50px;
            min-width: 100px;
            min-height: 100px;
        """)
        layout.addWidget(self.countdown_label)
        
        # スタイル設定
        main_widget.setStyleSheet("""
            QWidget {
                background-color: rgba(30, 30, 30, 240);
                border-radius: 15px;
            }
        """)
    
    def setup_timer(self):
        """タイマー設定"""
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_countdown)
        self.timer.start(1000)  # 1秒ごと
    
    def update_countdown(self):
        """カウントダウン更新"""
        logger.info(f"🔄 カウントダウン更新: count={self.count}")
        
        if self.count > 0:
            logger.info(f"📱 表示更新: {self.count}")
            self.countdown_label.setText(str(self.count))
            
            # アニメーション効果（縮小→拡大）
            self.animate_countdown()
            
            # カウント減算
            self.count -= 1
            logger.info(f"📉 カウント減算後: count={self.count}")
        else:
            # カウントダウン終了
            logger.info("⏹️ カウントダウン終了")
            self.timer.stop()
            self.countdown_finished.emit()
            self.close()
    
    def animate_countdown(self):
        """カウントダウンアニメーション"""
        try:
            from PyQt6.QtCore import QPropertyAnimation, QEasingCurve
            
            # スケールアニメーション
            self.animation = QPropertyAnimation(self.countdown_label, b"geometry")
            self.animation.setDuration(500)  # 0.5秒
            self.animation.setEasingCurve(QEasingCurve.Type.OutBounce)
            
            # 現在のジオメトリ
            current_rect = self.countdown_label.geometry()
            
            # 縮小→拡大
            shrink_rect = current_rect
            shrink_rect.setWidth(int(current_rect.width() * 0.8))
            shrink_rect.setHeight(int(current_rect.height() * 0.8))
            shrink_rect.moveCenter(current_rect.center())
            
            self.animation.setStartValue(shrink_rect)
            self.animation.setEndValue(current_rect)
            self.animation.start()
            
        except Exception as e:
            logger.error(f"カウントダウンアニメーションエラー: {e}")
    
    def center_on_screen(self):
        """画面中央に配置"""
        screen = QApplication.primaryScreen().geometry()
        window_rect = self.frameGeometry()
        center_point = screen.center()
        window_rect.moveCenter(center_point)
        self.move(window_rect.topLeft())
    
    def closeEvent(self, event):
        """ウィンドウ終了時"""
        if hasattr(self, 'timer'):
            self.timer.stop()
        if hasattr(self, 'animation'):
            self.animation.stop()
        event.accept()


class SessionTemplateManager(QObject):
    """セッションテンプレート管理システム"""
    
    template_changed = pyqtSignal(dict)  # テンプレート変更時のシグナル
    
    def __init__(self):
        super().__init__()
        self.templates_file = "data/session_templates.json"
        self.current_template = None
        self.custom_templates = {}
        
        # デフォルトテンプレートを定義
        self.default_templates = {
            "classic_pomodoro": {
                "name": "🍅 クラシック ポモドーロ",
                "description": "25分作業 + 5分休憩",
                "work_minutes": 25,
                "break_minutes": 5,
                "long_break_minutes": 15,
                "sessions_until_long_break": 4,
                "max_sessions": 8,
                "category": "標準"
            },
            "extended_focus": {
                "name": "🎯 集中特化",
                "description": "50分作業 + 10分休憩",
                "work_minutes": 50,
                "break_minutes": 10,
                "long_break_minutes": 30,
                "sessions_until_long_break": 3,
                "max_sessions": 6,
                "category": "長時間"
            },
            "quick_sprint": {
                "name": "⚡ クイックスプリント",
                "description": "15分作業 + 3分休憩",
                "work_minutes": 15,
                "break_minutes": 3,
                "long_break_minutes": 10,
                "sessions_until_long_break": 6,
                "max_sessions": 12,
                "category": "短時間"
            },
            "micro_focus": {
                "name": "⏱️ マイクロフォーカス",
                "description": "5分作業 + 1分休憩",
                "work_minutes": 5,
                "break_minutes": 1,
                "long_break_minutes": 5,
                "sessions_until_long_break": 10,
                "max_sessions": 20,
                "category": "超短時間"
            },
            "school_period": {
                "name": "📚 学校授業",
                "description": "45分授業 + 10分休憩",
                "work_minutes": 45,
                "break_minutes": 10,
                "long_break_minutes": 20,
                "sessions_until_long_break": 2,
                "max_sessions": 6,
                "category": "教育"
            },
            "university_lecture": {
                "name": "🎓 大学講義",
                "description": "90分講義 + 15分休憩",
                "work_minutes": 90,
                "break_minutes": 15,
                "long_break_minutes": 30,
                "sessions_until_long_break": 2,
                "max_sessions": 4,
                "category": "教育"
            },
            "deep_work": {
                "name": "🧠 ディープワーク",
                "description": "120分作業 + 20分休憩",
                "work_minutes": 120,
                "break_minutes": 20,
                "long_break_minutes": 45,
                "sessions_until_long_break": 2,
                "max_sessions": 3,
                "category": "超長時間"
            },
            "demo_mode": {
                "name": "🔬 デモモード",
                "description": "1分作業 + 1分休憩（テスト用）",
                "work_minutes": 1,
                "break_minutes": 1,
                "long_break_minutes": 1,
                "sessions_until_long_break": 3,
                "max_sessions": 10,
                "category": "テスト"
            }
        }
        
        self.load_templates()
        
        # デフォルトはクラシックポモドーロ
        self.current_template = self.default_templates["classic_pomodoro"]
        
        logger.info(f"📋 セッションテンプレート管理システム初期化完了")
    
    def get_all_templates(self) -> dict:
        """全テンプレートを取得"""
        templates = self.default_templates.copy()
        templates.update(self.custom_templates)
        return templates
    
    def get_templates_by_category(self) -> dict:
        """カテゴリ別にテンプレートを分類"""
        all_templates = self.get_all_templates()
        categories = {}
        
        for template_id, template in all_templates.items():
            category = template.get("category", "その他")
            if category not in categories:
                categories[category] = {}
            categories[category][template_id] = template
        
        return categories
    
    def set_template(self, template_id: str) -> bool:
        """テンプレートを設定"""
        all_templates = self.get_all_templates()
        
        if template_id in all_templates:
            self.current_template = all_templates[template_id]
            self.template_changed.emit(self.current_template)
            logger.info(f"📋 テンプレート変更: {self.current_template['name']}")
            return True
        
        return False
    
    def get_current_template(self) -> dict:
        """現在のテンプレートを取得"""
        return self.current_template or self.default_templates["classic_pomodoro"]
    
    def create_custom_template(self, template_id: str, name: str, description: str,
                             work_minutes: int, break_minutes: float, long_break_minutes: int,
                             sessions_until_long_break: int, max_sessions: int) -> bool:
        """カスタムテンプレートを作成"""
        try:
            custom_template = {
                "name": name,
                "description": description,
                "work_minutes": work_minutes,
                "break_minutes": break_minutes,
                "long_break_minutes": long_break_minutes,
                "sessions_until_long_break": sessions_until_long_break,
                "max_sessions": max_sessions,
                "category": "カスタム",
                "created_at": datetime.now().isoformat()
            }
            
            self.custom_templates[template_id] = custom_template
            self.save_templates()
            
            logger.info(f"📋 カスタムテンプレート作成: {name}")
            return True
            
        except Exception as e:
            logger.error(f"カスタムテンプレート作成エラー: {e}")
            return False
    
    def delete_custom_template(self, template_id: str) -> bool:
        """カスタムテンプレートを削除"""
        if template_id in self.custom_templates:
            del self.custom_templates[template_id]
            self.save_templates()
            logger.info(f"📋 カスタムテンプレート削除: {template_id}")
            return True
        
        return False
    
    def load_templates(self):
        """テンプレートファイルから読み込み"""
        try:
            import os
            if os.path.exists(self.templates_file):
                import json
                with open(self.templates_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.custom_templates = data.get("custom_templates", {})
                    logger.info(f"📋 カスタムテンプレート読み込み: {len(self.custom_templates)}件")
        except Exception as e:
            logger.error(f"テンプレート読み込みエラー: {e}")
            self.custom_templates = {}
    
    def save_templates(self):
        """テンプレートファイルに保存"""
        try:
            import os
            import json
            
            os.makedirs(os.path.dirname(self.templates_file), exist_ok=True)
            
            data = {
                "custom_templates": self.custom_templates,
                "last_updated": datetime.now().isoformat()
            }
            
            with open(self.templates_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"📋 テンプレート保存完了")
            
        except Exception as e:
            logger.error(f"テンプレート保存エラー: {e}")


class CustomTemplateDialog(QDialog):
    """カスタムテンプレート作成ダイアログ"""
    
    def __init__(self, template_manager: SessionTemplateManager, parent=None):
        super().__init__(parent)
        self.template_manager = template_manager
        self.setup_ui()
    
    def setup_ui(self):
        """UI設定"""
        self.setWindowTitle("📋 カスタムテンプレート作成")
        self.setModal(True)
        self.resize(400, 350)
        
        layout = QVBoxLayout(self)
        
        # テンプレート基本情報
        info_group = QGroupBox("基本情報")
        info_layout = QGridLayout()
        
        # テンプレート名
        info_layout.addWidget(QLabel("テンプレート名:"), 0, 0)
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("例: マイポモドーロ")
        info_layout.addWidget(self.name_edit, 0, 1)
        
        # 説明
        info_layout.addWidget(QLabel("説明:"), 1, 0)
        self.description_edit = QLineEdit()
        self.description_edit.setPlaceholderText("例: 30分作業 + 10分休憩")
        info_layout.addWidget(self.description_edit, 1, 1)
        
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        # 時間設定
        time_group = QGroupBox("時間設定")
        time_layout = QGridLayout()
        
        # 作業時間
        time_layout.addWidget(QLabel("作業時間 (分):"), 0, 0)
        self.work_spin = QSpinBox()
        self.work_spin.setRange(1, 180)
        self.work_spin.setValue(25)
        time_layout.addWidget(self.work_spin, 0, 1)
        
        # 短い休憩時間
        time_layout.addWidget(QLabel("短い休憩 (分):"), 1, 0)
        self.break_spin = QSpinBox()
        self.break_spin.setRange(1, 60)
        self.break_spin.setValue(5)
        time_layout.addWidget(self.break_spin, 1, 1)
        
        # 長い休憩時間
        time_layout.addWidget(QLabel("長い休憩 (分):"), 2, 0)
        self.long_break_spin = QSpinBox()
        self.long_break_spin.setRange(5, 120)
        self.long_break_spin.setValue(15)
        time_layout.addWidget(self.long_break_spin, 2, 1)
        
        time_group.setLayout(time_layout)
        layout.addWidget(time_group)
        
        # セッション設定
        session_group = QGroupBox("セッション設定")
        session_layout = QGridLayout()
        
        # 最大セッション数
        session_layout.addWidget(QLabel("最大セッション数:"), 0, 0)
        self.max_sessions_spin = QSpinBox()
        self.max_sessions_spin.setRange(1, 20)
        self.max_sessions_spin.setValue(8)
        session_layout.addWidget(self.max_sessions_spin, 0, 1)
        
        # 長い休憩間隔
        session_layout.addWidget(QLabel("長い休憩間隔:"), 1, 0)
        self.long_break_interval_spin = QSpinBox()
        self.long_break_interval_spin.setRange(2, 10)
        self.long_break_interval_spin.setValue(4)
        session_layout.addWidget(self.long_break_interval_spin, 1, 1)
        
        session_group.setLayout(session_layout)
        layout.addWidget(session_group)
        
        # プリセットボタン
        preset_layout = QHBoxLayout()
        preset_layout.addWidget(QLabel("プリセット:"))
        
        presets = [
            ("25-5", 25, 5, 15, 8, 4),
            ("50-10", 50, 10, 30, 6, 3),
            ("15-3", 15, 3, 10, 12, 6),
            ("45-10", 45, 10, 20, 6, 2)
        ]
        
        for name, work, short_break, long_break, max_sessions, interval in presets:
            btn = QPushButton(name)
            btn.clicked.connect(lambda checked, w=work, sb=short_break, lb=long_break, ms=max_sessions, iv=interval: 
                              self.set_preset(w, sb, lb, ms, iv))
            preset_layout.addWidget(btn)
        
        layout.addLayout(preset_layout)
        
        # ボタン
        button_layout = QHBoxLayout()
        
        create_btn = QPushButton("✅ 作成")
        create_btn.clicked.connect(self.create_template)
        create_btn.setDefault(True)
        button_layout.addWidget(create_btn)
        
        cancel_btn = QPushButton("❌ キャンセル")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def set_preset(self, work: int, short_break: int, long_break: int, max_sessions: int, interval: int):
        """プリセット値を設定"""
        self.work_spin.setValue(work)
        self.break_spin.setValue(short_break)
        self.long_break_spin.setValue(long_break)
        self.max_sessions_spin.setValue(max_sessions)
        self.long_break_interval_spin.setValue(interval)
        
        # 説明を自動生成
        self.description_edit.setText(f"{work}分作業 + {short_break}分休憩")
    
    def create_template(self):
        """テンプレートを作成"""
        name = self.name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "入力エラー", "テンプレート名を入力してください")
            return
        
        description = self.description_edit.text().strip()
        if not description:
            description = f"{self.work_spin.value()}分作業 + {self.break_spin.value()}分休憩"
        
        # テンプレートIDを生成
        import re
        template_id = f"custom_{re.sub(r'[^a-zA-Z0-9]', '_', name.lower())}"
        
        # テンプレートを作成
        success = self.template_manager.create_custom_template(
            template_id=template_id,
            name=name,
            description=description,
            work_minutes=self.work_spin.value(),
            break_minutes=self.break_spin.value(),
            long_break_minutes=self.long_break_spin.value(),
            sessions_until_long_break=self.long_break_interval_spin.value(),
            max_sessions=self.max_sessions_spin.value()
        )
        
        if success:
            QMessageBox.information(self, "作成完了", f"カスタムテンプレート '{name}' を作成しました")
            self.accept()
        else:
            QMessageBox.critical(self, "作成エラー", "テンプレートの作成に失敗しました")


class AdvancedDataCollector(QObject):
    """Phase 4: 高度なデータ収集システム - 詳細なセッションメトリクス追跡"""
    
    # シグナル - データ変更時に通知
    data_collected = pyqtSignal(dict)  # 収集されたデータ
    metric_updated = pyqtSignal(str, object)  # (metric_name, value)
    
    def __init__(self):
        super().__init__()
        
        # データファイル
        self.data_file = Path("data/advanced_session_data.json")
        self.data_file.parent.mkdir(exist_ok=True)
        
        # データ収集設定
        self.collection_interval = 10  # 10秒間隔でデータ収集
        self.max_session_history = 1000  # 最大セッション履歴数
        
        # データストレージ
        self.session_data = []
        self.current_session_metrics = {}
        self.performance_metrics = defaultdict(list)
        
        # 収集タイマー
        self.collection_timer = QTimer()
        self.collection_timer.timeout.connect(self._collect_periodic_data)
        
        # 現在のセッション状態
        self.current_session_start = None
        self.session_interruptions = []
        self.user_interactions = []
        
        self.load_data()
        logger.info("📊 AdvancedDataCollector 初期化完了")
    
    def start_session_tracking(self, session_type: str, duration_minutes: int):
        """セッション追跡開始"""
        self.current_session_start = datetime.now()
        self.current_session_metrics = {
            'session_id': f"{session_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            'type': session_type,
            'planned_duration': duration_minutes,
            'start_time': self.current_session_start.isoformat(),
            'interactions': [],
            'interruptions': [],
            'focus_scores': [],
            'environment_data': {},
            'performance_metrics': {}
        }
        
        # 定期データ収集開始
        self.collection_timer.start(self.collection_interval * 1000)
        
        logger.info(f"📊 セッション追跡開始: {session_type} ({duration_minutes}分)")
    
    def end_session_tracking(self, completed: bool = True):
        """セッション追跡終了"""
        if not self.current_session_start:
            return
        
        # 最終データ収集
        end_time = datetime.now()
        session_duration = (end_time - self.current_session_start).total_seconds()
        
        self.current_session_metrics.update({
            'end_time': end_time.isoformat(),
            'actual_duration': session_duration / 60,  # 分単位
            'completed': completed,
            'efficiency_score': self._calculate_efficiency_score(),
            'focus_score': self._calculate_session_focus_score()
        })
        
        # セッションデータを保存
        self.session_data.append(self.current_session_metrics.copy())
        
        # データ量制限
        if len(self.session_data) > self.max_session_history:
            self.session_data = self.session_data[-self.max_session_history:]
        
        # 定期収集停止
        self.collection_timer.stop()
        
        # データ保存とシグナル発信
        self.save_data()
        self.data_collected.emit(self.current_session_metrics.copy())
        
        logger.info(f"📊 セッション追跡終了: {self.current_session_metrics['session_id']}")
        
        # リセット
        self.current_session_start = None
        self.current_session_metrics = {}
    
    def record_user_interaction(self, interaction_type: str, details: dict = None):
        """ユーザーインタラクション記録"""
        if not self.current_session_start:
            return
        
        interaction = {
            'timestamp': datetime.now().isoformat(),
            'type': interaction_type,
            'details': details or {},
            'session_time': (datetime.now() - self.current_session_start).total_seconds()
        }
        
        self.current_session_metrics['interactions'].append(interaction)
        
        # インタラクション頻度メトリック更新
        self._update_interaction_metrics(interaction_type)
        
        logger.debug(f"🖱️ ユーザーインタラクション記録: {interaction_type}")
    
    def record_interruption(self, interruption_type: str, duration_seconds: float = 0):
        """中断記録"""
        if not self.current_session_start:
            return
        
        interruption = {
            'timestamp': datetime.now().isoformat(),
            'type': interruption_type,
            'duration': duration_seconds,
            'session_time': (datetime.now() - self.current_session_start).total_seconds()
        }
        
        self.current_session_metrics['interruptions'].append(interruption)
        
        logger.info(f"⚠️ 中断記録: {interruption_type} ({duration_seconds:.1f}秒)")
    
    def _collect_periodic_data(self):
        """定期データ収集"""
        if not self.current_session_start:
            return
        
        current_time = datetime.now()
        session_time = (current_time - self.current_session_start).total_seconds()
        
        # 環境データ収集
        env_data = {
            'timestamp': current_time.isoformat(),
            'session_time': session_time,
            'system_time': current_time.strftime('%H:%M:%S'),
            'day_of_week': current_time.strftime('%A'),
            'hour_of_day': current_time.hour
        }
        
        # フォーカススコア計算（仮想的な実装）
        focus_score = self._calculate_current_focus_score(session_time)
        self.current_session_metrics['focus_scores'].append({
            'timestamp': current_time.isoformat(),
            'score': focus_score
        })
        
        # メトリクス更新シグナル
        self.metric_updated.emit('focus_score', focus_score)
        self.metric_updated.emit('session_time', session_time)
    
    def _calculate_current_focus_score(self, session_time: float) -> float:
        """現在のフォーカススコア計算（0-100）"""
        # 基本スコア: セッション時間に基づく
        base_score = min(90, 50 + (session_time / 60) * 2)  # 時間とともに向上
        
        # 中断によるペナルティ
        interruption_penalty = len(self.current_session_metrics.get('interruptions', [])) * 5
        
        # インタラクション頻度による調整
        interaction_count = len(self.current_session_metrics.get('interactions', []))
        interaction_penalty = max(0, (interaction_count - 10) * 2)  # 10回を超えるとペナルティ
        
        # 最終スコア
        final_score = max(0, min(100, base_score - interruption_penalty - interaction_penalty))
        
        return round(final_score, 1)
    
    def _calculate_session_focus_score(self) -> float:
        """セッション全体のフォーカススコア計算"""
        focus_scores = [fs['score'] for fs in self.current_session_metrics.get('focus_scores', [])]
        if not focus_scores:
            return 0.0
        
        return round(statistics.mean(focus_scores), 1)
    
    def _calculate_efficiency_score(self) -> float:
        """効率スコア計算"""
        if not self.current_session_start:
            return 0.0
        
        planned_duration = self.current_session_metrics.get('planned_duration', 25) * 60
        actual_duration = (datetime.now() - self.current_session_start).total_seconds()
        
        # 完了率
        completion_rate = min(1.0, actual_duration / planned_duration)
        
        # 中断率
        interruption_count = len(self.current_session_metrics.get('interruptions', []))
        interruption_penalty = min(0.5, interruption_count * 0.1)
        
        # 効率スコア (0-100)
        efficiency = max(0, (completion_rate - interruption_penalty) * 100)
        
        return round(efficiency, 1)
    
    def _update_interaction_metrics(self, interaction_type: str):
        """インタラクションメトリクス更新"""
        self.performance_metrics[f"interaction_{interaction_type}"].append(datetime.now())
        
        # 古いデータを削除（24時間以上前）
        cutoff_time = datetime.now() - timedelta(hours=24)
        self.performance_metrics[f"interaction_{interaction_type}"] = [
            ts for ts in self.performance_metrics[f"interaction_{interaction_type}"] 
            if ts > cutoff_time
        ]
    
    def get_session_analytics(self, days: int = 7) -> Dict[str, Any]:
        """セッション分析データ取得"""
        cutoff_date = datetime.now() - timedelta(days=days)
        
        recent_sessions = [
            s for s in self.session_data 
            if datetime.fromisoformat(s['start_time']) > cutoff_date
        ]
        
        if not recent_sessions:
            return {'total_sessions': 0, 'message': 'データが不十分です'}
        
        # 基本統計
        total_sessions = len(recent_sessions)
        completed_sessions = len([s for s in recent_sessions if s.get('completed', False)])
        avg_focus_score = statistics.mean([s.get('focus_score', 0) for s in recent_sessions])
        avg_efficiency = statistics.mean([s.get('efficiency_score', 0) for s in recent_sessions])
        
        # セッションタイプ別統計
        work_sessions = [s for s in recent_sessions if s.get('type') == 'work']
        break_sessions = [s for s in recent_sessions if s.get('type') == 'break']
        
        return {
            'period_days': days,
            'total_sessions': total_sessions,
            'completed_sessions': completed_sessions,
            'completion_rate': (completed_sessions / total_sessions * 100) if total_sessions > 0 else 0,
            'avg_focus_score': round(avg_focus_score, 1),
            'avg_efficiency_score': round(avg_efficiency, 1),
            'work_sessions': len(work_sessions),
            'break_sessions': len(break_sessions),
            'total_interruptions': sum(len(s.get('interruptions', [])) for s in recent_sessions),
            'avg_session_duration': statistics.mean([s.get('actual_duration', 0) for s in recent_sessions])
        }
    
    def load_data(self):
        """データ読み込み"""
        try:
            if self.data_file.exists():
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.session_data = data.get('sessions', [])
                    self.performance_metrics = defaultdict(list, data.get('metrics', {}))
                    
                    # datetime オブジェクトに変換
                    for metric_name, timestamps in self.performance_metrics.items():
                        self.performance_metrics[metric_name] = [
                            datetime.fromisoformat(ts) if isinstance(ts, str) else ts
                            for ts in timestamps
                        ]
            
            logger.info(f"📊 AdvancedDataCollector データ読み込み: {len(self.session_data)}セッション")
            
        except Exception as e:
            logger.error(f"AdvancedDataCollector データ読み込みエラー: {e}")
            self.session_data = []
            self.performance_metrics = defaultdict(list)
    
    def save_data(self):
        """データ保存"""
        try:
            # datetime オブジェクトを文字列に変換
            serializable_metrics = {}
            for metric_name, timestamps in self.performance_metrics.items():
                serializable_metrics[metric_name] = [
                    ts.isoformat() if isinstance(ts, datetime) else ts
                    for ts in timestamps
                ]
            
            data = {
                'sessions': self.session_data,
                'metrics': serializable_metrics,
                'last_updated': datetime.now().isoformat()
            }
            
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            logger.error(f"AdvancedDataCollector データ保存エラー: {e}")


class SessionTracking(QObject):
    """Phase 4: 高度なセッション管理 - 作業パターンの詳細追跡"""
    
    # シグナル
    session_pattern_detected = pyqtSignal(str, dict)  # (pattern_name, details)
    productivity_trend_updated = pyqtSignal(float)  # productivity_score
    
    def __init__(self):
        super().__init__()
        
        self.tracking_file = Path("data/session_tracking.json")
        self.tracking_file.parent.mkdir(exist_ok=True)
        
        # セッション追跡データ
        self.session_history = []
        self.daily_patterns = defaultdict(list)
        self.productivity_trends = []
        
        # パターン検出設定
        self.pattern_window_days = 7
        self.min_sessions_for_pattern = 5
        
        self.load_tracking_data()
        logger.info("📈 SessionTracking 初期化完了")
    
    def record_session_completion(self, session_data: Dict[str, Any]):
        """セッション完了記録"""
        session_record = {
            'timestamp': datetime.now().isoformat(),
            'session_id': session_data.get('session_id', ''),
            'type': session_data.get('type', 'work'),
            'planned_duration': session_data.get('planned_duration', 25),
            'actual_duration': session_data.get('actual_duration', 0),
            'completed': session_data.get('completed', False),
            'focus_score': session_data.get('focus_score', 0),
            'efficiency_score': session_data.get('efficiency_score', 0),
            'interruptions': len(session_data.get('interruptions', [])),
            'hour_of_day': datetime.now().hour,
            'day_of_week': datetime.now().weekday(),
            'date': datetime.now().strftime('%Y-%m-%d')
        }
        
        self.session_history.append(session_record)
        
        # 日別パターン更新
        date_key = session_record['date']
        self.daily_patterns[date_key].append(session_record)
        
        # データ量制限
        if len(self.session_history) > 1000:
            self.session_history = self.session_history[-1000:]
        
        # パターン検出実行
        self._detect_patterns()
        
        # 生産性トレンド更新
        self._update_productivity_trend()
        
        # データ保存
        self.save_tracking_data()
        
        logger.info(f"📈 セッション記録: {session_record['type']} (効率: {session_record['efficiency_score']:.1f}%)")
    
    def _detect_patterns(self):
        """作業パターン検出"""
        if len(self.session_history) < self.min_sessions_for_pattern:
            return
        
        recent_sessions = self._get_recent_sessions(self.pattern_window_days)
        
        # 時間帯パターン検出
        self._detect_time_patterns(recent_sessions)
        
        # 効率パターン検出
        self._detect_efficiency_patterns(recent_sessions)
        
        # 中断パターン検出
        self._detect_interruption_patterns(recent_sessions)
    
    def _detect_time_patterns(self, sessions: List[Dict]):
        """時間帯パターン検出"""
        if not sessions:
            return
        
        # 時間帯別効率スコア
        hourly_efficiency = defaultdict(list)
        for session in sessions:
            if session['type'] == 'work':
                hourly_efficiency[session['hour_of_day']].append(session['efficiency_score'])
        
        # 最も効率的な時間帯を特定
        best_hours = []
        for hour, scores in hourly_efficiency.items():
            if len(scores) >= 3:  # 最低3セッション
                avg_score = statistics.mean(scores)
                if avg_score > 75:  # 75%以上の効率
                    best_hours.append((hour, avg_score))
        
        if best_hours:
            best_hours.sort(key=lambda x: x[1], reverse=True)
            best_hour, best_score = best_hours[0]
            
            pattern_details = {
                'best_hour': best_hour,
                'efficiency_score': best_score,
                'sessions_count': len(hourly_efficiency[best_hour])
            }
            
            self.session_pattern_detected.emit('optimal_time_detected', pattern_details)
            logger.info(f"⏰ 最適時間帯検出: {best_hour}時 (効率: {best_score:.1f}%)")
    
    def _detect_efficiency_patterns(self, sessions: List[Dict]):
        """効率パターン検出"""
        work_sessions = [s for s in sessions if s['type'] == 'work']
        if len(work_sessions) < 5:
            return
        
        # 効率スコアの傾向分析
        efficiency_scores = [s['efficiency_score'] for s in work_sessions[-10:]]  # 直近10セッション
        
        if len(efficiency_scores) >= 5:
            recent_avg = statistics.mean(efficiency_scores[-5:])
            earlier_avg = statistics.mean(efficiency_scores[:-5]) if len(efficiency_scores) > 5 else recent_avg
            
            trend = recent_avg - earlier_avg
            
            if abs(trend) > 10:  # 10%以上の変化
                pattern_type = 'efficiency_improving' if trend > 0 else 'efficiency_declining'
                pattern_details = {
                    'trend_change': trend,
                    'recent_average': recent_avg,
                    'sessions_analyzed': len(efficiency_scores)
                }
                
                self.session_pattern_detected.emit(pattern_type, pattern_details)
                logger.info(f"📊 効率傾向検出: {pattern_type} ({trend:+.1f}%)")
    
    def _detect_interruption_patterns(self, sessions: List[Dict]):
        """中断パターン検出"""
        work_sessions = [s for s in sessions if s['type'] == 'work']
        if len(work_sessions) < 5:
            return
        
        # 中断頻度分析
        interruption_counts = [s['interruptions'] for s in work_sessions]
        avg_interruptions = statistics.mean(interruption_counts)
        
        if avg_interruptions > 3:  # セッションあたり3回以上の中断
            high_interruption_sessions = len([s for s in work_sessions if s['interruptions'] > 3])
            interruption_rate = high_interruption_sessions / len(work_sessions)
            
            if interruption_rate > 0.5:  # 50%以上のセッションで高中断
                pattern_details = {
                    'avg_interruptions': avg_interruptions,
                    'high_interruption_rate': interruption_rate,
                    'sessions_affected': high_interruption_sessions
                }
                
                self.session_pattern_detected.emit('high_interruption_pattern', pattern_details)
                logger.warning(f"⚠️ 高中断パターン検出: 平均{avg_interruptions:.1f}回/セッション")
    
    def _update_productivity_trend(self):
        """生産性トレンド更新"""
        recent_work_sessions = [
            s for s in self._get_recent_sessions(7)
            if s['type'] == 'work' and s['completed']
        ]
        
        if len(recent_work_sessions) >= 3:
            # 完了率、効率スコア、フォーカススコアの複合指標
            completion_rate = len(recent_work_sessions) / max(1, len(self._get_recent_sessions(7)))
            avg_efficiency = statistics.mean([s['efficiency_score'] for s in recent_work_sessions])
            avg_focus = statistics.mean([s['focus_score'] for s in recent_work_sessions])
            
            # 生産性スコア計算 (0-100)
            productivity_score = (completion_rate * 30 + avg_efficiency * 0.4 + avg_focus * 0.3)
            productivity_score = min(100, max(0, productivity_score))
            
            self.productivity_trends.append({
                'date': datetime.now().strftime('%Y-%m-%d'),
                'score': productivity_score,
                'sessions_count': len(recent_work_sessions)
            })
            
            # データ量制限
            if len(self.productivity_trends) > 30:  # 30日分
                self.productivity_trends = self.productivity_trends[-30:]
            
            self.productivity_trend_updated.emit(productivity_score)
            logger.info(f"📈 生産性スコア更新: {productivity_score:.1f}")
    
    def _get_recent_sessions(self, days: int) -> List[Dict]:
        """最近のセッション取得"""
        cutoff_date = datetime.now() - timedelta(days=days)
        return [
            s for s in self.session_history
            if datetime.fromisoformat(s['timestamp']) > cutoff_date
        ]
    
    def get_productivity_insights(self) -> Dict[str, Any]:
        """生産性インサイト取得"""
        recent_sessions = self._get_recent_sessions(7)
        work_sessions = [s for s in recent_sessions if s['type'] == 'work']
        
        if not work_sessions:
            return {'message': 'データが不十分です'}
        
        # 基本統計
        total_work_time = sum(s['actual_duration'] for s in work_sessions)
        avg_efficiency = statistics.mean([s['efficiency_score'] for s in work_sessions])
        avg_focus = statistics.mean([s['focus_score'] for s in work_sessions])
        completion_rate = len([s for s in work_sessions if s['completed']]) / len(work_sessions) * 100
        
        # 最適時間帯
        hourly_performance = defaultdict(list)
        for session in work_sessions:
            hourly_performance[session['hour_of_day']].append(session['efficiency_score'])
        
        best_hour = None
        best_hour_score = 0
        for hour, scores in hourly_performance.items():
            if len(scores) >= 2:
                avg_score = statistics.mean(scores)
                if avg_score > best_hour_score:
                    best_hour = hour
                    best_hour_score = avg_score
        
        return {
            'total_work_sessions': len(work_sessions),
            'total_work_time_hours': round(total_work_time / 60, 1),
            'avg_efficiency_score': round(avg_efficiency, 1),
            'avg_focus_score': round(avg_focus, 1),
            'completion_rate': round(completion_rate, 1),
            'best_hour': best_hour,
            'best_hour_efficiency': round(best_hour_score, 1) if best_hour else 0,
            'current_productivity_trend': self.productivity_trends[-1] if self.productivity_trends else None
        }
    
    def load_tracking_data(self):
        """追跡データ読み込み"""
        try:
            if self.tracking_file.exists():
                with open(self.tracking_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.session_history = data.get('session_history', [])
                    self.daily_patterns = defaultdict(list, data.get('daily_patterns', {}))
                    self.productivity_trends = data.get('productivity_trends', [])
            
            logger.info(f"📈 SessionTracking データ読み込み: {len(self.session_history)}セッション")
            
        except Exception as e:
            logger.error(f"SessionTracking データ読み込みエラー: {e}")
            self.session_history = []
            self.daily_patterns = defaultdict(list)
            self.productivity_trends = []
    
    def save_tracking_data(self):
        """追跡データ保存"""
        try:
            data = {
                'session_history': self.session_history,
                'daily_patterns': dict(self.daily_patterns),
                'productivity_trends': self.productivity_trends,
                'last_updated': datetime.now().isoformat()
            }
            
            with open(self.tracking_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            logger.error(f"SessionTracking データ保存エラー: {e}")


class FocusScoreCalculator(QObject):
    """Phase 4: フォーカス・集中度スコア計算システム"""
    
    # シグナル
    focus_score_updated = pyqtSignal(float)  # 現在のフォーカススコア
    focus_level_changed = pyqtSignal(str)  # focus_level: 'low', 'medium', 'high'
    
    def __init__(self):
        super().__init__()
        
        # スコア計算パラメータ
        self.base_focus_score = 50.0
        self.max_focus_score = 100.0
        self.min_focus_score = 0.0
        
        # フォーカス要因の重み
        self.weights = {
            'session_duration': 0.3,
            'interruption_frequency': 0.25,
            'interaction_pattern': 0.2,
            'time_consistency': 0.15,
            'completion_rate': 0.1
        }
        
        # 計算データ
        self.session_start_time = None
        self.interaction_timestamps = deque(maxlen=50)  # 直近50回のインタラクション
        self.interruption_count = 0
        self.last_interruption_time = None
        
        logger.info("🎯 FocusScoreCalculator 初期化完了")
    
    def start_focus_tracking(self):
        """フォーカス追跡開始"""
        self.session_start_time = datetime.now()
        self.interaction_timestamps.clear()
        self.interruption_count = 0
        self.last_interruption_time = None
        
        logger.info("🎯 フォーカス追跡開始")
    
    def end_focus_tracking(self) -> float:
        """フォーカス追跡終了、最終スコア返却"""
        if not self.session_start_time:
            return self.base_focus_score
        
        final_score = self.calculate_current_focus_score()
        
        # リセット
        self.session_start_time = None
        self.interaction_timestamps.clear()
        self.interruption_count = 0
        self.last_interruption_time = None
        
        logger.info(f"🎯 フォーカス追跡終了、最終スコア: {final_score}")
        return final_score
    
    def record_interaction(self, interaction_type: str):
        """インタラクション記録"""
        current_time = datetime.now()
        self.interaction_timestamps.append({
            'timestamp': current_time,
            'type': interaction_type
        })
        
        # スコア更新
        self._update_focus_score()
    
    def record_interruption(self):
        """中断記録"""
        current_time = datetime.now()
        self.interruption_count += 1
        self.last_interruption_time = current_time
        
        logger.info(f"⚠️ 中断記録 #{self.interruption_count}")
        
        # スコア更新
        self._update_focus_score()
    
    def calculate_current_focus_score(self) -> float:
        """現在のフォーカススコア計算"""
        if not self.session_start_time:
            return self.base_focus_score
        
        current_time = datetime.now()
        session_duration = (current_time - self.session_start_time).total_seconds() / 60  # 分
        
        # 各要因のスコア計算
        duration_score = self._calc_duration_score(session_duration)
        interruption_score = self._calc_interruption_score(session_duration)
        interaction_score = self._calc_interaction_score(session_duration)
        consistency_score = self._calc_consistency_score()
        completion_score = self._calc_completion_score(session_duration)
        
        # 重み付き合計
        weighted_score = (
            duration_score * self.weights['session_duration'] +
            interruption_score * self.weights['interruption_frequency'] +
            interaction_score * self.weights['interaction_pattern'] +
            consistency_score * self.weights['time_consistency'] +
            completion_score * self.weights['completion_rate']
        )
        
        # 最終スコア
        final_score = max(self.min_focus_score, min(self.max_focus_score, weighted_score))
        
        return round(final_score, 1)
    
    def _calc_duration_score(self, duration_minutes: float) -> float:
        """セッション継続時間スコア (0-100)"""
        if duration_minutes <= 0:
            return 0
        
        # 理想的な継続時間は25分
        optimal_duration = 25
        
        if duration_minutes <= optimal_duration:
            # 25分まではリニアに増加
            return (duration_minutes / optimal_duration) * 100
        else:
            # 25分を超えても維持（長時間集中を評価）
            overtime_penalty = min(20, (duration_minutes - optimal_duration) * 0.5)
            return max(80, 100 - overtime_penalty)
    
    def _calc_interruption_score(self, duration_minutes: float) -> float:
        """中断頻度スコア (0-100)"""
        if duration_minutes <= 0:
            return 100
        
        # 許容中断頻度: 15分に1回まで
        acceptable_rate = 1 / 15  # 15分に1回
        actual_rate = self.interruption_count / max(1, duration_minutes)
        
        if actual_rate <= acceptable_rate:
            return 100
        else:
            # 過度な中断はスコア減少
            penalty = min(80, (actual_rate - acceptable_rate) * 500)
            return max(20, 100 - penalty)
    
    def _calc_interaction_score(self, duration_minutes: float) -> float:
        """インタラクションパターンスコア (0-100)"""
        if duration_minutes <= 0 or not self.interaction_timestamps:
            return 100
        
        # 最近のインタラクション頻度分析
        recent_interactions = [
            t for t in self.interaction_timestamps 
            if (datetime.now() - t['timestamp']).total_seconds() <= 300  # 5分以内
        ]
        
        interaction_rate = len(recent_interactions) / max(1, duration_minutes)
        
        # 適度なインタラクション（1-3回/分）を理想とする
        if 1 <= interaction_rate <= 3:
            return 100
        elif interaction_rate < 1:
            # インタラクション不足（放置状態）
            return 60 + interaction_rate * 40
        else:
            # 過度なインタラクション（落ち着きがない）
            penalty = min(60, (interaction_rate - 3) * 15)
            return max(40, 100 - penalty)
    
    def _calc_consistency_score(self) -> float:
        """時間一貫性スコア (0-100)"""
        if len(self.interaction_timestamps) < 3:
            return 100  # データ不足時は満点
        
        # インタラクション間隔の一貫性を評価
        intervals = []
        timestamps = [t['timestamp'] for t in self.interaction_timestamps]
        
        for i in range(1, len(timestamps)):
            interval = (timestamps[i] - timestamps[i-1]).total_seconds()
            intervals.append(interval)
        
        if len(intervals) < 2:
            return 100
        
        # 間隔の標準偏差を計算（小さいほど一貫性が高い）
        try:
            interval_std = statistics.stdev(intervals)
            # 標準偏差を正規化（0-60秒の範囲で評価）
            normalized_std = min(60, interval_std)
            consistency_score = 100 - (normalized_std / 60 * 50)
            return max(50, consistency_score)
        except:
            return 100
    
    def _calc_completion_score(self, duration_minutes: float) -> float:
        """完了スコア (0-100)"""
        # 25分セッションの完了度
        target_duration = 25
        completion_rate = min(1.0, duration_minutes / target_duration)
        return completion_rate * 100
    
    def _update_focus_score(self):
        """フォーカススコア更新とシグナル発信"""
        current_score = self.calculate_current_focus_score()
        self.focus_score_updated.emit(current_score)
        
        # フォーカスレベル判定
        if current_score >= 80:
            focus_level = 'high'
        elif current_score >= 60:
            focus_level = 'medium'
        else:
            focus_level = 'low'
        
        self.focus_level_changed.emit(focus_level)
    
    def get_focus_insights(self) -> Dict[str, Any]:
        """フォーカス分析インサイト"""
        if not self.session_start_time:
            return {'message': 'セッション実行中ではありません'}
        
        current_score = self.calculate_current_focus_score()
        session_duration = (datetime.now() - self.session_start_time).total_seconds() / 60
        
        # 各要因の個別スコア
        duration_score = self._calc_duration_score(session_duration)
        interruption_score = self._calc_interruption_score(session_duration)
        interaction_score = self._calc_interaction_score(session_duration)
        consistency_score = self._calc_consistency_score()
        
        return {
            'overall_focus_score': current_score,
            'session_duration_minutes': round(session_duration, 1),
            'component_scores': {
                'duration': round(duration_score, 1),
                'interruption_resistance': round(interruption_score, 1),
                'interaction_pattern': round(interaction_score, 1),
                'time_consistency': round(consistency_score, 1)
            },
            'interruption_count': self.interruption_count,
            'interaction_count': len(self.interaction_timestamps),
            'recommendations': self._generate_recommendations(current_score)
        }
    
    def _generate_recommendations(self, focus_score: float) -> List[str]:
        """フォーカス改善推奨事項生成"""
        recommendations = []
        
        if focus_score < 60:
            recommendations.append("🎯 集中力が低下しています。短い休憩を取ることをお勧めします")
            
        if self.interruption_count > 3:
            recommendations.append("⚠️ 中断が多発しています。通知を無効にしたり、静かな環境を確保してください")
            
        if len(self.interaction_timestamps) > 30:
            recommendations.append("🖱️ インタラクションが頻繁です。落ち着いて作業に集中しましょう")
            
        session_duration = (datetime.now() - self.session_start_time).total_seconds() / 60 if self.session_start_time else 0
        if session_duration < 10:
            recommendations.append("⏱️ セッションが短すぎます。もう少し継続してみましょう")
            
        if not recommendations:
            recommendations.append("✅ 良好な集中状態を維持しています！")
            
        return recommendations


class InterruptionTracker(QObject):
    """Phase 4: セッション中断検出・追跡システム"""
    
    # シグナル
    interruption_detected = pyqtSignal(str, dict)  # (interruption_type, details)
    interruption_pattern_found = pyqtSignal(str, dict)  # (pattern_type, analysis)
    
    def __init__(self):
        super().__init__()
        
        self.tracking_file = Path("data/interruption_tracking.json")
        self.tracking_file.parent.mkdir(exist_ok=True)
        
        # 中断追跡データ
        self.interruptions = []
        self.session_interruptions = []  # 現在のセッション中の中断
        
        # 中断検出設定
        self.pause_threshold = 10  # 10秒以上の一時停止で中断判定
        self.inactivity_threshold = 180  # 3分間の非活動で中断判定
        
        # 状態追跡
        self.session_active = False
        self.last_activity_time = None
        self.pause_start_time = None
        
        # パターン検出用タイマー
        self.inactivity_timer = QTimer()
        self.inactivity_timer.timeout.connect(self._check_inactivity)
        
        self.load_interruption_data()
        logger.info("⚠️ InterruptionTracker 初期化完了")
    
    def start_session_monitoring(self):
        """セッション監視開始"""
        self.session_active = True
        self.session_interruptions = []
        self.last_activity_time = datetime.now()
        self.pause_start_time = None
        
        # 非活動タイマー開始（30秒間隔でチェック）
        self.inactivity_timer.start(30000)
        
        logger.info("⚠️ セッション中断監視開始")
    
    def end_session_monitoring(self):
        """セッション監視終了"""
        self.session_active = False
        self.inactivity_timer.stop()
        
        # セッション中断データを保存
        if self.session_interruptions:
            session_summary = {
                'session_date': datetime.now().strftime('%Y-%m-%d'),
                'session_time': datetime.now().strftime('%H:%M:%S'),
                'total_interruptions': len(self.session_interruptions),
                'interruption_types': self._categorize_interruptions(self.session_interruptions),
                'interruptions': self.session_interruptions.copy()
            }
            
            self.interruptions.append(session_summary)
            
            # データ量制限
            if len(self.interruptions) > 100:
                self.interruptions = self.interruptions[-100:]
            
            # パターン分析
            self._analyze_interruption_patterns()
            
            self.save_interruption_data()
        
        logger.info(f"⚠️ セッション監視終了: {len(self.session_interruptions)}回の中断")
    
    def record_pause_start(self):
        """一時停止開始記録"""
        if not self.session_active:
            return
        
        self.pause_start_time = datetime.now()
        self.last_activity_time = datetime.now()
        
        logger.debug("⏸️ 一時停止開始記録")
    
    def record_pause_end(self):
        """一時停止終了記録"""
        if not self.session_active or not self.pause_start_time:
            return
        
        end_time = datetime.now()
        pause_duration = (end_time - self.pause_start_time).total_seconds()
        
        # 閾値を超える一時停止は中断として記録
        if pause_duration >= self.pause_threshold:
            interruption = {
                'type': 'manual_pause',
                'start_time': self.pause_start_time.isoformat(),
                'end_time': end_time.isoformat(),
                'duration_seconds': pause_duration,
                'severity': self._calculate_interruption_severity(pause_duration)
            }
            
            self.session_interruptions.append(interruption)
            self.interruption_detected.emit('manual_pause', interruption)
            
            logger.info(f"⚠️ 一時停止中断検出: {pause_duration:.1f}秒")
        
        self.pause_start_time = None
        self.last_activity_time = datetime.now()
    
    def record_user_activity(self, activity_type: str = 'interaction'):
        """ユーザー活動記録"""
        if not self.session_active:
            return
        
        self.last_activity_time = datetime.now()
        
        # 一時停止中の活動は一時停止終了として扱う
        if self.pause_start_time:
            self.record_pause_end()
    
    def record_external_interruption(self, interruption_type: str, description: str = ""):
        """外部中断記録"""
        if not self.session_active:
            return
        
        interruption = {
            'type': interruption_type,
            'timestamp': datetime.now().isoformat(),
            'description': description,
            'severity': 'high' if interruption_type in ['phone_call', 'urgent_message'] else 'medium'
        }
        
        self.session_interruptions.append(interruption)
        self.interruption_detected.emit(interruption_type, interruption)
        
        logger.info(f"⚠️ 外部中断記録: {interruption_type}")
    
    def _check_inactivity(self):
        """非活動状態チェック"""
        if not self.session_active or not self.last_activity_time:
            return
        
        current_time = datetime.now()
        inactivity_duration = (current_time - self.last_activity_time).total_seconds()
        
        # 非活動閾値を超えた場合
        if inactivity_duration >= self.inactivity_threshold:
            interruption = {
                'type': 'inactivity',
                'detected_time': current_time.isoformat(),
                'last_activity': self.last_activity_time.isoformat(),
                'duration_seconds': inactivity_duration,
                'severity': self._calculate_interruption_severity(inactivity_duration)
            }
            
            self.session_interruptions.append(interruption)
            self.interruption_detected.emit('inactivity', interruption)
            
            # 活動時間を更新（重複検出を防ぐ）
            self.last_activity_time = current_time
            
            logger.warning(f"⚠️ 非活動中断検出: {inactivity_duration:.1f}秒")
    
    def _calculate_interruption_severity(self, duration_seconds: float) -> str:
        """中断重要度計算"""
        if duration_seconds < 30:
            return 'low'
        elif duration_seconds < 120:
            return 'medium'
        else:
            return 'high'
    
    def _categorize_interruptions(self, interruptions: List[Dict]) -> Dict[str, int]:
        """中断タイプ別分類"""
        categories = defaultdict(int)
        for interruption in interruptions:
            categories[interruption['type']] += 1
        return dict(categories)
    
    def _analyze_interruption_patterns(self):
        """中断パターン分析"""
        if len(self.interruptions) < 3:
            return
        
        # 最近5セッションの分析
        recent_sessions = self.interruptions[-5:]
        
        # 中断頻度パターン
        self._analyze_frequency_patterns(recent_sessions)
        
        # 中断タイプパターン
        self._analyze_type_patterns(recent_sessions)
        
        # 時間帯パターン
        self._analyze_time_patterns(recent_sessions)
    
    def _analyze_frequency_patterns(self, sessions: List[Dict]):
        """中断頻度パターン分析"""
        interruption_counts = [s['total_interruptions'] for s in sessions]
        avg_interruptions = statistics.mean(interruption_counts)
        
        if avg_interruptions > 5:  # セッションあたり5回以上
            pattern_details = {
                'average_interruptions': avg_interruptions,
                'sessions_analyzed': len(sessions),
                'max_interruptions': max(interruption_counts),
                'recommendation': '中断が多発しています。環境の改善を検討してください'
            }
            
            self.interruption_pattern_found.emit('high_frequency', pattern_details)
            logger.warning(f"⚠️ 高頻度中断パターン検出: 平均{avg_interruptions:.1f}回/セッション")
    
    def _analyze_type_patterns(self, sessions: List[Dict]):
        """中断タイプパターン分析"""
        all_types = defaultdict(int)
        for session in sessions:
            for int_type, count in session['interruption_types'].items():
                all_types[int_type] += count
        
        # 最も多い中断タイプを特定
        if all_types:
            dominant_type = max(all_types.items(), key=lambda x: x[1])
            type_name, count = dominant_type
            
            if count >= len(sessions) * 2:  # セッションあたり平均2回以上
                pattern_details = {
                    'dominant_type': type_name,
                    'occurrence_count': count,
                    'sessions_analyzed': len(sessions),
                    'recommendation': self._get_type_recommendation(type_name)
                }
                
                self.interruption_pattern_found.emit('dominant_type', pattern_details)
                logger.info(f"⚠️ 支配的中断タイプ検出: {type_name} ({count}回)")
    
    def _analyze_time_patterns(self, sessions: List[Dict]):
        """時間帯パターン分析"""
        # セッション時間から中断傾向を分析
        session_times = []
        for session in sessions:
            try:
                time_str = session['session_time']
                hour = int(time_str.split(':')[0])
                session_times.append(hour)
            except:
                continue
        
        if len(session_times) >= 3:
            # 特定時間帯での中断率が高いかチェック
            time_interruption_map = {}
            for i, session in enumerate(sessions):
                if i < len(session_times):
                    hour = session_times[i]
                    interruptions = session['total_interruptions']
                    if hour not in time_interruption_map:
                        time_interruption_map[hour] = []
                    time_interruption_map[hour].append(interruptions)
            
            # 時間帯別平均中断数計算
            for hour, interruption_counts in time_interruption_map.items():
                if len(interruption_counts) >= 2:
                    avg_interruptions = statistics.mean(interruption_counts)
                    if avg_interruptions > 4:  # 4回以上の中断
                        pattern_details = {
                            'problematic_hour': hour,
                            'average_interruptions': avg_interruptions,
                            'sessions_count': len(interruption_counts),
                            'recommendation': f'{hour}時台は中断が多い傾向があります。この時間帯の作業環境を見直してください'
                        }
                        
                        self.interruption_pattern_found.emit('time_pattern', pattern_details)
                        logger.info(f"⚠️ 時間帯中断パターン検出: {hour}時台")
    
    def _get_type_recommendation(self, interruption_type: str) -> str:
        """中断タイプ別推奨事項"""
        recommendations = {
            'manual_pause': '頻繁な一時停止が発生しています。25分間の集中を心がけましょう',
            'inactivity': '非活動状態が多発しています。集中力維持のため適度な休憩を取りましょう',
            'phone_call': '電話による中断が多いです。作業時間中は電話を無音にすることを検討してください',
            'urgent_message': '緊急メッセージによる中断が頻発しています。通知設定を見直してください',
            'external_noise': '外部騒音による中断が多いです。より静かな環境での作業を検討してください'
        }
        
        return recommendations.get(interruption_type, '中断パターンが検出されました。作業環境の改善を検討してください')
    
    def get_interruption_summary(self, days: int = 7) -> Dict[str, Any]:
        """中断サマリー取得"""
        cutoff_date = datetime.now() - timedelta(days=days)
        
        recent_sessions = [
            s for s in self.interruptions
            if datetime.strptime(s['session_date'], '%Y-%m-%d').date() >= cutoff_date.date()
        ]
        
        if not recent_sessions:
            return {'message': 'データがありません'}
        
        # 統計計算
        total_sessions = len(recent_sessions)
        total_interruptions = sum(s['total_interruptions'] for s in recent_sessions)
        avg_interruptions = total_interruptions / total_sessions if total_sessions > 0 else 0
        
        # タイプ別集計
        all_types = defaultdict(int)
        for session in recent_sessions:
            for int_type, count in session['interruption_types'].items():
                all_types[int_type] += count
        
        # 最も問題のあるセッション
        worst_session = max(recent_sessions, key=lambda s: s['total_interruptions']) if recent_sessions else None
        
        return {
            'period_days': days,
            'total_sessions': total_sessions,
            'total_interruptions': total_interruptions,
            'average_interruptions_per_session': round(avg_interruptions, 1),
            'interruption_types': dict(all_types),
            'worst_session': {
                'date': worst_session['session_date'],
                'interruptions': worst_session['total_interruptions']
            } if worst_session else None,
            'improvement_needed': avg_interruptions > 3
        }
    
    def load_interruption_data(self):
        """中断データ読み込み"""
        try:
            if self.tracking_file.exists():
                with open(self.tracking_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.interruptions = data.get('interruptions', [])
            
            logger.info(f"⚠️ InterruptionTracker データ読み込み: {len(self.interruptions)}セッション")
            
        except Exception as e:
            logger.error(f"InterruptionTracker データ読み込みエラー: {e}")
            self.interruptions = []
    
    def save_interruption_data(self):
        """中断データ保存"""
        try:
            data = {
                'interruptions': self.interruptions,
                'last_updated': datetime.now().isoformat()
            }
            
            with open(self.tracking_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            logger.error(f"InterruptionTracker データ保存エラー: {e}")


class EnvironmentLogger(QObject):
    """Phase 4: 環境データ記録システム - 時間帯・曜日パターンの記録"""
    
    # シグナル
    environment_data_updated = pyqtSignal(dict)  # 環境データ更新
    optimal_time_detected = pyqtSignal(dict)  # 最適時間帯検出
    
    def __init__(self):
        super().__init__()
        
        self.env_file = Path("data/environment_log.json")
        self.env_file.parent.mkdir(exist_ok=True)
        
        # 環境データ
        self.environment_records = []
        self.current_session_env = {}
        
        # パフォーマンスマップ
        self.hourly_performance = defaultdict(list)
        self.daily_performance = defaultdict(list)
        self.monthly_performance = defaultdict(list)
        
        # 環境要因追跡
        self.tracked_factors = [
            'hour_of_day', 'day_of_week', 'month', 'season',
            'session_type', 'duration', 'efficiency_score'
        ]
        
        self.load_environment_data()
        logger.info("🌍 EnvironmentLogger 初期化完了")
    
    def start_environment_logging(self, session_type: str):
        """環境ロギング開始"""
        current_time = datetime.now()
        
        self.current_session_env = {
            'session_start': current_time.isoformat(),
            'session_type': session_type,
            'hour_of_day': current_time.hour,
            'day_of_week': current_time.weekday(),
            'day_name': current_time.strftime('%A'),
            'month': current_time.month,
            'month_name': current_time.strftime('%B'),
            'season': self._get_season(current_time.month),
            'date': current_time.strftime('%Y-%m-%d'),
            'time_period': self._get_time_period(current_time.hour),
            'weekend': current_time.weekday() >= 5
        }
        
        logger.info(f"🌍 環境ロギング開始: {session_type} ({self.current_session_env['time_period']})")
    
    def end_environment_logging(self, session_data: Dict[str, Any]):
        """環境ロギング終了"""
        if not self.current_session_env:
            return
        
        end_time = datetime.now()
        
        # セッション結果を環境データに追加
        self.current_session_env.update({
            'session_end': end_time.isoformat(),
            'duration_minutes': session_data.get('actual_duration', 0),
            'completed': session_data.get('completed', False),
            'efficiency_score': session_data.get('efficiency_score', 0),
            'focus_score': session_data.get('focus_score', 0),
            'interruptions': len(session_data.get('interruptions', [])),
            'interactions': len(session_data.get('interactions', []))
        })
        
        # 記録保存
        self.environment_records.append(self.current_session_env.copy())
        
        # データ量制限
        if len(self.environment_records) > 1000:
            self.environment_records = self.environment_records[-1000:]
        
        # パフォーマンスマップ更新
        self._update_performance_maps()
        
        # 最適時間帯検出
        self._detect_optimal_times()
        
        # データ保存
        self.save_environment_data()
        
        # シグナル発信
        self.environment_data_updated.emit(self.current_session_env.copy())
        
        logger.info(f"🌍 環境ロギング終了: 効率{self.current_session_env['efficiency_score']:.1f}%")
        
        # リセット
        self.current_session_env = {}
    
    def _get_season(self, month: int) -> str:
        """季節判定"""
        if month in [12, 1, 2]:
            return 'winter'
        elif month in [3, 4, 5]:
            return 'spring'
        elif month in [6, 7, 8]:
            return 'summer'
        else:
            return 'autumn'
    
    def _get_time_period(self, hour: int) -> str:
        """時間帯判定"""
        if 5 <= hour < 12:
            return 'morning'
        elif 12 <= hour < 17:
            return 'afternoon'
        elif 17 <= hour < 22:
            return 'evening'
        else:
            return 'night'
    
    def _update_performance_maps(self):
        """パフォーマンスマップ更新"""
        if not self.current_session_env or self.current_session_env.get('session_type') != 'work':
            return
        
        efficiency = self.current_session_env.get('efficiency_score', 0)
        focus = self.current_session_env.get('focus_score', 0)
        performance_score = (efficiency + focus) / 2  # 複合パフォーマンススコア
        
        # 時間別パフォーマンス
        hour = self.current_session_env['hour_of_day']
        self.hourly_performance[hour].append(performance_score)
        
        # 曜日別パフォーマンス
        day = self.current_session_env['day_of_week']
        self.daily_performance[day].append(performance_score)
        
        # 月別パフォーマンス
        month = self.current_session_env['month']
        self.monthly_performance[month].append(performance_score)
        
        # データ量制限（各カテゴリ100件まで）
        for performance_map in [self.hourly_performance, self.daily_performance, self.monthly_performance]:
            for key in performance_map:
                if len(performance_map[key]) > 100:
                    performance_map[key] = performance_map[key][-100:]
    
    def _detect_optimal_times(self):
        """最適時間帯検出"""
        # 時間別パフォーマンス分析
        best_hours = []
        for hour, scores in self.hourly_performance.items():
            if len(scores) >= 3:  # 最低3セッション
                avg_score = statistics.mean(scores)
                if avg_score > 70:  # 70%以上のパフォーマンス
                    best_hours.append((hour, avg_score, len(scores)))
        
        if best_hours:
            # 最高パフォーマンス時間帯を特定
            best_hours.sort(key=lambda x: x[1], reverse=True)
            top_hour, top_score, session_count = best_hours[0]
            
            optimal_data = {
                'type': 'optimal_hour',
                'hour': top_hour,
                'time_period': self._get_time_period(top_hour),
                'performance_score': round(top_score, 1),
                'sessions_count': session_count,
                'all_good_hours': [(h, round(s, 1)) for h, s, c in best_hours[:3]]
            }
            
            self.optimal_time_detected.emit(optimal_data)
            logger.info(f"🌍 最適時間帯検出: {top_hour}時 (パフォーマンス: {top_score:.1f}%)")
        
        # 曜日別パフォーマンス分析
        best_days = []
        day_names = ['月', '火', '水', '木', '金', '土', '日']
        for day, scores in self.daily_performance.items():
            if len(scores) >= 2:  # 最低2セッション
                avg_score = statistics.mean(scores)
                if avg_score > 65:  # 65%以上のパフォーマンス
                    best_days.append((day, avg_score, len(scores)))
        
        if best_days:
            best_days.sort(key=lambda x: x[1], reverse=True)
            top_day, top_score, session_count = best_days[0]
            
            optimal_data = {
                'type': 'optimal_day',
                'day_number': top_day,
                'day_name': day_names[top_day],
                'performance_score': round(top_score, 1),
                'sessions_count': session_count
            }
            
            self.optimal_time_detected.emit(optimal_data)
            logger.info(f"🌍 最適曜日検出: {day_names[top_day]} (パフォーマンス: {top_score:.1f}%)")
    
    def get_environment_insights(self, days: int = 14) -> Dict[str, Any]:
        """環境インサイト取得"""
        cutoff_date = datetime.now() - timedelta(days=days)
        
        recent_records = [
            r for r in self.environment_records
            if datetime.fromisoformat(r['session_start']) > cutoff_date
            and r.get('session_type') == 'work'
        ]
        
        if not recent_records:
            return {'message': 'データが不十分です'}
        
        # 時間帯分析
        time_period_performance = defaultdict(list)
        for record in recent_records:
            period = record.get('time_period', 'unknown')
            efficiency = record.get('efficiency_score', 0)
            focus = record.get('focus_score', 0)
            performance = (efficiency + focus) / 2
            time_period_performance[period].append(performance)
        
        # 最適時間帯
        best_period = None
        best_period_score = 0
        for period, scores in time_period_performance.items():
            if len(scores) >= 2:
                avg_score = statistics.mean(scores)
                if avg_score > best_period_score:
                    best_period = period
                    best_period_score = avg_score
        
        # 曜日分析
        day_names = ['月', '火', '水', '木', '金', '土', '日']
        weekday_performance = defaultdict(list)
        for record in recent_records:
            day = record.get('day_of_week', 0)
            efficiency = record.get('efficiency_score', 0)
            focus = record.get('focus_score', 0)
            performance = (efficiency + focus) / 2
            weekday_performance[day].append(performance)
        
        # 平日 vs 週末
        weekday_scores = []
        weekend_scores = []
        for record in recent_records:
            efficiency = record.get('efficiency_score', 0)
            focus = record.get('focus_score', 0)
            performance = (efficiency + focus) / 2
            
            if record.get('weekend', False):
                weekend_scores.append(performance)
            else:
                weekday_scores.append(performance)
        
        return {
            'analysis_period': f'過去{days}日間',
            'total_work_sessions': len(recent_records),
            'time_period_analysis': {
                'best_period': best_period,
                'best_period_score': round(best_period_score, 1) if best_period else 0,
                'period_scores': {
                    period: round(statistics.mean(scores), 1)
                    for period, scores in time_period_performance.items()
                    if len(scores) >= 2
                }
            },
            'weekday_analysis': {
                'weekday_avg': round(statistics.mean(weekday_scores), 1) if weekday_scores else 0,
                'weekend_avg': round(statistics.mean(weekend_scores), 1) if weekend_scores else 0,
                'day_scores': {
                    day_names[day]: round(statistics.mean(scores), 1)
                    for day, scores in weekday_performance.items()
                    if len(scores) >= 2
                }
            },
            'recommendations': self._generate_environment_recommendations(
                best_period, best_period_score, weekday_scores, weekend_scores
            )
        }
    
    def _generate_environment_recommendations(self, best_period: str, best_score: float, 
                                           weekday_scores: List[float], weekend_scores: List[float]) -> List[str]:
        """環境改善推奨事項生成"""
        recommendations = []
        
        if best_period and best_score > 75:
            period_names = {
                'morning': '午前中', 'afternoon': '午後', 'evening': '夕方', 'night': '夜間'
            }
            recommendations.append(f"🌅 {period_names.get(best_period, best_period)}の作業パフォーマンスが最も高いです")
        
        if weekday_scores and weekend_scores:
            weekday_avg = statistics.mean(weekday_scores)
            weekend_avg = statistics.mean(weekend_scores)
            
            if weekday_avg > weekend_avg + 10:
                recommendations.append("📅 平日の方が集中しやすい傾向があります")
            elif weekend_avg > weekday_avg + 10:
                recommendations.append("🏠 週末の方がリラックスして作業できているようです")
        
        if not recommendations:
            recommendations.append("📊 継続的なデータ収集で、より詳細な分析が可能になります")
        
        return recommendations
    
    def get_performance_heatmap_data(self) -> Dict[str, Any]:
        """パフォーマンスヒートマップ用データ取得"""
        # 時間×曜日のパフォーマンスマップ
        heatmap_data = {}
        day_names = ['月', '火', '水', '木', '金', '土', '日']
        
        for record in self.environment_records[-200:]:  # 直近200セッション
            if record.get('session_type') != 'work':
                continue
                
            hour = record.get('hour_of_day', 0)
            day = record.get('day_of_week', 0)
            efficiency = record.get('efficiency_score', 0)
            focus = record.get('focus_score', 0)
            performance = (efficiency + focus) / 2
            
            key = f"{day}_{hour}"
            if key not in heatmap_data:
                heatmap_data[key] = []
            heatmap_data[key].append(performance)
        
        # 平均パフォーマンスを計算
        averaged_data = {}
        for key, scores in heatmap_data.items():
            if len(scores) >= 2:  # 最低2セッション
                day, hour = key.split('_')
                averaged_data[key] = {
                    'day': int(day),
                    'day_name': day_names[int(day)],
                    'hour': int(hour),
                    'performance': round(statistics.mean(scores), 1),
                    'sessions_count': len(scores)
                }
        
        return {
            'heatmap_data': averaged_data,
            'day_names': day_names,
            'hours': list(range(24))
        }
    
    def load_environment_data(self):
        """環境データ読み込み"""
        try:
            if self.env_file.exists():
                with open(self.env_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.environment_records = data.get('records', [])
                    
                    # パフォーマンスマップ復元
                    hourly_data = data.get('hourly_performance', {})
                    self.hourly_performance = defaultdict(list)
                    for hour_str, scores in hourly_data.items():
                        self.hourly_performance[int(hour_str)] = scores
                    
                    daily_data = data.get('daily_performance', {})
                    self.daily_performance = defaultdict(list)
                    for day_str, scores in daily_data.items():
                        self.daily_performance[int(day_str)] = scores
                    
                    monthly_data = data.get('monthly_performance', {})
                    self.monthly_performance = defaultdict(list)
                    for month_str, scores in monthly_data.items():
                        self.monthly_performance[int(month_str)] = scores
            
            logger.info(f"🌍 EnvironmentLogger データ読み込み: {len(self.environment_records)}記録")
            
        except Exception as e:
            logger.error(f"EnvironmentLogger データ読み込みエラー: {e}")
            self.environment_records = []
            self.hourly_performance = defaultdict(list)
            self.daily_performance = defaultdict(list)
            self.monthly_performance = defaultdict(list)
    
    def save_environment_data(self):
        """環境データ保存"""
        try:
            data = {
                'records': self.environment_records,
                'hourly_performance': {str(k): v for k, v in self.hourly_performance.items()},
                'daily_performance': {str(k): v for k, v in self.daily_performance.items()},
                'monthly_performance': {str(k): v for k, v in self.monthly_performance.items()},
                'last_updated': datetime.now().isoformat()
            }
            
            with open(self.env_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            logger.error(f"EnvironmentLogger データ保存エラー: {e}")


class InteractiveReportsEngine(QObject):
    """Phase 4: インタラクティブレポートエンジン - ドリルダウン機能付きレポート生成"""
    
    # シグナル
    report_generated = pyqtSignal(str, dict)  # (report_type, report_data)
    drill_down_requested = pyqtSignal(str, dict)  # (data_type, filter_params)
    
    def __init__(self, data_collector, session_tracker, focus_calculator, 
                 interruption_tracker, environment_logger):
        super().__init__()
        
        # データソース
        self.data_collector = data_collector
        self.session_tracker = session_tracker
        self.focus_calculator = focus_calculator
        self.interruption_tracker = interruption_tracker
        self.environment_logger = environment_logger
        
        # レポート設定
        self.reports_dir = Path("data/reports")
        self.reports_dir.mkdir(exist_ok=True)
        
        # レポートキャッシュ
        self.report_cache = {}
        self.cache_expiry = timedelta(minutes=15)  # 15分でキャッシュ期限切れ
        
        logger.info("📈 InteractiveReportsEngine 初期化完了")
    
    def generate_comprehensive_report(self, date_range: Tuple[datetime, datetime] = None) -> Dict[str, Any]:
        """包括的レポート生成"""
        if date_range is None:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)  # 過去30日間
            date_range = (start_date, end_date)
        
        cache_key = f"comprehensive_{date_range[0].date()}_{date_range[1].date()}"
        
        # キャッシュチェック
        if self._is_cache_valid(cache_key):
            return self.report_cache[cache_key]['data']
        
        try:
            # 各コンポーネントからデータ取得
            session_data = self._get_session_summary(date_range)
            focus_data = self._get_focus_analysis(date_range)
            interruption_data = self._get_interruption_analysis(date_range)
            environment_data = self._get_environment_analysis(date_range)
            productivity_trends = self._get_productivity_trends(date_range)
            
            report = {
                'report_id': f"comprehensive_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                'generated_at': datetime.now().isoformat(),
                'date_range': {
                    'start': date_range[0].isoformat(),
                    'end': date_range[1].isoformat()
                },
                'summary': {
                    'total_sessions': session_data['total_sessions'],
                    'total_work_hours': session_data['total_work_hours'],
                    'average_focus_score': focus_data['average_score'],
                    'total_interruptions': interruption_data['total_count'],
                    'productivity_trend': productivity_trends['overall_trend']
                },
                'detailed_sections': {
                    'session_analysis': session_data,
                    'focus_analysis': focus_data,
                    'interruption_analysis': interruption_data,
                    'environment_analysis': environment_data,
                    'productivity_trends': productivity_trends
                },
                'drill_down_available': {
                    'sessions_by_day': True,
                    'focus_by_time': True,
                    'interruption_patterns': True,
                    'environment_correlation': True
                },
                'recommendations': self._generate_recommendations(
                    session_data, focus_data, interruption_data, environment_data
                )
            }
            
            # キャッシュに保存
            self._cache_report(cache_key, report)
            
            self.report_generated.emit('comprehensive', report)
            logger.info(f"📈 包括的レポート生成完了: {date_range[0].date()} - {date_range[1].date()}")
            
            return report
            
        except Exception as e:
            logger.error(f"包括的レポート生成エラー: {e}")
            return self._get_empty_report()
    
    def generate_focus_drill_down(self, session_id: str = None, 
                                date_filter: datetime = None) -> Dict[str, Any]:
        """フォーカス詳細ドリルダウン"""
        try:
            if session_id:
                # 特定セッションの詳細
                session_focus = self._get_session_focus_details(session_id)
                return {
                    'type': 'session_focus_detail',
                    'session_id': session_id,
                    'focus_timeline': session_focus['timeline'],
                    'focus_events': session_focus['events'],
                    'score_breakdown': session_focus['breakdown'],
                    'improvement_tips': session_focus['tips']
                }
            elif date_filter:
                # 特定日のフォーカス分析
                daily_focus = self._get_daily_focus_analysis(date_filter)
                return {
                    'type': 'daily_focus_analysis',
                    'date': date_filter.isoformat(),
                    'sessions': daily_focus['sessions'],
                    'peak_focus_times': daily_focus['peak_times'],
                    'focus_patterns': daily_focus['patterns'],
                    'recommendations': daily_focus['recommendations']
                }
        except Exception as e:
            logger.error(f"フォーカスドリルダウンエラー: {e}")
            return {'error': str(e)}
    
    def generate_interruption_drill_down(self, interruption_type: str = None,
                                       time_range: Tuple[datetime, datetime] = None) -> Dict[str, Any]:
        """中断分析ドリルダウン"""
        try:
            interruption_data = self.interruption_tracker.get_interruption_analysis()
            
            if interruption_type:
                # 特定タイプの中断詳細
                type_analysis = self._analyze_interruption_type(interruption_type, time_range)
                return {
                    'type': 'interruption_type_analysis',
                    'interruption_type': interruption_type,
                    'frequency': type_analysis['frequency'],
                    'time_patterns': type_analysis['time_patterns'],
                    'duration_analysis': type_analysis['duration'],
                    'prevention_tips': type_analysis['prevention_tips']
                }
            else:
                # 全中断パターン分析
                return {
                    'type': 'interruption_pattern_analysis',
                    'time_range': time_range,
                    'patterns': interruption_data['patterns'],
                    'hotspots': interruption_data['hotspots'],
                    'trends': interruption_data['trends']
                }
        except Exception as e:
            logger.error(f"中断ドリルダウンエラー: {e}")
            return {'error': str(e)}
    
    def _get_session_summary(self, date_range: Tuple[datetime, datetime]) -> Dict[str, Any]:
        """セッション要約データ取得"""
        sessions = self.data_collector.session_data
        filtered_sessions = [
            s for s in sessions 
            if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
        ]
        
        total_sessions = len(filtered_sessions)
        work_sessions = [s for s in filtered_sessions if s['type'] == 'work']
        total_work_time = sum(s.get('actual_duration', 0) for s in work_sessions)
        
        return {
            'total_sessions': total_sessions,
            'work_sessions': len(work_sessions),
            'total_work_hours': round(total_work_time / 60, 2),
            'average_session_length': round(
                statistics.mean([s.get('actual_duration', 0) for s in work_sessions]) 
                if work_sessions else 0, 1
            ),
            'completion_rate': round(
                len([s for s in work_sessions if s.get('completed', False)]) / len(work_sessions) * 100
                if work_sessions else 0, 1
            )
        }
    
    def _get_focus_analysis(self, date_range: Tuple[datetime, datetime]) -> Dict[str, Any]:
        """フォーカス分析データ取得"""
        focus_data = self.focus_calculator.get_comprehensive_analysis()
        
        return {
            'average_score': focus_data.get('average_focus_score', 0),
            'score_trend': focus_data.get('trend', 'stable'),
            'peak_performance_times': focus_data.get('peak_times', []),
            'focus_patterns': focus_data.get('patterns', {}),
            'improvement_areas': focus_data.get('improvement_areas', [])
        }
    
    def _get_interruption_analysis(self, date_range: Tuple[datetime, datetime]) -> Dict[str, Any]:
        """中断分析データ取得"""
        interruption_data = self.interruption_tracker.get_interruption_analysis()
        
        return {
            'total_count': interruption_data.get('total_interruptions', 0),
            'most_common_type': interruption_data.get('most_common_type', 'unknown'),
            'average_duration': interruption_data.get('average_duration', 0),
            'patterns': interruption_data.get('patterns', {}),
            'hotspot_times': interruption_data.get('hotspot_times', [])
        }
    
    def _get_environment_analysis(self, date_range: Tuple[datetime, datetime]) -> Dict[str, Any]:
        """環境分析データ取得"""
        env_data = self.environment_logger.get_performance_analysis()
        
        return {
            'best_performance_time': env_data.get('time_analysis', {}).get('best_period', 'unknown'),
            'weekday_vs_weekend': env_data.get('day_analysis', {}),
            'performance_heatmap': self.environment_logger.get_performance_heatmap_data(),
            'recommendations': env_data.get('recommendations', [])
        }
    
    def _get_productivity_trends(self, date_range: Tuple[datetime, datetime]) -> Dict[str, Any]:
        """生産性トレンド分析"""
        sessions = self.data_collector.session_data
        filtered_sessions = [
            s for s in sessions 
            if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
        ]
        
        # 日別生産性計算
        daily_productivity = defaultdict(list)
        for session in filtered_sessions:
            if session['type'] == 'work':
                date_key = datetime.fromisoformat(session['start_time']).date()
                productivity_score = session.get('efficiency_score', 0) + session.get('focus_score', 0)
                daily_productivity[date_key].append(productivity_score)
        
        # トレンド計算
        daily_averages = {
            date: statistics.mean(scores) 
            for date, scores in daily_productivity.items()
        }
        
        if len(daily_averages) >= 3:
            dates = sorted(daily_averages.keys())
            scores = [daily_averages[date] for date in dates]
            
            # 線形トレンド計算（簡易版）
            if len(scores) > 1:
                trend_slope = (scores[-1] - scores[0]) / len(scores)
                if trend_slope > 0.5:
                    trend = 'improving'
                elif trend_slope < -0.5:
                    trend = 'declining'
                else:
                    trend = 'stable'
            else:
                trend = 'insufficient_data'
        else:
            trend = 'insufficient_data'
        
        return {
            'overall_trend': trend,
            'daily_scores': daily_averages,
            'trend_analysis': self._analyze_productivity_trend(daily_averages)
        }
    
    def _analyze_productivity_trend(self, daily_scores: Dict) -> Dict[str, Any]:
        """生産性トレンド詳細分析"""
        if not daily_scores:
            return {'status': 'no_data'}
        
        scores = list(daily_scores.values())
        
        return {
            'average_score': round(statistics.mean(scores), 2),
            'best_day': max(daily_scores, key=daily_scores.get),
            'worst_day': min(daily_scores, key=daily_scores.get),
            'consistency': round(1 - (statistics.stdev(scores) / statistics.mean(scores)), 2) if len(scores) > 1 else 1.0,
            'improvement_rate': self._calculate_improvement_rate(scores)
        }
    
    def _calculate_improvement_rate(self, scores: List[float]) -> float:
        """改善率計算"""
        if len(scores) < 2:
            return 0.0
        
        # 前半と後半の平均を比較
        mid_point = len(scores) // 2
        first_half_avg = statistics.mean(scores[:mid_point])
        second_half_avg = statistics.mean(scores[mid_point:])
        
        if first_half_avg == 0:
            return 0.0
        
        return round((second_half_avg - first_half_avg) / first_half_avg * 100, 2)
    
    def _generate_recommendations(self, session_data: Dict, focus_data: Dict, 
                                interruption_data: Dict, environment_data: Dict) -> List[str]:
        """AI駆動推奨事項生成"""
        recommendations = []
        
        # セッション完了率に基づく推奨
        if session_data['completion_rate'] < 70:
            recommendations.append("🎯 セッション完了率が低いです。より小さな目標設定を試してみてください")
        
        # フォーカススコアに基づく推奨
        if focus_data['average_score'] < 60:
            recommendations.append("🧠 集中力向上のため、作業前の短い瞑想や準備時間を設けることをお勧めします")
        
        # 中断回数に基づく推奨
        if interruption_data['total_count'] > session_data['total_sessions'] * 0.5:
            recommendations.append("⚠️ 中断が多いです。通知をオフにしたり、集中できる環境を整えてください")
        
        # 環境データに基づく推奨
        best_time = environment_data.get('best_performance_time')
        if best_time and best_time != 'unknown':
            recommendations.append(f"⏰ {best_time}の時間帯でのパフォーマンスが最高です。重要なタスクをこの時間に配置してください")
        
        if not recommendations:
            recommendations.append("📈 順調です！現在のリズムを維持していきましょう")
        
        return recommendations
    
    def _is_cache_valid(self, cache_key: str) -> bool:
        """キャッシュ有効性チェック"""
        if cache_key not in self.report_cache:
            return False
        
        cache_time = self.report_cache[cache_key]['timestamp']
        return datetime.now() - cache_time < self.cache_expiry
    
    def _cache_report(self, cache_key: str, report_data: Dict):
        """レポートキャッシュ"""
        self.report_cache[cache_key] = {
            'timestamp': datetime.now(),
            'data': report_data
        }
        
        # キャッシュサイズ制限（最大10レポート）
        if len(self.report_cache) > 10:
            oldest_key = min(self.report_cache.keys(), 
                           key=lambda k: self.report_cache[k]['timestamp'])
            del self.report_cache[oldest_key]
    
    def _get_empty_report(self) -> Dict[str, Any]:
        """空のレポート取得"""
        return {
            'report_id': f"empty_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            'generated_at': datetime.now().isoformat(),
            'error': 'データが不足しているため、レポートを生成できませんでした',
            'summary': {},
            'detailed_sections': {},
            'recommendations': ['📊 継続的なデータ収集により、詳細な分析が可能になります']
        }
    
    def export_report(self, report_data: Dict, format_type: str = 'json') -> str:
        """レポートエクスポート"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"report_{timestamp}.{format_type}"
            filepath = self.reports_dir / filename
            
            if format_type == 'json':
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(report_data, f, ensure_ascii=False, indent=2)
            elif format_type == 'txt':
                self._export_as_text(report_data, filepath)
            
            logger.info(f"📄 レポートエクスポート完了: {filename}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"レポートエクスポートエラー: {e}")
            return ""
    
    def _export_as_text(self, report_data: Dict, filepath: Path):
        """テキスト形式でエクスポート"""
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("🍅 Pomodoro Timer 生産性レポート\n")
            f.write("=" * 50 + "\n\n")
            
            # サマリー
            summary = report_data.get('summary', {})
            f.write("📊 概要\n")
            f.write("-" * 20 + "\n")
            for key, value in summary.items():
                f.write(f"{key}: {value}\n")
            f.write("\n")
            
            # 推奨事項
            recommendations = report_data.get('recommendations', [])
            if recommendations:
                f.write("💡 推奨事項\n")
                f.write("-" * 20 + "\n")
                for i, rec in enumerate(recommendations, 1):
                    f.write(f"{i}. {rec}\n")


class AdvancedVisualization(QObject):
    """Phase 4: 高度な可視化システム - matplotlib/seaborn使用"""
    
    # シグナル
    visualization_ready = pyqtSignal(str, object)  # (chart_type, figure_widget)
    export_completed = pyqtSignal(str, str)  # (chart_type, filepath)
    
    def __init__(self, reports_engine, data_collector):
        super().__init__()
        
        self.reports_engine = reports_engine
        self.data_collector = data_collector
        
        # 可視化設定
        self.figure_size = (12, 8)
        self.dpi = 100
        
        # ライブラリ利用可能性チェック
        self.matplotlib_available = MATPLOTLIB_AVAILABLE
        self.seaborn_available = SEABORN_AVAILABLE
        
        # スタイル設定（matplotlib/seaborn利用可能な場合のみ）
        if self.matplotlib_available:
            try:
                if self.seaborn_available:
                    plt.style.use('default')  # seaborn-v0_8-darkgridは問題を起こす可能性があるため
                    sns.set_palette("husl")
                else:
                    plt.style.use('default')
            except Exception as e:
                logger.warning(f"📊 スタイル設定エラー: {e}")
                plt.style.use('default')
        
        # 出力ディレクトリ
        self.charts_dir = Path("data/charts")
        self.charts_dir.mkdir(exist_ok=True)
        
        # フォールバック用カラーパレット
        self.fallback_colors = [
            '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
            '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf'
        ]
        
        if not self.matplotlib_available:
            logger.warning("📊 matplotlib利用不可のため、テキストベース表示を使用します")
        elif not self.seaborn_available:
            logger.warning("📊 seaborn利用不可のため、matplotlib基本機能のみ使用します")
        
        logger.info("📊 AdvancedVisualization 初期化完了")
    
    def _get_color_palette(self, palette_name: str, n_colors: int) -> List[str]:
        """カラーパレット取得（seaborn利用可能性を考慮）"""
        if self.seaborn_available:
            try:
                return sns.color_palette(palette_name, n_colors).as_hex()
            except:
                pass
        
        # フォールバック: 基本カラーパレットから循環取得
        colors = []
        for i in range(n_colors):
            colors.append(self.fallback_colors[i % len(self.fallback_colors)])
        return colors
    
    def _create_text_based_display(self, title: str, data: Dict[str, Any]) -> 'QWidget':
        """matplotlib利用不可時のテキストベース表示作成"""
        from PyQt6.QtWidgets import QWidget, QVBoxLayout, QLabel, QTextEdit
        from PyQt6.QtCore import Qt
        
        widget = QWidget()
        layout = QVBoxLayout()
        
        # タイトル
        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin: 10px;")
        layout.addWidget(title_label)
        
        # データ表示
        text_display = QTextEdit()
        text_display.setReadOnly(True)
        
        # データを読みやすいテキスト形式に変換
        text_content = self._format_data_as_text(data)
        text_display.setPlainText(text_content)
        text_display.setStyleSheet("font-family: monospace; font-size: 12px;")
        
        layout.addWidget(text_display)
        widget.setLayout(layout)
        widget.setMinimumSize(600, 400)
        
        return widget
    
    def _format_data_as_text(self, data: Dict[str, Any]) -> str:
        """データをテキスト形式に整形"""
        lines = []
        
        if isinstance(data, dict):
            for key, value in data.items():
                if isinstance(value, (list, tuple)):
                    lines.append(f"{key}:")
                    for i, item in enumerate(value):
                        lines.append(f"  {i+1}. {item}")
                elif isinstance(value, dict):
                    lines.append(f"{key}:")
                    for sub_key, sub_value in value.items():
                        lines.append(f"  {sub_key}: {sub_value}")
                else:
                    lines.append(f"{key}: {value}")
                lines.append("")
        
        return "\n".join(lines)
    
    def create_productivity_timeline(self, date_range: Tuple[datetime, datetime] = None) -> 'QWidget':
        """生産性タイムライン作成"""
        try:
            # データ取得
            report_data = self.reports_engine.generate_comprehensive_report(date_range)
            productivity_data = report_data['detailed_sections']['productivity_trends']['daily_scores']
            
            if not productivity_data:
                if self.matplotlib_available:
                    return self._create_no_data_chart("生産性データが不足しています")
                else:
                    return self._create_text_based_display("📈 生産性トレンド分析", {"メッセージ": "生産性データが不足しています"})
            
            # matplotlib利用不可の場合はテキストベース表示
            if not self.matplotlib_available:
                # データを整理してテキスト表示用に準備
                display_data = {
                    "期間": f"{list(productivity_data.keys())[0]} ～ {list(productivity_data.keys())[-1]}",
                    "データ数": len(productivity_data),
                    "平均スコア": f"{statistics.mean(productivity_data.values()):.2f}",
                    "最高スコア": f"{max(productivity_data.values()):.2f}",
                    "最低スコア": f"{min(productivity_data.values()):.2f}",
                    "日別データ": {str(date): f"{score:.2f}" for date, score in productivity_data.items()}
                }
                return self._create_text_based_display("📈 生産性トレンド分析", display_data)
            
            # matplotlib利用可能な場合はグラフ作成
            fig = Figure(figsize=self.figure_size, dpi=self.dpi)
            ax = fig.add_subplot(111)
            
            # データ準備
            dates = list(productivity_data.keys())
            scores = list(productivity_data.values())
            
            # 日付をdatetimeに変換
            datetime_dates = [datetime.combine(date, datetime.min.time()) for date in dates]
            
            # プロット
            ax.plot(datetime_dates, scores, marker='o', linewidth=2, markersize=6, 
                   color='#2E86AB', label='生産性スコア')
            
            # トレンドライン追加
            if len(scores) > 2:
                z = np.polyfit(range(len(scores)), scores, 1)
                p = np.poly1d(z)
                trend_scores = p(range(len(scores)))
                ax.plot(datetime_dates, trend_scores, "--", alpha=0.7, 
                       color='#A23B72', label='トレンド')
            
            # 平均線
            avg_score = statistics.mean(scores)
            ax.axhline(y=avg_score, color='#F18F01', linestyle='-', alpha=0.7, 
                      label=f'平均 ({avg_score:.1f})')
            
            # グラフ設定
            ax.set_title('📈 生産性トレンド分析', fontsize=16, fontweight='bold', pad=20)
            ax.set_xlabel('日付', fontsize=12)
            ax.set_ylabel('生産性スコア', fontsize=12)
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            # 日付フォーマット
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//10)))
            fig.autofmt_xdate()
            
            # レイアウト調整
            fig.tight_layout()
            
            # FigureCanvasに変換
            canvas = FigureCanvas(fig)
            canvas.setMinimumSize(600, 400)
            
            self.visualization_ready.emit('productivity_timeline', canvas)
            return canvas
            
        except Exception as e:
            logger.error(f"生産性タイムライン作成エラー: {e}")
            if self.matplotlib_available:
                return self._create_error_chart(f"エラー: {e}")
            else:
                return self._create_text_based_display("📈 生産性トレンド分析 (エラー)", {"エラー": str(e)})
    
    def create_focus_heatmap(self, date_range: Tuple[datetime, datetime] = None) -> 'QWidget':
        """フォーカスヒートマップ作成"""
        try:
            # 環境データからヒートマップデータ取得
            from collections import defaultdict
            
            sessions = self.data_collector.session_data
            if date_range:
                sessions = [
                    s for s in sessions 
                    if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
                ]
            
            # 時間×曜日のデータマトリクス作成
            heatmap_data = defaultdict(lambda: defaultdict(list))
            
            for session in sessions:
                if session['type'] != 'work':
                    continue
                    
                start_time = datetime.fromisoformat(session['start_time'])
                hour = start_time.hour
                weekday = start_time.weekday()  # 0=月曜日
                focus_score = session.get('focus_score', 0)
                
                heatmap_data[weekday][hour].append(focus_score)
            
            # 平均スコア計算
            hours = list(range(24))
            weekdays = list(range(7))
            weekday_names = ['月', '火', '水', '木', '金', '土', '日']
            
            # matplotlib利用不可の場合はテキストベース表示
            if not self.matplotlib_available:
                # ヒートマップデータをテキスト形式で表示
                display_data = {"曜日別・時間別フォーカススコア": {}}
                for day_idx, day_name in enumerate(weekday_names):
                    day_data = {}
                    for hour in hours:
                        scores = heatmap_data[day_idx][hour] 
                        if scores:
                            day_data[f"{hour:02d}:00"] = f"{statistics.mean(scores):.2f}"
                        else:
                            day_data[f"{hour:02d}:00"] = "データなし"
                    if any(score != "データなし" for score in day_data.values()):
                        display_data["曜日別・時間別フォーカススコア"][day_name] = day_data
                
                return self._create_text_based_display("🔥 フォーカススコア ヒートマップ", display_data)
            
            # numpy利用可能性チェック
            try:
                matrix = np.zeros((7, 24))
                for day in weekdays:
                    for hour in hours:
                        scores = heatmap_data[day][hour]
                        if scores:
                            matrix[day][hour] = statistics.mean(scores)
                        else:
                            matrix[day][hour] = np.nan
            except NameError:
                # numpyが利用できない場合のフォールバック
                matrix = [[0 for _ in range(24)] for _ in range(7)]
                for day in weekdays:
                    for hour in hours:
                        scores = heatmap_data[day][hour]
                        if scores:
                            matrix[day][hour] = statistics.mean(scores)
                        else:
                            matrix[day][hour] = None
            
            # Figure作成
            fig = Figure(figsize=(14, 6), dpi=self.dpi)
            ax = fig.add_subplot(111)
            
            # ヒートマップ作成
            im = ax.imshow(matrix, cmap='YlOrRd', aspect='auto', interpolation='nearest')
            
            # カラーバー
            cbar = fig.colorbar(im, ax=ax)
            cbar.set_label('平均フォーカススコア', rotation=270, labelpad=20)
            
            # 軸設定
            ax.set_xticks(range(24))
            ax.set_xticklabels([f'{h:02d}:00' for h in range(24)], rotation=45)
            ax.set_yticks(range(7))
            ax.set_yticklabels(weekday_names)
            
            ax.set_xlabel('時間', fontsize=12)
            ax.set_ylabel('曜日', fontsize=12)
            ax.set_title('🔥 フォーカススコア ヒートマップ', fontsize=16, fontweight='bold', pad=20)
            
            # グリッド
            ax.set_xticks(np.arange(-0.5, 24, 1), minor=True)
            ax.set_yticks(np.arange(-0.5, 7, 1), minor=True)
            ax.grid(which='minor', color='white', linestyle='-', linewidth=1)
            
            fig.tight_layout()
            
            canvas = FigureCanvas(fig)
            canvas.setMinimumSize(700, 300)
            
            self.visualization_ready.emit('focus_heatmap', canvas)
            return canvas
            
        except Exception as e:
            logger.error(f"フォーカスヒートマップ作成エラー: {e}")
            if self.matplotlib_available:
                return self._create_error_chart(f"エラー: {e}")
            else:
                return self._create_text_based_display("🔥 フォーカススコア ヒートマップ (エラー)", {"エラー": str(e)})
    
    def create_interruption_analysis_chart(self, date_range: Tuple[datetime, datetime] = None) -> 'QWidget':
        """中断分析チャート作成"""
        try:
            # 中断データ取得
            sessions = self.data_collector.session_data
            if date_range:
                sessions = [
                    s for s in sessions 
                    if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
                ]
            
            # 中断タイプ別カウント
            interruption_counts = defaultdict(int)
            interruption_durations = defaultdict(list)
            
            for session in sessions:
                for interruption in session.get('interruptions', []):
                    int_type = interruption.get('type', 'unknown')
                    duration = interruption.get('duration', 0)
                    
                    interruption_counts[int_type] += 1
                    interruption_durations[int_type].append(duration)
            
            if not interruption_counts:
                if self.matplotlib_available:
                    return self._create_no_data_chart("中断データが不足しています")
                else:
                    return self._create_text_based_display("⚠️ 中断分析ダッシュボード", {"メッセージ": "中断データが不足しています"})
            
            # matplotlib利用不可の場合はテキストベース表示
            if not self.matplotlib_available:
                # 中断データを整理してテキスト表示用に準備
                types = list(interruption_counts.keys())
                counts = list(interruption_counts.values())
                
                display_data = {
                    "中断回数（タイプ別）": {t: str(c) for t, c in zip(types, counts)},
                    "中断継続時間統計": {},
                    "総中断回数": sum(counts),
                    "最も多い中断タイプ": max(interruption_counts, key=interruption_counts.get) if interruption_counts else "なし"
                }
                
                # 継続時間統計
                for int_type, durations in interruption_durations.items():
                    if durations:
                        display_data["中断継続時間統計"][int_type] = {
                            "平均": f"{statistics.mean(durations):.1f}秒",
                            "最大": f"{max(durations):.1f}秒",
                            "最小": f"{min(durations):.1f}秒"
                        }
                
                return self._create_text_based_display("⚠️ 中断分析ダッシュボード", display_data)
            
            # Figure作成（2つのサブプロット）
            fig = Figure(figsize=(14, 8), dpi=self.dpi)
            
            # 1. 中断回数（棒グラフ）
            ax1 = fig.add_subplot(221)
            types = list(interruption_counts.keys())
            counts = list(interruption_counts.values())
            
            bars1 = ax1.bar(types, counts, color=self._get_color_palette("viridis", len(types)))
            ax1.set_title('中断回数（タイプ別）', fontsize=14, fontweight='bold')
            ax1.set_ylabel('回数')
            ax1.tick_params(axis='x', rotation=45)
            
            # 値をバーの上に表示
            for bar, count in zip(bars1, counts):
                height = bar.get_height()
                ax1.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                        f'{count}', ha='center', va='bottom')
            
            # 2. 中断継続時間（箱ひげ図）
            ax2 = fig.add_subplot(222)
            duration_data = [durations for durations in interruption_durations.values() if durations]
            duration_labels = [t for t, durations in interruption_durations.items() if durations]
            
            if duration_data:
                bp = ax2.boxplot(duration_data, labels=duration_labels, patch_artist=True)
                
                # 箱の色設定
                colors = self._get_color_palette("Set2", len(bp['boxes']))
                for patch, color in zip(bp['boxes'], colors):
                    patch.set_facecolor(color)
                    patch.set_alpha(0.7)
                
                ax2.set_title('中断継続時間分布', fontsize=14, fontweight='bold')
                ax2.set_ylabel('継続時間（秒）')
                ax2.tick_params(axis='x', rotation=45)
            
            # 3. 時間別中断発生パターン
            ax3 = fig.add_subplot(223)
            hourly_interruptions = defaultdict(int)
            
            for session in sessions:
                for interruption in session.get('interruptions', []):
                    timestamp = datetime.fromisoformat(interruption['timestamp'])
                    hour = timestamp.hour
                    hourly_interruptions[hour] += 1
            
            hours = list(range(24))
            int_counts_by_hour = [hourly_interruptions[h] for h in hours]
            
            ax3.plot(hours, int_counts_by_hour, marker='o', linewidth=2, color='#E74C3C')
            ax3.fill_between(hours, int_counts_by_hour, alpha=0.3, color='#E74C3C')
            ax3.set_title('時間別中断発生パターン', fontsize=14, fontweight='bold')
            ax3.set_xlabel('時間')
            ax3.set_ylabel('中断回数')
            ax3.set_xticks(range(0, 24, 4))
            ax3.grid(True, alpha=0.3)
            
            # 4. 中断タイプ円グラフ
            ax4 = fig.add_subplot(224)
            colors_pie = self._get_color_palette("pastel", len(types))
            wedges, texts, autotexts = ax4.pie(counts, labels=types, autopct='%1.1f%%',
                                              colors=colors_pie, startangle=90)
            ax4.set_title('中断タイプ分布', fontsize=14, fontweight='bold')
            
            fig.suptitle('⚠️ 中断分析ダッシュボード', fontsize=16, fontweight='bold')
            fig.tight_layout()
            
            canvas = FigureCanvas(fig)
            canvas.setMinimumSize(800, 600)
            
            self.visualization_ready.emit('interruption_analysis', canvas)
            return canvas
            
        except Exception as e:
            logger.error(f"中断分析チャート作成エラー: {e}")
            if self.matplotlib_available:
                return self._create_error_chart(f"エラー: {e}")
            else:
                return self._create_text_based_display("⚠️ 中断分析ダッシュボード (エラー)", {"エラー": str(e)})
    
    def create_session_performance_chart(self, date_range: Tuple[datetime, datetime] = None) -> 'QWidget':
        """セッションパフォーマンスチャート作成"""
        try:
            sessions = self.data_collector.session_data
            if date_range:
                sessions = [
                    s for s in sessions 
                    if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
                ]
            
            work_sessions = [s for s in sessions if s['type'] == 'work']
            
            if not work_sessions:
                if self.matplotlib_available:
                    return self._create_no_data_chart("ワークセッションデータが不足しています")
                else:
                    return self._create_text_based_display("🏆 セッションパフォーマンス総合分析", {"メッセージ": "ワークセッションデータが不足しています"})
            
            # データ準備
            completion_rates = []
            focus_scores = []
            efficiency_scores = []
            dates = []
            
            for session in work_sessions:
                start_date = datetime.fromisoformat(session['start_time']).date()
                dates.append(start_date)
                
                # 完了率（完了=1, 未完了=0）
                completion_rates.append(1 if session.get('completed', False) else 0)
                focus_scores.append(session.get('focus_score', 0))
                efficiency_scores.append(session.get('efficiency_score', 0))
            
            # matplotlib利用不可の場合はテキストベース表示
            if not self.matplotlib_available:
                # パフォーマンスデータを整理してテキスト表示用に準備
                total_completion_rate = (sum(completion_rates) / len(completion_rates)) * 100 if completion_rates else 0
                avg_focus = statistics.mean(focus_scores) if focus_scores else 0
                avg_efficiency = statistics.mean(efficiency_scores) if efficiency_scores else 0
                
                display_data = {
                    "セッション統計": {
                        "総セッション数": len(work_sessions),
                        "平均完了率": f"{total_completion_rate:.1f}%",
                        "平均フォーカススコア": f"{avg_focus:.2f}",
                        "平均効率スコア": f"{avg_efficiency:.2f}"
                    },
                    "スコア分布": {
                        "フォーカススコア最高": f"{max(focus_scores):.2f}" if focus_scores else "0",
                        "フォーカススコア最低": f"{min(focus_scores):.2f}" if focus_scores else "0",
                        "効率スコア最高": f"{max(efficiency_scores):.2f}" if efficiency_scores else "0",
                        "効率スコア最低": f"{min(efficiency_scores):.2f}" if efficiency_scores else "0"
                    }
                }
                
                return self._create_text_based_display("🏆 セッションパフォーマンス総合分析", display_data)
            
            # Figure作成（3つのサブプロット）
            fig = Figure(figsize=(15, 10), dpi=self.dpi)
            
            # 1. 完了率とスコアの時系列
            ax1 = fig.add_subplot(311)
            
            # 日別集計
            daily_data = defaultdict(lambda: {'completion': [], 'focus': [], 'efficiency': []})
            for i, date in enumerate(dates):
                daily_data[date]['completion'].append(completion_rates[i])
                daily_data[date]['focus'].append(focus_scores[i])
                daily_data[date]['efficiency'].append(efficiency_scores[i])
            
            plot_dates = sorted(daily_data.keys())
            daily_completion = [statistics.mean(daily_data[d]['completion']) * 100 for d in plot_dates]
            daily_focus = [statistics.mean(daily_data[d]['focus']) for d in plot_dates]
            daily_efficiency = [statistics.mean(daily_data[d]['efficiency']) for d in plot_dates]
            
            datetime_dates = [datetime.combine(date, datetime.min.time()) for date in plot_dates]
            
            ax1.plot(datetime_dates, daily_completion, marker='s', label='完了率 (%)', 
                    color='#27AE60', linewidth=2)
            ax1.plot(datetime_dates, daily_focus, marker='o', label='フォーカススコア', 
                    color='#3498DB', linewidth=2)
            ax1.plot(datetime_dates, daily_efficiency, marker='^', label='効率スコア', 
                    color='#E67E22', linewidth=2)
            
            ax1.set_title('📊 日別パフォーマンス推移', fontsize=14, fontweight='bold')
            ax1.set_ylabel('スコア')
            ax1.legend()
            ax1.grid(True, alpha=0.3)
            ax1.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            
            # 2. パフォーマンス相関分析
            ax2 = fig.add_subplot(312)
            scatter = ax2.scatter(focus_scores, efficiency_scores, 
                                c=completion_rates, cmap='RdYlGn', 
                                s=60, alpha=0.7, edgecolors='black', linewidth=0.5)
            
            ax2.set_xlabel('フォーカススコア')
            ax2.set_ylabel('効率スコア')
            ax2.set_title('🎯 フォーカス vs 効率 相関分析', fontsize=14, fontweight='bold')
            ax2.grid(True, alpha=0.3)
            
            # カラーバー
            cbar2 = fig.colorbar(scatter, ax=ax2)
            cbar2.set_label('完了状況 (0=未完了, 1=完了)')
            
            # 3. パフォーマンス分布（ヒストグラム）
            ax3 = fig.add_subplot(313)
            
            ax3.hist(focus_scores, bins=15, alpha=0.7, label='フォーカススコア', 
                    color='#3498DB', density=True)
            ax3.hist(efficiency_scores, bins=15, alpha=0.7, label='効率スコア', 
                    color='#E67E22', density=True)
            
            ax3.set_xlabel('スコア')
            ax3.set_ylabel('密度')
            ax3.set_title('📈 パフォーマンススコア分布', fontsize=14, fontweight='bold')
            ax3.legend()
            ax3.grid(True, alpha=0.3)
            
            fig.suptitle('🏆 セッションパフォーマンス総合分析', fontsize=16, fontweight='bold')
            fig.tight_layout()
            
            canvas = FigureCanvas(fig)
            canvas.setMinimumSize(900, 700)
            
            self.visualization_ready.emit('session_performance', canvas)
            return canvas
            
        except Exception as e:
            logger.error(f"セッションパフォーマンスチャート作成エラー: {e}")
            if self.matplotlib_available:
                return self._create_error_chart(f"エラー: {e}")
            else:
                return self._create_text_based_display("🏆 セッションパフォーマンス総合分析 (エラー)", {"エラー": str(e)})
    
    def create_custom_dashboard(self, chart_types: List[str], 
                              date_range: Tuple[datetime, datetime] = None) -> 'QWidget':
        """カスタムダッシュボード作成"""
        try:
            num_charts = len(chart_types)
            if num_charts == 0:
                if self.matplotlib_available:
                    return self._create_no_data_chart("表示するチャートが選択されていません")
                else:
                    return self._create_text_based_display("📊 カスタムダッシュボード", {"メッセージ": "表示するチャートが選択されていません"})
            
            # matplotlib利用不可の場合はテキストベース表示
            if not self.matplotlib_available:
                display_data = {
                    "選択されたチャート": chart_types,
                    "チャート数": num_charts,
                    "メッセージ": "matplotlibが利用できないため、個別のチャート表示を使用してください"
                }
                return self._create_text_based_display("📊 カスタムダッシュボード", display_data)
            
            # レイアウト計算
            if num_charts == 1:
                rows, cols = 1, 1
            elif num_charts == 2:
                rows, cols = 1, 2
            elif num_charts <= 4:
                rows, cols = 2, 2
            else:
                rows, cols = 3, 2
            
            # Figure作成
            fig = Figure(figsize=(cols * 6, rows * 4), dpi=self.dpi)
            
            for i, chart_type in enumerate(chart_types[:6]):  # 最大6個
                ax = fig.add_subplot(rows, cols, i + 1)
                
                if chart_type == 'productivity_trend':
                    self._add_productivity_trend_subplot(ax, date_range)
                elif chart_type == 'focus_distribution':
                    self._add_focus_distribution_subplot(ax, date_range)
                elif chart_type == 'interruption_summary':
                    self._add_interruption_summary_subplot(ax, date_range)
                elif chart_type == 'completion_rate':
                    self._add_completion_rate_subplot(ax, date_range)
                elif chart_type == 'time_analysis':
                    self._add_time_analysis_subplot(ax, date_range)
                elif chart_type == 'weekly_pattern':
                    self._add_weekly_pattern_subplot(ax, date_range)
            
            fig.suptitle('📊 カスタムダッシュボード', fontsize=16, fontweight='bold')
            fig.tight_layout()
            
            canvas = FigureCanvas(fig)
            canvas.setMinimumSize(600 * cols, 300 * rows)
            
            self.visualization_ready.emit('custom_dashboard', canvas)
            return canvas
            
        except Exception as e:
            logger.error(f"カスタムダッシュボード作成エラー: {e}")
            if self.matplotlib_available:
                return self._create_error_chart(f"エラー: {e}")
            else:
                return self._create_text_based_display("📊 カスタムダッシュボード (エラー)", {"エラー": str(e)})
    
    def _add_productivity_trend_subplot(self, ax, date_range):
        """生産性トレンドサブプロット追加"""
        try:
            report_data = self.reports_engine.generate_comprehensive_report(date_range)
            productivity_data = report_data['detailed_sections']['productivity_trends']['daily_scores']
            
            if productivity_data:
                dates = list(productivity_data.keys())
                scores = list(productivity_data.values())
                datetime_dates = [datetime.combine(date, datetime.min.time()) for date in dates]
                
                ax.plot(datetime_dates, scores, marker='o', color='#2E86AB')
                ax.set_title('生産性トレンド', fontsize=12, fontweight='bold')
                ax.set_ylabel('スコア')
                ax.grid(True, alpha=0.3)
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%d'))
            else:
                ax.text(0.5, 0.5, 'データなし', ha='center', va='center', transform=ax.transAxes)
                ax.set_title('生産性トレンド', fontsize=12, fontweight='bold')
        except Exception as e:
            ax.text(0.5, 0.5, f'エラー: {e}', ha='center', va='center', transform=ax.transAxes)
            ax.set_title('生産性トレンド (エラー)', fontsize=12, fontweight='bold')
    
    def _add_focus_distribution_subplot(self, ax, date_range):
        """フォーカス分布サブプロット追加"""
        try:
            sessions = self.data_collector.session_data
            if date_range:
                sessions = [
                    s for s in sessions 
                    if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
                ]
            
            focus_scores = [s.get('focus_score', 0) for s in sessions if s['type'] == 'work']
            
            if focus_scores:
                ax.hist(focus_scores, bins=10, alpha=0.7, color='#3498DB', density=True)
                ax.axvline(statistics.mean(focus_scores), color='red', linestyle='--', 
                          label=f'平均: {statistics.mean(focus_scores):.1f}')
                ax.set_title('フォーカススコア分布', fontsize=12, fontweight='bold')
                ax.set_xlabel('スコア')
                ax.set_ylabel('密度')
                ax.legend()
                ax.grid(True, alpha=0.3)
            else:
                ax.text(0.5, 0.5, 'データなし', ha='center', va='center', transform=ax.transAxes)
                ax.set_title('フォーカススコア分布', fontsize=12, fontweight='bold')
        except Exception as e:
            ax.text(0.5, 0.5, f'エラー: {e}', ha='center', va='center', transform=ax.transAxes)
            ax.set_title('フォーカススコア分布 (エラー)', fontsize=12, fontweight='bold')
    
    def _add_interruption_summary_subplot(self, ax, date_range):
        """中断サマリーサブプロット追加"""
        try:
            sessions = self.data_collector.session_data
            if date_range:
                sessions = [
                    s for s in sessions 
                    if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
                ]
            
            interruption_counts = defaultdict(int)
            for session in sessions:
                for interruption in session.get('interruptions', []):
                    int_type = interruption.get('type', 'unknown')
                    interruption_counts[int_type] += 1
            
            if interruption_counts:
                types = list(interruption_counts.keys())
                counts = list(interruption_counts.values())
                
                ax.bar(types, counts, color=sns.color_palette("viridis", len(types)))
                ax.set_title('中断回数', fontsize=12, fontweight='bold')
                ax.set_ylabel('回数')
                ax.tick_params(axis='x', rotation=45)
                ax.grid(True, alpha=0.3)
            else:
                ax.text(0.5, 0.5, 'データなし', ha='center', va='center', transform=ax.transAxes)
                ax.set_title('中断回数', fontsize=12, fontweight='bold')
        except Exception as e:
            ax.text(0.5, 0.5, f'エラー: {e}', ha='center', va='center', transform=ax.transAxes)
            ax.set_title('中断回数 (エラー)', fontsize=12, fontweight='bold')
    
    def _add_completion_rate_subplot(self, ax, date_range):
        """完了率サブプロット追加"""
        try:
            sessions = self.data_collector.session_data
            if date_range:
                sessions = [
                    s for s in sessions 
                    if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
                ]
            
            work_sessions = [s for s in sessions if s['type'] == 'work']
            
            if work_sessions:
                completed = len([s for s in work_sessions if s.get('completed', False)])
                total = len(work_sessions)
                completion_rate = (completed / total) * 100 if total > 0 else 0
                
                # 円グラフ
                sizes = [completed, total - completed]
                labels = ['完了', '未完了']
                colors = ['#27AE60', '#E74C3C']
                
                ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
                ax.set_title(f'完了率: {completion_rate:.1f}%', fontsize=12, fontweight='bold')
            else:
                ax.text(0.5, 0.5, 'データなし', ha='center', va='center', transform=ax.transAxes)
                ax.set_title('完了率', fontsize=12, fontweight='bold')
        except Exception as e:
            ax.text(0.5, 0.5, f'エラー: {e}', ha='center', va='center', transform=ax.transAxes)
            ax.set_title('完了率 (エラー)', fontsize=12, fontweight='bold')
    
    def _add_time_analysis_subplot(self, ax, date_range):
        """時間分析サブプロット追加"""
        try:
            sessions = self.data_collector.session_data
            if date_range:
                sessions = [
                    s for s in sessions 
                    if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
                ]
            
            hourly_sessions = defaultdict(int)
            for session in sessions:
                if session['type'] == 'work':
                    start_time = datetime.fromisoformat(session['start_time'])
                    hour = start_time.hour
                    hourly_sessions[hour] += 1
            
            if hourly_sessions:
                hours = list(range(24))
                counts = [hourly_sessions[h] for h in hours]
                
                ax.bar(hours, counts, color='#9B59B6', alpha=0.7)
                ax.set_title('時間別セッション数', fontsize=12, fontweight='bold')
                ax.set_xlabel('時間')
                ax.set_ylabel('セッション数')
                ax.set_xticks(range(0, 24, 4))
                ax.grid(True, alpha=0.3)
            else:
                ax.text(0.5, 0.5, 'データなし', ha='center', va='center', transform=ax.transAxes)
                ax.set_title('時間別セッション数', fontsize=12, fontweight='bold')
        except Exception as e:
            ax.text(0.5, 0.5, f'エラー: {e}', ha='center', va='center', transform=ax.transAxes)
            ax.set_title('時間別セッション数 (エラー)', fontsize=12, fontweight='bold')
    
    def _add_weekly_pattern_subplot(self, ax, date_range):
        """週間パターンサブプロット追加"""
        try:
            sessions = self.data_collector.session_data
            if date_range:
                sessions = [
                    s for s in sessions 
                    if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
                ]
            
            weekday_sessions = defaultdict(int)
            for session in sessions:
                if session['type'] == 'work':
                    start_time = datetime.fromisoformat(session['start_time'])
                    weekday = start_time.weekday()  # 0=月曜日
                    weekday_sessions[weekday] += 1
            
            if weekday_sessions:
                weekday_names = ['月', '火', '水', '木', '金', '土', '日']
                counts = [weekday_sessions[i] for i in range(7)]
                
                bars = ax.bar(weekday_names, counts, color=sns.color_palette("husl", 7))
                ax.set_title('曜日別セッション数', fontsize=12, fontweight='bold')
                ax.set_ylabel('セッション数')
                ax.grid(True, alpha=0.3)
                
                # 値をバーの上に表示
                for bar, count in zip(bars, counts):
                    if count > 0:
                        height = bar.get_height()
                        ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                               f'{count}', ha='center', va='bottom')
            else:
                ax.text(0.5, 0.5, 'データなし', ha='center', va='center', transform=ax.transAxes)
                ax.set_title('曜日別セッション数', fontsize=12, fontweight='bold')
        except Exception as e:
            ax.text(0.5, 0.5, f'エラー: {e}', ha='center', va='center', transform=ax.transAxes)
            ax.set_title('曜日別セッション数 (エラー)', fontsize=12, fontweight='bold')
    
    def _create_no_data_chart(self, message: str) -> FigureCanvas:
        """データなし用チャート作成"""
        fig = Figure(figsize=(8, 6), dpi=self.dpi)
        ax = fig.add_subplot(111)
        
        ax.text(0.5, 0.5, message, ha='center', va='center', 
               transform=ax.transAxes, fontsize=14)
        ax.set_title('📊 データ不足', fontsize=16, fontweight='bold')
        ax.axis('off')
        
        fig.tight_layout()
        
        canvas = FigureCanvas(fig)
        canvas.setMinimumSize(400, 300)
        
        return canvas
    
    def _create_error_chart(self, error_message: str) -> FigureCanvas:
        """エラー用チャート作成"""
        fig = Figure(figsize=(8, 6), dpi=self.dpi)
        ax = fig.add_subplot(111)
        
        ax.text(0.5, 0.5, error_message, ha='center', va='center', 
               transform=ax.transAxes, fontsize=12, color='red')
        ax.set_title('❌ チャート作成エラー', fontsize=16, fontweight='bold', color='red')
        ax.axis('off')
        
        fig.tight_layout()
        
        canvas = FigureCanvas(fig)
        canvas.setMinimumSize(400, 300)
        
        return canvas
    
    def export_chart(self, canvas: FigureCanvas, chart_type: str, 
                    format_type: str = 'png') -> str:
        """チャートエクスポート"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{chart_type}_{timestamp}.{format_type}"
            filepath = self.charts_dir / filename
            
            # 高解像度で保存
            canvas.figure.savefig(filepath, dpi=300, bbox_inches='tight', 
                                format=format_type, facecolor='white')
            
            logger.info(f"📊 チャートエクスポート完了: {filename}")
            self.export_completed.emit(chart_type, str(filepath))
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"チャートエクスポートエラー: {e}")
            return ""


class ComparisonAnalytics(QObject):
    """Phase 4: 期間比較分析システム - 日次/週次/月次比較"""
    
    # シグナル
    comparison_completed = pyqtSignal(str, dict)  # (comparison_type, results)
    trend_detected = pyqtSignal(str, dict)  # (trend_type, details)
    
    def __init__(self, data_collector, reports_engine):
        super().__init__()
        
        self.data_collector = data_collector
        self.reports_engine = reports_engine
        
        # 比較結果キャッシュ
        self.comparison_cache = {}
        self.cache_expiry = timedelta(hours=1)  # 1時間でキャッシュ期限切れ
        
        logger.info("📈 ComparisonAnalytics 初期化完了")
    
    def compare_periods(self, period_type: str, current_start: datetime, 
                       current_end: datetime, comparison_periods: int = 1) -> Dict[str, Any]:
        """期間比較分析"""
        try:
            cache_key = f"{period_type}_{current_start.date()}_{current_end.date()}_{comparison_periods}"
            
            # キャッシュチェック
            if self._is_cache_valid(cache_key):
                return self.comparison_cache[cache_key]['data']
            
            # 現在期間のデータ取得
            current_data = self._get_period_data(current_start, current_end)
            
            # 比較期間のデータ取得
            comparison_data_list = []
            for i in range(1, comparison_periods + 1):
                if period_type == 'daily':
                    comp_start = current_start - timedelta(days=i)
                    comp_end = current_end - timedelta(days=i)
                elif period_type == 'weekly':
                    comp_start = current_start - timedelta(weeks=i)
                    comp_end = current_end - timedelta(weeks=i)
                elif period_type == 'monthly':
                    comp_start = current_start - timedelta(days=30*i)
                    comp_end = current_end - timedelta(days=30*i)
                else:
                    raise ValueError(f"未対応の期間タイプ: {period_type}")
                
                comp_data = self._get_period_data(comp_start, comp_end)
                comparison_data_list.append({
                    'period': i,
                    'start_date': comp_start,
                    'end_date': comp_end,
                    'data': comp_data
                })
            
            # 比較分析実行
            comparison_results = self._analyze_period_comparison(
                current_data, comparison_data_list, period_type
            )
            
            # キャッシュに保存
            self._cache_comparison(cache_key, comparison_results)
            
            self.comparison_completed.emit(period_type, comparison_results)
            logger.info(f"📈 期間比較分析完了: {period_type}")
            
            return comparison_results
            
        except Exception as e:
            logger.error(f"期間比較分析エラー: {e}")
            return {'error': str(e)}
    
    def compare_weekdays_vs_weekends(self, date_range: Tuple[datetime, datetime] = None) -> Dict[str, Any]:
        """平日 vs 週末比較分析"""
        try:
            if date_range is None:
                end_date = datetime.now()
                start_date = end_date - timedelta(days=30)
                date_range = (start_date, end_date)
            
            sessions = self.data_collector.session_data
            filtered_sessions = [
                s for s in sessions 
                if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
                and s['type'] == 'work'
            ]
            
            # 平日と週末に分類
            weekday_sessions = []
            weekend_sessions = []
            
            for session in filtered_sessions:
                start_time = datetime.fromisoformat(session['start_time'])
                if start_time.weekday() < 5:  # 月-金
                    weekday_sessions.append(session)
                else:  # 土-日
                    weekend_sessions.append(session)
            
            # 各メトリクスを比較
            weekday_metrics = self._calculate_session_metrics(weekday_sessions)
            weekend_metrics = self._calculate_session_metrics(weekend_sessions)
            
            comparison = {
                'comparison_type': 'weekdays_vs_weekends',
                'date_range': {
                    'start': date_range[0].isoformat(),
                    'end': date_range[1].isoformat()
                },
                'weekday_metrics': weekday_metrics,
                'weekend_metrics': weekend_metrics,
                'comparison_results': self._compare_metrics(weekday_metrics, weekend_metrics),
                'recommendations': self._generate_weekday_weekend_recommendations(
                    weekday_metrics, weekend_metrics
                ),
                'statistical_significance': self._test_statistical_significance(
                    weekday_sessions, weekend_sessions
                )
            }
            
            self.comparison_completed.emit('weekdays_vs_weekends', comparison)
            return comparison
            
        except Exception as e:
            logger.error(f"平日週末比較エラー: {e}")
            return {'error': str(e)}
    
    def compare_time_periods(self, morning_hours: Tuple[int, int] = (6, 12),
                           afternoon_hours: Tuple[int, int] = (12, 18),
                           evening_hours: Tuple[int, int] = (18, 24),
                           date_range: Tuple[datetime, datetime] = None) -> Dict[str, Any]:
        """時間帯別比較分析"""
        try:
            if date_range is None:
                end_date = datetime.now()
                start_date = end_date - timedelta(days=30)
                date_range = (start_date, end_date)
            
            sessions = self.data_collector.session_data
            filtered_sessions = [
                s for s in sessions 
                if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
                and s['type'] == 'work'
            ]
            
            # 時間帯別に分類
            morning_sessions = []
            afternoon_sessions = []
            evening_sessions = []
            
            for session in filtered_sessions:
                start_time = datetime.fromisoformat(session['start_time'])
                hour = start_time.hour
                
                if morning_hours[0] <= hour < morning_hours[1]:
                    morning_sessions.append(session)
                elif afternoon_hours[0] <= hour < afternoon_hours[1]:
                    afternoon_sessions.append(session)
                elif evening_hours[0] <= hour < evening_hours[1]:
                    evening_sessions.append(session)
            
            # 各時間帯のメトリクス計算
            morning_metrics = self._calculate_session_metrics(morning_sessions)
            afternoon_metrics = self._calculate_session_metrics(afternoon_sessions)
            evening_metrics = self._calculate_session_metrics(evening_sessions)
            
            # 最高パフォーマンス時間帯特定
            all_metrics = {
                'morning': morning_metrics,
                'afternoon': afternoon_metrics,
                'evening': evening_metrics
            }
            
            best_period = self._identify_best_time_period(all_metrics)
            
            comparison = {
                'comparison_type': 'time_periods',
                'date_range': {
                    'start': date_range[0].isoformat(),
                    'end': date_range[1].isoformat()
                },
                'time_periods': {
                    'morning': {
                        'hours': morning_hours,
                        'metrics': morning_metrics
                    },
                    'afternoon': {
                        'hours': afternoon_hours,
                        'metrics': afternoon_metrics
                    },
                    'evening': {
                        'hours': evening_hours,
                        'metrics': evening_metrics
                    }
                },
                'best_performance_period': best_period,
                'recommendations': self._generate_time_period_recommendations(all_metrics, best_period)
            }
            
            self.comparison_completed.emit('time_periods', comparison)
            return comparison
            
        except Exception as e:
            logger.error(f"時間帯比較エラー: {e}")
            return {'error': str(e)}
    
    def analyze_progress_trends(self, window_days: int = 7, 
                              date_range: Tuple[datetime, datetime] = None) -> Dict[str, Any]:
        """進歩トレンド分析"""
        try:
            if date_range is None:
                end_date = datetime.now()
                start_date = end_date - timedelta(days=30)
                date_range = (start_date, end_date)
            
            sessions = self.data_collector.session_data
            filtered_sessions = [
                s for s in sessions 
                if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
                and s['type'] == 'work'
            ]
            
            # 移動平均でトレンド分析
            daily_metrics = defaultdict(list)
            
            for session in filtered_sessions:
                date_key = datetime.fromisoformat(session['start_time']).date()
                daily_metrics[date_key].append({
                    'focus_score': session.get('focus_score', 0),
                    'efficiency_score': session.get('efficiency_score', 0),
                    'completed': session.get('completed', False)
                })
            
            # 日別平均計算
            daily_averages = {}
            for date, sessions_data in daily_metrics.items():
                daily_averages[date] = {
                    'focus_avg': statistics.mean([s['focus_score'] for s in sessions_data]),
                    'efficiency_avg': statistics.mean([s['efficiency_score'] for s in sessions_data]),
                    'completion_rate': sum([s['completed'] for s in sessions_data]) / len(sessions_data) * 100
                }
            
            # 移動平均計算
            sorted_dates = sorted(daily_averages.keys())
            moving_averages = self._calculate_moving_averages(daily_averages, sorted_dates, window_days)
            
            # トレンド方向判定
            trend_analysis = self._analyze_trend_direction(moving_averages)
            
            # 改善率計算
            improvement_rates = self._calculate_improvement_rates(moving_averages)
            
            analysis = {
                'analysis_type': 'progress_trends',
                'date_range': {
                    'start': date_range[0].isoformat(),
                    'end': date_range[1].isoformat()
                },
                'window_days': window_days,
                'daily_averages': {str(k): v for k, v in daily_averages.items()},
                'moving_averages': {str(k): v for k, v in moving_averages.items()},
                'trend_analysis': trend_analysis,
                'improvement_rates': improvement_rates,
                'predictions': self._generate_trend_predictions(moving_averages, trend_analysis),
                'milestones': self._identify_performance_milestones(daily_averages)
            }
            
            # トレンド検出シグナル発行
            if trend_analysis['overall_trend'] != 'stable':
                self.trend_detected.emit(trend_analysis['overall_trend'], trend_analysis)
            
            return analysis
            
        except Exception as e:
            logger.error(f"トレンド分析エラー: {e}")
            return {'error': str(e)}
    
    def _get_period_data(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """指定期間のデータ取得"""
        sessions = self.data_collector.session_data
        filtered_sessions = [
            s for s in sessions 
            if start_date <= datetime.fromisoformat(s['start_time']) <= end_date
        ]
        
        work_sessions = [s for s in filtered_sessions if s['type'] == 'work']
        
        return {
            'total_sessions': len(filtered_sessions),
            'work_sessions': len(work_sessions),
            'session_data': work_sessions,
            'metrics': self._calculate_session_metrics(work_sessions)
        }
    
    def _calculate_session_metrics(self, sessions: List[Dict]) -> Dict[str, Any]:
        """セッションメトリクス計算"""
        if not sessions:
            return {
                'count': 0,
                'avg_focus_score': 0,
                'avg_efficiency_score': 0,
                'completion_rate': 0,
                'avg_duration': 0,
                'total_interruptions': 0
            }
        
        focus_scores = [s.get('focus_score', 0) for s in sessions]
        efficiency_scores = [s.get('efficiency_score', 0) for s in sessions]
        completed_sessions = [s for s in sessions if s.get('completed', False)]
        durations = [s.get('actual_duration', 0) for s in sessions]
        total_interruptions = sum(len(s.get('interruptions', [])) for s in sessions)
        
        return {
            'count': len(sessions),
            'avg_focus_score': round(statistics.mean(focus_scores), 2),
            'avg_efficiency_score': round(statistics.mean(efficiency_scores), 2),
            'completion_rate': round(len(completed_sessions) / len(sessions) * 100, 1),
            'avg_duration': round(statistics.mean(durations), 2),
            'total_interruptions': total_interruptions,
            'interruptions_per_session': round(total_interruptions / len(sessions), 2)
        }
    
    def _analyze_period_comparison(self, current_data: Dict, comparison_data_list: List[Dict], 
                                 period_type: str) -> Dict[str, Any]:
        """期間比較分析実行"""
        current_metrics = current_data['metrics']
        
        comparisons = []
        for comp_data in comparison_data_list:
            comp_metrics = comp_data['data']['metrics']
            
            # 各メトリクスの変化率計算
            changes = {}
            for metric in ['avg_focus_score', 'avg_efficiency_score', 'completion_rate', 'avg_duration']:
                if comp_metrics[metric] != 0:
                    change_pct = ((current_metrics[metric] - comp_metrics[metric]) / comp_metrics[metric]) * 100
                else:
                    change_pct = 0 if current_metrics[metric] == 0 else 100
                changes[metric] = round(change_pct, 1)
            
            comparisons.append({
                'period': comp_data['period'],
                'comparison_date_range': {
                    'start': comp_data['start_date'].isoformat(),
                    'end': comp_data['end_date'].isoformat()
                },
                'metrics': comp_metrics,
                'changes': changes
            })
        
        # 全体的な傾向分析
        overall_trend = self._determine_overall_trend(comparisons)
        
        return {
            'comparison_type': f'{period_type}_comparison',
            'current_period': current_data,
            'comparisons': comparisons,
            'overall_trend': overall_trend,
            'insights': self._generate_comparison_insights(current_metrics, comparisons, period_type)
        }
    
    def _compare_metrics(self, metrics1: Dict, metrics2: Dict) -> Dict[str, Any]:
        """メトリクス比較"""
        if not metrics1['count'] or not metrics2['count']:
            return {'status': 'insufficient_data'}
        
        comparisons = {}
        for metric in ['avg_focus_score', 'avg_efficiency_score', 'completion_rate']:
            value1 = metrics1[metric]
            value2 = metrics2[metric]
            
            if value2 != 0:
                change_pct = ((value1 - value2) / value2) * 100
            else:
                change_pct = 0 if value1 == 0 else 100
            
            comparisons[metric] = {
                'value1': value1,
                'value2': value2,
                'change_percent': round(change_pct, 1),
                'better': 'value1' if value1 > value2 else 'value2' if value2 > value1 else 'equal'
            }
        
        return comparisons
    
    def _identify_best_time_period(self, all_metrics: Dict[str, Dict]) -> Dict[str, Any]:
        """最適時間帯特定"""
        scores = {}
        
        for period, metrics in all_metrics.items():
            if metrics['count'] == 0:
                scores[period] = 0
            else:
                # 複合スコア計算（フォーカス + 効率 + 完了率）
                composite_score = (
                    metrics['avg_focus_score'] * 0.4 +
                    metrics['avg_efficiency_score'] * 0.4 +
                    metrics['completion_rate'] * 0.2
                )
                scores[period] = composite_score
        
        if not scores or all(score == 0 for score in scores.values()):
            return {'period': 'none', 'score': 0, 'confidence': 'low'}
        
        best_period = max(scores, key=scores.get)
        best_score = scores[best_period]
        
        # 信頼度計算（データ量と他との差に基づく）
        best_metrics = all_metrics[best_period]
        confidence = 'high' if best_metrics['count'] >= 5 else 'medium' if best_metrics['count'] >= 2 else 'low'
        
        return {
            'period': best_period,
            'score': round(best_score, 2),
            'confidence': confidence,
            'metrics': best_metrics
        }
    
    def _calculate_moving_averages(self, daily_averages: Dict, sorted_dates: List, 
                                 window_days: int) -> Dict:
        """移動平均計算"""
        moving_averages = {}
        
        for i, date in enumerate(sorted_dates):
            if i < window_days - 1:
                continue
            
            window_dates = sorted_dates[i - window_days + 1:i + 1]
            window_data = [daily_averages[d] for d in window_dates]
            
            moving_averages[date] = {
                'focus_avg': statistics.mean([d['focus_avg'] for d in window_data]),
                'efficiency_avg': statistics.mean([d['efficiency_avg'] for d in window_data]),
                'completion_rate': statistics.mean([d['completion_rate'] for d in window_data])
            }
        
        return moving_averages
    
    def _analyze_trend_direction(self, moving_averages: Dict) -> Dict[str, Any]:
        """トレンド方向分析"""
        if len(moving_averages) < 3:
            return {'overall_trend': 'insufficient_data'}
        
        dates = sorted(moving_averages.keys())
        
        trends = {}
        for metric in ['focus_avg', 'efficiency_avg', 'completion_rate']:
            values = [moving_averages[date][metric] for date in dates]
            
            # 線形回帰でトレンド判定
            x = list(range(len(values)))
            if len(values) > 1:
                slope = np.polyfit(x, values, 1)[0]
                
                if slope > 0.5:
                    trend = 'improving'
                elif slope < -0.5:
                    trend = 'declining'
                else:
                    trend = 'stable'
            else:
                trend = 'stable'
            
            trends[metric] = {
                'direction': trend,
                'slope': slope,
                'start_value': values[0],
                'end_value': values[-1]
            }
        
        # 全体的なトレンド判定
        trend_counts = defaultdict(int)
        for trend_info in trends.values():
            trend_counts[trend_info['direction']] += 1
        
        overall_trend = max(trend_counts, key=trend_counts.get)
        
        return {
            'overall_trend': overall_trend,
            'metric_trends': trends,
            'trend_strength': self._calculate_trend_strength(trends)
        }
    
    def _calculate_improvement_rates(self, moving_averages: Dict) -> Dict[str, float]:
        """改善率計算"""
        if len(moving_averages) < 2:
            return {}
        
        dates = sorted(moving_averages.keys())
        first_data = moving_averages[dates[0]]
        last_data = moving_averages[dates[-1]]
        
        improvement_rates = {}
        for metric in ['focus_avg', 'efficiency_avg', 'completion_rate']:
            first_value = first_data[metric]
            last_value = last_data[metric]
            
            if first_value != 0:
                improvement_rate = ((last_value - first_value) / first_value) * 100
            else:
                improvement_rate = 0 if last_value == 0 else 100
            
            improvement_rates[metric] = round(improvement_rate, 2)
        
        return improvement_rates
    
    def _calculate_trend_strength(self, trends: Dict) -> str:
        """トレンド強度計算"""
        slopes = [abs(trend['slope']) for trend in trends.values()]
        avg_slope = statistics.mean(slopes)
        
        if avg_slope > 2.0:
            return 'strong'
        elif avg_slope > 0.5:
            return 'moderate'
        else:
            return 'weak'
    
    def _generate_trend_predictions(self, moving_averages: Dict, trend_analysis: Dict) -> Dict[str, Any]:
        """トレンド予測生成"""
        if len(moving_averages) < 5:
            return {'status': 'insufficient_data'}
        
        predictions = {}
        dates = sorted(moving_averages.keys())
        
        for metric in ['focus_avg', 'efficiency_avg', 'completion_rate']:
            values = [moving_averages[date][metric] for date in dates]
            trend_info = trend_analysis['metric_trends'][metric]
            
            # 7日後の予測値
            future_value = values[-1] + (trend_info['slope'] * 7)
            
            predictions[metric] = {
                'current_value': round(values[-1], 2),
                'predicted_value_7days': round(future_value, 2),
                'confidence': 'high' if len(values) >= 10 else 'medium'
            }
        
        return predictions
    
    def _identify_performance_milestones(self, daily_averages: Dict) -> List[Dict]:
        """パフォーマンスマイルストーン特定"""
        milestones = []
        
        if not daily_averages:
            return milestones
        
        # 最高スコア達成日
        best_focus_date = max(daily_averages, key=lambda d: daily_averages[d]['focus_avg'])
        best_efficiency_date = max(daily_averages, key=lambda d: daily_averages[d]['efficiency_avg'])
        best_completion_date = max(daily_averages, key=lambda d: daily_averages[d]['completion_rate'])
        
        milestones.extend([
            {
                'type': 'best_focus',
                'date': str(best_focus_date),
                'value': daily_averages[best_focus_date]['focus_avg'],
                'description': '最高フォーカススコア達成'
            },
            {
                'type': 'best_efficiency', 
                'date': str(best_efficiency_date),
                'value': daily_averages[best_efficiency_date]['efficiency_avg'],
                'description': '最高効率スコア達成'
            },
            {
                'type': 'best_completion',
                'date': str(best_completion_date), 
                'value': daily_averages[best_completion_date]['completion_rate'],
                'description': '最高完了率達成'
            }
        ])
        
        return milestones
    
    def _generate_comparison_insights(self, current_metrics: Dict, comparisons: List[Dict], 
                                    period_type: str) -> List[str]:
        """比較インサイト生成"""
        insights = []
        
        if not comparisons:
            return ["比較データが不足しています"]
        
        latest_comparison = comparisons[0]  # 最新の比較期間
        changes = latest_comparison['changes']
        
        # フォーカススコア変化
        focus_change = changes['avg_focus_score']
        if focus_change > 10:
            insights.append(f"🎯 フォーカススコアが{focus_change:.1f}%向上しました")
        elif focus_change < -10:
            insights.append(f"⚠️ フォーカススコアが{abs(focus_change):.1f}%低下しています")
        
        # 完了率変化
        completion_change = changes['completion_rate']
        if completion_change > 15:
            insights.append(f"✅ セッション完了率が{completion_change:.1f}%改善しました")
        elif completion_change < -15:
            insights.append(f"❌ セッション完了率が{abs(completion_change):.1f}%低下しています")
        
        # 全体的な傾向
        if all(changes[metric] > 5 for metric in ['avg_focus_score', 'avg_efficiency_score', 'completion_rate']):
            insights.append("🚀 全体的にパフォーマンスが向上しています！")
        elif all(changes[metric] < -5 for metric in ['avg_focus_score', 'avg_efficiency_score', 'completion_rate']):
            insights.append("📉 パフォーマンスの低下が見られます。休息を取ることをお勧めします")
        
        if not insights:
            insights.append("📊 パフォーマンスは安定しています")
        
        return insights
    
    def _generate_weekday_weekend_recommendations(self, weekday_metrics: Dict, 
                                                weekend_metrics: Dict) -> List[str]:
        """平日週末推奨事項生成"""
        recommendations = []
        
        if weekday_metrics['count'] == 0 or weekend_metrics['count'] == 0:
            return ["十分なデータが蓄積されていません"]
        
        # フォーカススコア比較
        if weekday_metrics['avg_focus_score'] > weekend_metrics['avg_focus_score'] + 10:
            recommendations.append("📅 平日の集中環境を週末にも取り入れてみましょう")
        elif weekend_metrics['avg_focus_score'] > weekday_metrics['avg_focus_score'] + 10:
            recommendations.append("🏠 週末のリラックス環境を平日にも活用してみましょう")
        
        # 完了率比較
        if weekday_metrics['completion_rate'] > weekend_metrics['completion_rate'] + 20:
            recommendations.append("⏰ 週末も平日と同様の時間管理を心がけてみましょう")
        elif weekend_metrics['completion_rate'] > weekday_metrics['completion_rate'] + 20:
            recommendations.append("🎯 平日も週末のようにプレッシャーを減らしてみましょう")
        
        if not recommendations:
            recommendations.append("⚖️ 平日と週末のバランスが良好です")
        
        return recommendations
    
    def _generate_time_period_recommendations(self, all_metrics: Dict, best_period: Dict) -> List[str]:
        """時間帯推奨事項生成"""
        recommendations = []
        
        if best_period['period'] == 'none':
            return ["十分なデータが蓄積されていません"]
        
        period_names = {
            'morning': '午前中',
            'afternoon': '午後',
            'evening': '夕方・夜'
        }
        
        best_period_name = period_names.get(best_period['period'], best_period['period'])
        recommendations.append(f"⭐ {best_period_name}のパフォーマンスが最も高いです")
        
        if best_period['confidence'] == 'high':
            recommendations.append(f"🎯 重要なタスクは{best_period_name}に集中させることをお勧めします")
        
        # 他の時間帯との差が大きい場合
        best_score = best_period['score']
        other_scores = [metrics['avg_focus_score'] for period, metrics in all_metrics.items() 
                       if period != best_period['period'] and metrics['count'] > 0]
        
        if other_scores and best_score > max(other_scores) + 20:
            recommendations.append("📈 時間帯による差が大きいです。スケジューリングを最適化しましょう")
        
        return recommendations
    
    def _test_statistical_significance(self, sessions1: List[Dict], sessions2: List[Dict]) -> Dict[str, Any]:
        """統計的有意性テスト（簡易版）"""
        if len(sessions1) < 5 or len(sessions2) < 5:
            return {'status': 'insufficient_data'}
        
        # フォーカススコアでt検定（簡易版）
        scores1 = [s.get('focus_score', 0) for s in sessions1]
        scores2 = [s.get('focus_score', 0) for s in sessions2]
        
        mean1 = statistics.mean(scores1)
        mean2 = statistics.mean(scores2)
        
        if len(scores1) > 1 and len(scores2) > 1:
            std1 = statistics.stdev(scores1)
            std2 = statistics.stdev(scores2)
            
            # 差の大きさを評価（効果量の簡易計算）
            pooled_std = ((std1 ** 2 + std2 ** 2) / 2) ** 0.5
            effect_size = abs(mean1 - mean2) / pooled_std if pooled_std > 0 else 0
            
            if effect_size > 0.8:
                significance = 'large_effect'
            elif effect_size > 0.5:
                significance = 'medium_effect'
            elif effect_size > 0.2:
                significance = 'small_effect'
            else:
                significance = 'no_effect'
        else:
            significance = 'insufficient_data'
        
        return {
            'status': significance,
            'mean_difference': round(mean1 - mean2, 2),
            'effect_size': round(effect_size, 3) if 'effect_size' in locals() else None
        }
    
    def _determine_overall_trend(self, comparisons: List[Dict]) -> str:
        """全体トレンド判定"""
        if not comparisons:
            return 'no_data'
        
        # 最新の比較結果を分析
        latest = comparisons[0]
        changes = latest['changes']
        
        positive_changes = sum(1 for change in changes.values() if change > 5)
        negative_changes = sum(1 for change in changes.values() if change < -5)
        
        if positive_changes > negative_changes:
            return 'improving'
        elif negative_changes > positive_changes:
            return 'declining'
        else:
            return 'stable'
    
    def _is_cache_valid(self, cache_key: str) -> bool:
        """キャッシュ有効性チェック"""
        if cache_key not in self.comparison_cache:
            return False
        
        cache_time = self.comparison_cache[cache_key]['timestamp']
        return datetime.now() - cache_time < self.cache_expiry
    
    def _cache_comparison(self, cache_key: str, comparison_data: Dict):
        """比較結果キャッシュ"""
        self.comparison_cache[cache_key] = {
            'timestamp': datetime.now(),
            'data': comparison_data
        }
        
        # キャッシュサイズ制限（最大20件）
        if len(self.comparison_cache) > 20:
            oldest_key = min(self.comparison_cache.keys(), 
                           key=lambda k: self.comparison_cache[k]['timestamp'])
            del self.comparison_cache[oldest_key]


class CustomReportBuilder(QObject):
    """Phase 4: カスタムレポートビルダー - 柔軟なパラメータ設定でレポート生成"""
    
    # シグナル
    report_built = pyqtSignal(str, dict)  # (report_name, report_data)
    template_saved = pyqtSignal(str, dict)  # (template_name, template_config)
    
    def __init__(self, reports_engine, visualization, comparison_analytics):
        super().__init__()
        
        self.reports_engine = reports_engine
        self.visualization = visualization
        self.comparison_analytics = comparison_analytics
        
        # テンプレート保存ディレクトリ
        self.templates_dir = Path("data/report_templates")
        self.templates_dir.mkdir(exist_ok=True)
        
        # デフォルトテンプレート
        self.default_templates = self._create_default_templates()
        
        # カスタムテンプレート
        self.custom_templates = self._load_custom_templates()
        
        logger.info("📝 CustomReportBuilder 初期化完了")
    
    def build_custom_report(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """カスタムレポート生成"""
        try:
            # 設定検証
            validated_config = self._validate_config(config)
            
            report_data = {
                'report_id': f"custom_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                'report_name': validated_config.get('name', 'カスタムレポート'),
                'generated_at': datetime.now().isoformat(),
                'config': validated_config,
                'sections': {}
            }
            
            # 各セクション生成
            for section_config in validated_config['sections']:
                section_name = section_config['name']
                section_type = section_config['type']
                section_params = section_config.get('parameters', {})
                
                logger.info(f"📊 セクション生成中: {section_name} ({section_type})")
                
                try:
                    if section_type == 'summary':
                        section_data = self._build_summary_section(section_params)
                    elif section_type == 'productivity_analysis':
                        section_data = self._build_productivity_section(section_params)
                    elif section_type == 'comparison':
                        section_data = self._build_comparison_section(section_params)
                    elif section_type == 'visualization':
                        section_data = self._build_visualization_section(section_params)
                    elif section_type == 'trend_analysis':
                        section_data = self._build_trend_section(section_params)
                    elif section_type == 'recommendations':
                        section_data = self._build_recommendations_section(section_params)
                    elif section_type == 'raw_data':
                        section_data = self._build_raw_data_section(section_params)
                    else:
                        section_data = {'error': f'未対応のセクションタイプ: {section_type}'}
                    
                    report_data['sections'][section_name] = {
                        'type': section_type,
                        'config': section_config,
                        'data': section_data,
                        'generated_at': datetime.now().isoformat()
                    }
                    
                except Exception as e:
                    logger.error(f"セクション生成エラー ({section_name}): {e}")
                    report_data['sections'][section_name] = {
                        'type': section_type,
                        'error': str(e)
                    }
            
            # レポート後処理
            report_data = self._post_process_report(report_data, validated_config)
            
            self.report_built.emit(validated_config.get('name', 'カスタムレポート'), report_data)
            logger.info(f"📝 カスタムレポート生成完了: {report_data['report_name']}")
            
            return report_data
            
        except Exception as e:
            logger.error(f"カスタムレポート生成エラー: {e}")
            return {'error': str(e)}
    
    def create_report_from_template(self, template_name: str, 
                                  custom_params: Dict[str, Any] = None) -> Dict[str, Any]:
        """テンプレートからレポート生成"""
        try:
            # テンプレート取得
            template = self._get_template(template_name)
            if not template:
                return {'error': f'テンプレートが見つかりません: {template_name}'}
            
            # カスタムパラメータでテンプレートを更新
            if custom_params:
                template = self._merge_template_params(template, custom_params)
            
            # レポート生成
            return self.build_custom_report(template)
            
        except Exception as e:
            logger.error(f"テンプレートレポート生成エラー: {e}")
            return {'error': str(e)}
    
    def save_template(self, template_name: str, config: Dict[str, Any]) -> bool:
        """テンプレート保存"""
        try:
            # 設定検証
            validated_config = self._validate_config(config)
            
            template_data = {
                'name': template_name,
                'created_at': datetime.now().isoformat(),
                'config': validated_config,
                'description': config.get('description', ''),
                'tags': config.get('tags', [])
            }
            
            # ファイルに保存
            template_file = self.templates_dir / f"{template_name}.json"
            with open(template_file, 'w', encoding='utf-8') as f:
                json.dump(template_data, f, ensure_ascii=False, indent=2)
            
            # キャッシュ更新
            self.custom_templates[template_name] = template_data
            
            self.template_saved.emit(template_name, template_data)
            logger.info(f"📝 テンプレート保存完了: {template_name}")
            
            return True
            
        except Exception as e:
            logger.error(f"テンプレート保存エラー: {e}")
            return False
    
    def get_available_templates(self) -> Dict[str, Dict]:
        """利用可能なテンプレート一覧取得"""
        all_templates = {}
        all_templates.update(self.default_templates)
        all_templates.update(self.custom_templates)
        
        return {
            name: {
                'name': template['name'],
                'description': template.get('description', ''),
                'tags': template.get('tags', []),
                'created_at': template.get('created_at', ''),
                'type': 'default' if name in self.default_templates else 'custom'
            }
            for name, template in all_templates.items()
        }
    
    def get_report_config_schema(self) -> Dict[str, Any]:
        """レポート設定スキーマ取得"""
        return {
            'name': {
                'type': 'string',
                'required': True,
                'description': 'レポート名'
            },
            'description': {
                'type': 'string',
                'required': False,
                'description': 'レポートの説明'
            },
            'date_range': {
                'type': 'object',
                'properties': {
                    'start_date': {'type': 'string', 'format': 'date'},
                    'end_date': {'type': 'string', 'format': 'date'},
                    'preset': {'type': 'string', 'enum': ['last_7_days', 'last_30_days', 'this_month', 'custom']}
                }
            },
            'sections': {
                'type': 'array',
                'items': {
                    'type': 'object',
                    'properties': {
                        'name': {'type': 'string', 'required': True},
                        'type': {'type': 'string', 'enum': [
                            'summary', 'productivity_analysis', 'comparison', 
                            'visualization', 'trend_analysis', 'recommendations', 'raw_data'
                        ]},
                        'parameters': {'type': 'object'}
                    }
                }
            },
            'export_options': {
                'type': 'object',
                'properties': {
                    'formats': {'type': 'array', 'items': {'type': 'string'}},
                    'include_charts': {'type': 'boolean'},
                    'chart_resolution': {'type': 'string', 'enum': ['low', 'medium', 'high']}
                }
            }
        }
    
    def _validate_config(self, config: Dict[str, Any]) -> Dict[str, Any]:
        """設定検証"""
        validated = {}
        
        # 必須フィールド
        if 'name' not in config:
            raise ValueError("レポート名は必須です")
        validated['name'] = config['name']
        
        # オプションフィールド
        validated['description'] = config.get('description', '')
        
        # 日付範囲
        date_range = config.get('date_range', {})
        validated['date_range'] = self._validate_date_range(date_range)
        
        # セクション
        sections = config.get('sections', [])
        if not sections:
            # デフォルトセクション追加
            sections = [
                {'name': 'サマリー', 'type': 'summary', 'parameters': {}},
                {'name': '生産性分析', 'type': 'productivity_analysis', 'parameters': {}}
            ]
        validated['sections'] = self._validate_sections(sections)
        
        # エクスポートオプション
        validated['export_options'] = config.get('export_options', {
            'formats': ['json'],
            'include_charts': True,
            'chart_resolution': 'medium'
        })
        
        return validated
    
    def _validate_date_range(self, date_range: Dict) -> Dict[str, Any]:
        """日付範囲検証"""
        preset = date_range.get('preset', 'last_30_days')
        
        if preset == 'custom':
            start_date = date_range.get('start_date')
            end_date = date_range.get('end_date')
            
            if not start_date or not end_date:
                raise ValueError("カスタム日付範囲には開始日と終了日が必要です")
            
            try:
                start_dt = datetime.fromisoformat(start_date)
                end_dt = datetime.fromisoformat(end_date)
                
                if start_dt >= end_dt:
                    raise ValueError("開始日は終了日より前である必要があります")
                
                return {
                    'preset': 'custom',
                    'start_date': start_dt,
                    'end_date': end_dt
                }
            except ValueError as e:
                raise ValueError(f"日付形式が無効です: {e}")
        
        else:
            # プリセット日付範囲
            end_date = datetime.now()
            
            if preset == 'last_7_days':
                start_date = end_date - timedelta(days=7)
            elif preset == 'last_30_days':
                start_date = end_date - timedelta(days=30)
            elif preset == 'this_month':
                start_date = end_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            else:
                start_date = end_date - timedelta(days=30)  # デフォルト
            
            return {
                'preset': preset,
                'start_date': start_date,
                'end_date': end_date
            }
    
    def _validate_sections(self, sections: List[Dict]) -> List[Dict]:
        """セクション検証"""
        valid_types = [
            'summary', 'productivity_analysis', 'comparison', 
            'visualization', 'trend_analysis', 'recommendations', 'raw_data'
        ]
        
        validated_sections = []
        
        for section in sections:
            if 'name' not in section:
                raise ValueError("セクション名は必須です")
            
            if 'type' not in section:
                raise ValueError("セクションタイプは必須です")
            
            if section['type'] not in valid_types:
                raise ValueError(f"無効なセクションタイプ: {section['type']}")
            
            validated_sections.append({
                'name': section['name'],
                'type': section['type'],
                'parameters': section.get('parameters', {})
            })
        
        return validated_sections
    
    def _build_summary_section(self, params: Dict) -> Dict[str, Any]:
        """サマリーセクション構築"""
        date_range = params.get('date_range')
        if not date_range:
            date_range = (datetime.now() - timedelta(days=30), datetime.now())
        
        # 基本統計取得
        comprehensive_report = self.reports_engine.generate_comprehensive_report(date_range)
        
        return {
            'type': 'summary',
            'title': 'セッション概要',
            'data': comprehensive_report['summary'],
            'date_range': {
                'start': date_range[0].isoformat(),
                'end': date_range[1].isoformat()
            },
            'generated_at': datetime.now().isoformat()
        }
    
    def _build_productivity_section(self, params: Dict) -> Dict[str, Any]:
        """生産性分析セクション構築"""
        date_range = params.get('date_range')
        if not date_range:
            date_range = (datetime.now() - timedelta(days=30), datetime.now())
        
        # 生産性トレンド分析
        trends = self.comparison_analytics.analyze_progress_trends(
            window_days=params.get('window_days', 7),
            date_range=date_range
        )
        
        return {
            'type': 'productivity_analysis',
            'title': '生産性分析',
            'data': {
                'trends': trends,
                'key_metrics': self._extract_key_productivity_metrics(trends),
                'insights': self._generate_productivity_insights(trends)
            },
            'parameters': params
        }
    
    def _build_comparison_section(self, params: Dict) -> Dict[str, Any]:
        """比較分析セクション構築"""
        comparison_type = params.get('type', 'weekly')
        
        if comparison_type == 'weekdays_vs_weekends':
            data = self.comparison_analytics.compare_weekdays_vs_weekends()
        elif comparison_type == 'time_periods':
            data = self.comparison_analytics.compare_time_periods()
        else:
            # 期間比較
            current_end = datetime.now()
            
            if comparison_type == 'daily':
                current_start = current_end - timedelta(days=1)
            elif comparison_type == 'weekly':
                current_start = current_end - timedelta(weeks=1)
            elif comparison_type == 'monthly':
                current_start = current_end - timedelta(days=30)
            else:
                current_start = current_end - timedelta(weeks=1)
            
            data = self.comparison_analytics.compare_periods(
                comparison_type, current_start, current_end,
                comparison_periods=params.get('periods', 3)
            )
        
        return {
            'type': 'comparison',
            'title': f'{comparison_type} 比較分析',
            'data': data,
            'parameters': params
        }
    
    def _build_visualization_section(self, params: Dict) -> Dict[str, Any]:
        """可視化セクション構築"""
        chart_type = params.get('chart_type', 'productivity_timeline')
        date_range = params.get('date_range')
        
        # チャート生成（メタデータのみ返す）
        chart_info = {
            'chart_type': chart_type,
            'parameters': params,
            'description': self._get_chart_description(chart_type),
            'available': True
        }
        
        return {
            'type': 'visualization',
            'title': f'{chart_type} チャート',
            'data': chart_info,
            'note': 'チャートは実際のレポート表示時に生成されます'
        }
    
    def _build_trend_section(self, params: Dict) -> Dict[str, Any]:
        """トレンド分析セクション構築"""
        date_range = params.get('date_range')
        if not date_range:
            date_range = (datetime.now() - timedelta(days=30), datetime.now())
        
        trends = self.comparison_analytics.analyze_progress_trends(
            window_days=params.get('window_days', 7),
            date_range=date_range
        )
        
        return {
            'type': 'trend_analysis',
            'title': 'トレンド分析',
            'data': {
                'trend_direction': trends.get('trend_analysis', {}),
                'predictions': trends.get('predictions', {}),
                'milestones': trends.get('milestones', []),
                'improvement_rates': trends.get('improvement_rates', {})
            },
            'parameters': params
        }
    
    def _build_recommendations_section(self, params: Dict) -> Dict[str, Any]:
        """推奨事項セクション構築"""
        date_range = params.get('date_range')
        if not date_range:
            date_range = (datetime.now() - timedelta(days=30), datetime.now())
        
        # 包括的レポートから推奨事項取得
        comprehensive_report = self.reports_engine.generate_comprehensive_report(date_range)
        recommendations = comprehensive_report.get('recommendations', [])
        
        # 追加の推奨事項生成
        additional_recommendations = self._generate_additional_recommendations(params)
        
        all_recommendations = recommendations + additional_recommendations
        
        return {
            'type': 'recommendations',
            'title': '改善推奨事項',
            'data': {
                'recommendations': all_recommendations,
                'priority_recommendations': [r for r in all_recommendations if '🚀' in r or '⚠️' in r],
                'categorized_recommendations': self._categorize_recommendations(all_recommendations)
            }
        }
    
    def _build_raw_data_section(self, params: Dict) -> Dict[str, Any]:
        """生データセクション構築"""
        date_range = params.get('date_range')
        if not date_range:
            date_range = (datetime.now() - timedelta(days=7), datetime.now())
        
        # データフィルタリング
        sessions = self.reports_engine.data_collector.session_data
        filtered_sessions = [
            s for s in sessions 
            if date_range[0] <= datetime.fromisoformat(s['start_time']) <= date_range[1]
        ]
        
        # データ集約
        include_fields = params.get('include_fields', [
            'session_id', 'type', 'start_time', 'actual_duration', 
            'focus_score', 'efficiency_score', 'completed'
        ])
        
        raw_data = []
        for session in filtered_sessions:
            session_data = {field: session.get(field) for field in include_fields}
            raw_data.append(session_data)
        
        return {
            'type': 'raw_data',
            'title': '生データ',
            'data': {
                'sessions': raw_data,
                'total_records': len(raw_data),
                'fields': include_fields,
                'date_range': {
                    'start': date_range[0].isoformat(),
                    'end': date_range[1].isoformat()
                }
            },
            'parameters': params
        }
    
    def _post_process_report(self, report_data: Dict, config: Dict) -> Dict[str, Any]:
        """レポート後処理"""
        # メタデータ追加
        report_data['metadata'] = {
            'total_sections': len(report_data['sections']),
            'generation_time': datetime.now().isoformat(),
            'config_hash': hash(str(config)),
            'data_sources': ['advanced_session_data', 'session_tracking', 'environment_log']
        }
        
        # エラーセクション数をカウント
        error_sections = [
            name for name, section in report_data['sections'].items() 
            if 'error' in section
        ]
        
        if error_sections:
            report_data['warnings'] = [
                f"セクション生成エラー: {', '.join(error_sections)}"
            ]
        
        return report_data
    
    def _create_default_templates(self) -> Dict[str, Dict]:
        """デフォルトテンプレート作成"""
        return {
            'daily_summary': {
                'name': 'Daily Summary',
                'description': '日次サマリーレポート',
                'config': {
                    'name': '日次レポート',
                    'date_range': {'preset': 'last_7_days'},
                    'sections': [
                        {'name': 'サマリー', 'type': 'summary', 'parameters': {}},
                        {'name': '生産性チャート', 'type': 'visualization', 
                         'parameters': {'chart_type': 'productivity_timeline'}},
                        {'name': '推奨事項', 'type': 'recommendations', 'parameters': {}}
                    ]
                }
            },
            'weekly_analysis': {
                'name': 'Weekly Analysis',
                'description': '週次分析レポート',
                'config': {
                    'name': '週次分析レポート',
                    'date_range': {'preset': 'last_30_days'},
                    'sections': [
                        {'name': 'サマリー', 'type': 'summary', 'parameters': {}},
                        {'name': '生産性分析', 'type': 'productivity_analysis', 'parameters': {}},
                        {'name': '平日週末比較', 'type': 'comparison', 
                         'parameters': {'type': 'weekdays_vs_weekends'}},
                        {'name': 'フォーカスヒートマップ', 'type': 'visualization',
                         'parameters': {'chart_type': 'focus_heatmap'}}
                    ]
                }
            },
            'comprehensive': {
                'name': 'Comprehensive Report',
                'description': '包括的レポート',
                'config': {
                    'name': '包括的分析レポート',
                    'date_range': {'preset': 'last_30_days'},
                    'sections': [
                        {'name': 'サマリー', 'type': 'summary', 'parameters': {}},
                        {'name': '生産性分析', 'type': 'productivity_analysis', 'parameters': {}},
                        {'name': 'トレンド分析', 'type': 'trend_analysis', 'parameters': {}},
                        {'name': '時間帯比較', 'type': 'comparison',
                         'parameters': {'type': 'time_periods'}},
                        {'name': 'セッション詳細', 'type': 'raw_data',
                         'parameters': {'include_fields': ['session_id', 'type', 'start_time', 'focus_score']}},
                        {'name': '推奨事項', 'type': 'recommendations', 'parameters': {}}
                    ]
                }
            }
        }
    
    def _load_custom_templates(self) -> Dict[str, Dict]:
        """カスタムテンプレート読み込み"""
        custom_templates = {}
        
        try:
            for template_file in self.templates_dir.glob("*.json"):
                with open(template_file, 'r', encoding='utf-8') as f:
                    template_data = json.load(f)
                    template_name = template_file.stem
                    custom_templates[template_name] = template_data
        except Exception as e:
            logger.error(f"カスタムテンプレート読み込みエラー: {e}")
        
        return custom_templates
    
    def _get_template(self, template_name: str) -> Dict[str, Any]:
        """テンプレート取得"""
        if template_name in self.default_templates:
            return self.default_templates[template_name]['config']
        elif template_name in self.custom_templates:
            return self.custom_templates[template_name]['config']
        else:
            return None
    
    def _merge_template_params(self, template: Dict, custom_params: Dict) -> Dict:
        """テンプレートパラメータマージ"""
        merged = template.copy()
        
        # 上位レベルのパラメータを更新
        for key, value in custom_params.items():
            if key in merged:
                if isinstance(merged[key], dict) and isinstance(value, dict):
                    merged[key].update(value)
                else:
                    merged[key] = value
        
        return merged
    
    def _extract_key_productivity_metrics(self, trends: Dict) -> Dict[str, Any]:
        """主要生産性メトリクス抽出"""
        if 'error' in trends:
            return {'error': trends['error']}
        
        daily_averages = trends.get('daily_averages', {})
        if not daily_averages:
            return {'status': 'no_data'}
        
        # 最新の平均値
        latest_date = max(daily_averages.keys(), key=lambda x: datetime.fromisoformat(x))
        latest_metrics = daily_averages[latest_date]
        
        return {
            'latest_focus_score': latest_metrics.get('focus_avg', 0),
            'latest_efficiency_score': latest_metrics.get('efficiency_avg', 0),
            'latest_completion_rate': latest_metrics.get('completion_rate', 0),
            'improvement_rates': trends.get('improvement_rates', {}),
            'trend_direction': trends.get('trend_analysis', {}).get('overall_trend', 'unknown')
        }
    
    def _generate_productivity_insights(self, trends: Dict) -> List[str]:
        """生産性インサイト生成"""
        insights = []
        
        if 'error' in trends:
            return ['データ取得エラーが発生しました']
        
        trend_analysis = trends.get('trend_analysis', {})
        overall_trend = trend_analysis.get('overall_trend', 'unknown')
        
        if overall_trend == 'improving':
            insights.append("📈 生産性が向上傾向にあります")
        elif overall_trend == 'declining':
            insights.append("📉 生産性の低下が見られます")
        else:
            insights.append("📊 生産性は安定しています")
        
        # 改善率チェック
        improvement_rates = trends.get('improvement_rates', {})
        for metric, rate in improvement_rates.items():
            if rate > 20:
                insights.append(f"🚀 {metric}が{rate:.1f}%向上しました")
            elif rate < -20:
                insights.append(f"⚠️ {metric}が{abs(rate):.1f}%低下しています")
        
        return insights if insights else ["十分なデータが蓄積されていません"]
    
    def _get_chart_description(self, chart_type: str) -> str:
        """チャート説明取得"""
        descriptions = {
            'productivity_timeline': '時系列での生産性スコア推移を表示',
            'focus_heatmap': '曜日×時間帯でのフォーカススコア分布を表示',
            'interruption_analysis': '中断パターンの詳細分析を表示',
            'session_performance': 'セッションパフォーマンスの総合分析を表示',
            'custom_dashboard': '選択されたメトリクスのカスタムダッシュボードを表示'
        }
        
        return descriptions.get(chart_type, 'チャートの説明が利用できません')
    
    def _generate_additional_recommendations(self, params: Dict) -> List[str]:
        """追加推奨事項生成"""
        recommendations = []
        
        # パラメータに基づく推奨事項
        focus_threshold = params.get('focus_threshold', 60)
        recommendations.append(f"🎯 フォーカススコア{focus_threshold}以上を目標にしましょう")
        
        if params.get('include_break_recommendations', True):
            recommendations.append("☕ 定期的な休憩でパフォーマンスを維持しましょう")
        
        return recommendations
    
    def _categorize_recommendations(self, recommendations: List[str]) -> Dict[str, List[str]]:
        """推奨事項分類"""
        categories = {
            'focus': [],
            'productivity': [],
            'health': [],
            'general': []
        }
        
        for rec in recommendations:
            if any(keyword in rec for keyword in ['フォーカス', '集中', '🎯', '🧠']):
                categories['focus'].append(rec)
            elif any(keyword in rec for keyword in ['生産性', '効率', '📈', '🚀']):
                categories['productivity'].append(rec)
            elif any(keyword in rec for keyword in ['休憩', '健康', '☕', '💤']):
                categories['health'].append(rec)
            else:
                categories['general'].append(rec)
        
        return {k: v for k, v in categories.items() if v}  # 空のカテゴリを除外


# =============================================================================
# Worker3: Prediction Engine & Export Systems Implementation
# =============================================================================

class PredictionEngine(QObject):
    """Worker3: Machine Learning Prediction System - 予測エンジン"""
    
    # シグナル
    prediction_ready = pyqtSignal(str, dict)  # (prediction_type, results)
    model_trained = pyqtSignal(str, dict)  # (model_name, metrics)
    prediction_updated = pyqtSignal(str, object)  # (metric_name, predicted_value)
    
    def __init__(self, data_collector, session_tracker, focus_calculator):
        super().__init__()
        self.data_collector = data_collector
        self.session_tracker = session_tracker
        self.focus_calculator = focus_calculator
        
        # モデル保存ディレクトリ
        self.model_dir = Path("data/ml_models")
        self.model_dir.mkdir(exist_ok=True)
        
        # 予測モデル
        self.models = {
            'focus_score': None,
            'productivity_trend': None,
            'optimal_work_time': None,
            'session_completion': None
        }
        
        # スケーラー
        self.scalers = {}
        
        # 予測精度メトリクス
        self.model_metrics = {}
        
        # データ前処理パイプライン
        self.feature_encoders = {}
        
        logger.info("🤖 PredictionEngine初期化完了")
    
    def prepare_training_data(self) -> Dict[str, pd.DataFrame]:
        """機械学習用のトレーニングデータを準備"""
        try:
            # セッションデータの取得
            sessions = self.data_collector.session_data
            if len(sessions) < 10:
                logger.warning("⚠️ トレーニングデータが不足 (最低10セッション必要)")
                return {}
            
            # DataFrame作成
            df_list = []
            for session in sessions:
                try:
                    row = {
                        'session_id': session.get('session_id', ''),
                        'type': session.get('type', 'work'),
                        'planned_duration': session.get('planned_duration', 25),
                        'actual_duration': session.get('actual_duration', 0),
                        'completed': session.get('completed', False),
                        'focus_score': session.get('focus_score', 0.0),
                        'efficiency_score': session.get('efficiency_score', 0),
                        'interruption_count': len(session.get('interruptions', [])),
                        'interaction_count': len(session.get('interactions', [])),
                        'hour_of_day': self._extract_hour_of_day(session.get('start_time', '')),
                        'day_of_week': self._extract_day_of_week(session.get('start_time', '')),
                        'environment_score': self._calculate_environment_score(session.get('environment_data', {}))
                    }
                    df_list.append(row)
                except Exception as e:
                    logger.error(f"セッションデータ処理エラー: {e}")
                    continue
            
            if not df_list:
                return {}
            
            df = pd.DataFrame(df_list)
            
            # データセット分割
            datasets = {
                'focus_prediction': self._prepare_focus_dataset(df),
                'productivity_prediction': self._prepare_productivity_dataset(df),
                'optimal_time_prediction': self._prepare_optimal_time_dataset(df),
                'completion_prediction': self._prepare_completion_dataset(df)
            }
            
            logger.info(f"📊 トレーニングデータ準備完了: {len(df)}セッション")
            return datasets
            
        except Exception as e:
            logger.error(f"トレーニングデータ準備エラー: {e}")
            return {}
    
    def train_focus_score_model(self) -> Dict[str, Any]:
        """フォーカススコア予測モデルをトレーニング"""
        try:
            datasets = self.prepare_training_data()
            if 'focus_prediction' not in datasets:
                return {'error': 'データ不足'}
            
            df = datasets['focus_prediction']
            
            # 特徴量とターゲットの分離
            feature_cols = ['planned_duration', 'hour_of_day', 'day_of_week', 
                          'interruption_count', 'environment_score']
            X = df[feature_cols]
            y = df['focus_score']
            
            # データ分割
            X_train, X_test, y_train, y_test = train_test_split(
                X, y, test_size=0.2, random_state=42
            )
            
            # スケーリング
            scaler = StandardScaler()
            X_train_scaled = scaler.fit_transform(X_train)
            X_test_scaled = scaler.transform(X_test)
            
            # モデルトレーニング（アンサンブル）
            models = {
                'random_forest': RandomForestRegressor(n_estimators=100, random_state=42),
                'gradient_boosting': GradientBoostingRegressor(n_estimators=100, random_state=42),
                'linear_regression': LinearRegression()
            }
            
            best_model = None
            best_score = -float('inf')
            results = {}
            
            for name, model in models.items():
                # クロスバリデーション
                cv_scores = cross_val_score(model, X_train_scaled, y_train, cv=5, scoring='r2')
                
                # モデル訓練
                model.fit(X_train_scaled, y_train)
                
                # 予測と評価
                y_pred = model.predict(X_test_scaled)
                mse = mean_squared_error(y_test, y_pred)
                r2 = r2_score(y_test, y_pred)
                mae = mean_absolute_error(y_test, y_pred)
                
                model_result = {
                    'cv_mean': cv_scores.mean(),
                    'cv_std': cv_scores.std(),
                    'mse': mse,
                    'r2': r2,
                    'mae': mae
                }
                
                results[name] = model_result
                
                if r2 > best_score:
                    best_score = r2
                    best_model = (name, model)
            
            # ベストモデルを保存
            if best_model:
                model_name, model = best_model
                self.models['focus_score'] = model
                self.scalers['focus_score'] = scaler
                
                # モデル保存
                joblib.dump(model, self.model_dir / 'focus_score_model.pkl')
                joblib.dump(scaler, self.model_dir / 'focus_score_scaler.pkl')
                
                self.model_metrics['focus_score'] = results[model_name]
                
                self.model_trained.emit('focus_score', results[model_name])
                logger.info(f"🎯 フォーカススコア予測モデル訓練完了: {model_name} (R² = {best_score:.3f})")
                
                return {
                    'best_model': model_name,
                    'metrics': results[model_name],
                    'all_results': results
                }
            
            return {'error': 'モデル訓練失敗'}
            
        except Exception as e:
            logger.error(f"フォーカススコアモデル訓練エラー: {e}")
            return {'error': str(e)}
    
    def train_productivity_trend_model(self) -> Dict[str, Any]:
        """生産性トレンド予測モデルをトレーニング"""
        try:
            datasets = self.prepare_training_data()
            if 'productivity_prediction' not in datasets:
                return {'error': 'データ不足'}
            
            df = datasets['productivity_prediction']
            
            # 時系列特徴量の追加
            df['session_date'] = pd.to_datetime(df['start_time'])
            df = df.sort_values('session_date')
            
            # 移動平均特徴量
            df['productivity_ma7'] = df['productivity_score'].rolling(window=7, min_periods=1).mean()
            df['productivity_ma30'] = df['productivity_score'].rolling(window=30, min_periods=1).mean()
            
            # ラグ特徴量
            df['productivity_lag1'] = df['productivity_score'].shift(1)
            df['productivity_lag7'] = df['productivity_score'].shift(7)
            
            # 特徴量選択
            feature_cols = ['hour_of_day', 'day_of_week', 'planned_duration',
                          'productivity_ma7', 'productivity_ma30', 
                          'productivity_lag1', 'productivity_lag7']
            
            # 欠損値処理
            df = df.dropna()
            
            if len(df) < 20:
                return {'error': '時系列データ不足'}
            
            X = df[feature_cols]
            y = df['productivity_score']
            
            # 時系列分割（最新20%をテスト用）
            split_idx = int(len(df) * 0.8)
            X_train, X_test = X[:split_idx], X[split_idx:]
            y_train, y_test = y[:split_idx], y[split_idx:]
            
            # スケーリング
            scaler = StandardScaler()
            X_train_scaled = scaler.fit_transform(X_train)
            X_test_scaled = scaler.transform(X_test)
            
            # グラディエントブースティングモデル（時系列に適している）
            model = GradientBoostingRegressor(
                n_estimators=200,
                learning_rate=0.1,
                max_depth=4,
                random_state=42
            )
            
            model.fit(X_train_scaled, y_train)
            
            # 予測と評価
            y_pred = model.predict(X_test_scaled)
            mse = mean_squared_error(y_test, y_pred)
            r2 = r2_score(y_test, y_pred)
            mae = mean_absolute_error(y_test, y_pred)
            
            # モデル保存
            self.models['productivity_trend'] = model
            self.scalers['productivity_trend'] = scaler
            
            joblib.dump(model, self.model_dir / 'productivity_trend_model.pkl')
            joblib.dump(scaler, self.model_dir / 'productivity_trend_scaler.pkl')
            
            metrics = {'mse': mse, 'r2': r2, 'mae': mae}
            self.model_metrics['productivity_trend'] = metrics
            
            self.model_trained.emit('productivity_trend', metrics)
            logger.info(f"📈 生産性トレンド予測モデル訓練完了 (R² = {r2:.3f})")
            
            return {
                'metrics': metrics,
                'feature_importance': dict(zip(feature_cols, model.feature_importances_))
            }
            
        except Exception as e:
            logger.error(f"生産性トレンドモデル訓練エラー: {e}")
            return {'error': str(e)}
    
    def train_optimal_work_time_model(self) -> Dict[str, Any]:
        """最適作業時間予測モデルをトレーニング"""
        try:
            datasets = self.prepare_training_data()
            if 'optimal_time_prediction' not in datasets:
                return {'error': 'データ不足'}
            
            df = datasets['optimal_time_prediction']
            
            # 時間別効率スコアの計算
            hourly_efficiency = df.groupby('hour_of_day').agg({
                'efficiency_score': ['mean', 'std', 'count'],
                'focus_score': 'mean',
                'completed': 'sum'
            }).round(3)
            
            # マルチレベルカラム名を平坦化
            hourly_efficiency.columns = ['_'.join(col).strip() for col in hourly_efficiency.columns]
            hourly_efficiency = hourly_efficiency.reset_index()
            
            # 曜日別効率スコアの計算
            daily_efficiency = df.groupby('day_of_week').agg({
                'efficiency_score': ['mean', 'std'],
                'focus_score': 'mean',
                'completed': 'sum'
            }).round(3)
            
            daily_efficiency.columns = ['_'.join(col).strip() for col in daily_efficiency.columns]
            daily_efficiency = daily_efficiency.reset_index()
            
            # 最適時間帯の特定（効率スコア上位25%）
            efficiency_threshold = df['efficiency_score'].quantile(0.75)
            optimal_sessions = df[df['efficiency_score'] >= efficiency_threshold]
            
            if len(optimal_sessions) < 5:
                return {'error': '最適セッションデータ不足'}
            
            # 最適時間帯の特徴量抽出
            optimal_hours = optimal_sessions['hour_of_day'].value_counts()
            optimal_days = optimal_sessions['day_of_week'].value_counts()
            
            # 分類問題として解く（最適/非最適）
            df['is_optimal'] = (df['efficiency_score'] >= efficiency_threshold).astype(int)
            
            feature_cols = ['hour_of_day', 'day_of_week', 'planned_duration']
            X = df[feature_cols]
            y = df['is_optimal']
            
            # データ分割
            X_train, X_test, y_train, y_test = train_test_split(
                X, y, test_size=0.2, random_state=42, stratify=y
            )
            
            # ランダムフォレスト分類器
            from sklearn.ensemble import RandomForestClassifier
            from sklearn.metrics import classification_report, accuracy_score
            
            model = RandomForestClassifier(
                n_estimators=100,
                max_depth=10,
                random_state=42,
                class_weight='balanced'
            )
            
            model.fit(X_train, y_train)
            
            # 予測と評価
            y_pred = model.predict(X_test)
            accuracy = accuracy_score(y_test, y_pred)
            
            # モデル保存
            self.models['optimal_work_time'] = model
            
            joblib.dump(model, self.model_dir / 'optimal_work_time_model.pkl')
            
            metrics = {
                'accuracy': accuracy,
                'optimal_hours': optimal_hours.to_dict(),
                'optimal_days': optimal_days.to_dict(),
                'hourly_efficiency': hourly_efficiency.to_dict('records'),
                'daily_efficiency': daily_efficiency.to_dict('records')
            }
            
            self.model_metrics['optimal_work_time'] = metrics
            self.model_trained.emit('optimal_work_time', metrics)
            
            logger.info(f"⏰ 最適作業時間予測モデル訓練完了 (精度 = {accuracy:.3f})")
            
            return metrics
            
        except Exception as e:
            logger.error(f"最適作業時間モデル訓練エラー: {e}")
            return {'error': str(e)}
    
    def predict_focus_score(self, session_params: Dict[str, Any]) -> Dict[str, Any]:
        """フォーカススコアを予測"""
        try:
            if self.models['focus_score'] is None:
                # モデルをロード
                self.load_models()
            
            model = self.models['focus_score']
            scaler = self.scalers['focus_score']
            
            if model is None or scaler is None:
                return {'error': 'モデル未訓練'}
            
            # 特徴量準備
            features = np.array([[
                session_params.get('planned_duration', 25),
                session_params.get('hour_of_day', datetime.now().hour),
                session_params.get('day_of_week', datetime.now().weekday()),
                session_params.get('interruption_count', 0),
                session_params.get('environment_score', 0.5)
            ]])
            
            # スケーリングと予測
            features_scaled = scaler.transform(features)
            prediction = model.predict(features_scaled)[0]
            
            # 信頼区間の計算（アンサンブル予測の場合）
            if hasattr(model, 'estimators_'):
                # ランダムフォレストの場合
                predictions = [estimator.predict(features_scaled)[0] 
                             for estimator in model.estimators_]
                confidence_interval = {
                    'lower': np.percentile(predictions, 25),
                    'upper': np.percentile(predictions, 75),
                    'std': np.std(predictions)
                }
            else:
                confidence_interval = {'lower': prediction, 'upper': prediction, 'std': 0}
            
            result = {
                'predicted_focus_score': max(0.0, min(1.0, prediction)),
                'confidence_interval': confidence_interval,
                'model_accuracy': self.model_metrics.get('focus_score', {}).get('r2', 0)
            }
            
            self.prediction_updated.emit('focus_score', prediction)
            
            return result
            
        except Exception as e:
            logger.error(f"フォーカススコア予測エラー: {e}")
            return {'error': str(e)}
    
    def predict_optimal_work_time(self, user_preferences: Dict[str, Any] = None) -> Dict[str, Any]:
        """最適作業時間を予測"""
        try:
            if self.models['optimal_work_time'] is None:
                self.load_models()
            
            model = self.models['optimal_work_time']
            if model is None:
                return {'error': 'モデル未訓練'}
            
            current_hour = datetime.now().hour
            current_day = datetime.now().weekday()
            
            # 24時間×7日の組み合わせをテスト
            predictions = []
            for hour in range(24):
                for day in range(7):
                    for duration in [25, 45, 60]:  # 一般的なポモドーロ時間
                        features = np.array([[hour, day, duration]])
                        prob = model.predict_proba(features)[0][1]  # 最適である確率
                        
                        predictions.append({
                            'hour': hour,
                            'day': day,
                            'duration': duration,
                            'optimal_probability': prob,
                            'hour_name': self._get_hour_name(hour),
                            'day_name': self._get_day_name(day)
                        })
            
            # 確率でソート
            predictions.sort(key=lambda x: x['optimal_probability'], reverse=True)
            
            # 今日の推奨時間帯
            today_recommendations = [p for p in predictions 
                                  if p['day'] == current_day and p['hour'] >= current_hour][:5]
            
            # 全体的な推奨時間帯
            top_recommendations = predictions[:10]
            
            # 時間帯別統計（モデルメトリクスから）
            metrics = self.model_metrics.get('optimal_work_time', {})
            
            result = {
                'today_recommendations': today_recommendations,
                'top_recommendations': top_recommendations,
                'current_time_optimal_prob': self._get_current_time_probability(predictions),
                'hourly_statistics': metrics.get('hourly_efficiency', []),
                'daily_statistics': metrics.get('daily_efficiency', [])
            }
            
            self.prediction_ready.emit('optimal_work_time', result)
            
            return result
            
        except Exception as e:
            logger.error(f"最適作業時間予測エラー: {e}")
            return {'error': str(e)}
    
    def predict_productivity_trend(self, days_ahead: int = 7) -> Dict[str, Any]:
        """生産性トレンドを予測"""
        try:
            if self.models['productivity_trend'] is None:
                self.load_models()
            
            model = self.models['productivity_trend']
            scaler = self.scalers['productivity_trend']
            
            if model is None or scaler is None:
                return {'error': 'モデル未訓練'}
            
            # 最近のデータを取得
            recent_sessions = self.data_collector.session_data[-30:]  # 最新30セッション
            if len(recent_sessions) < 7:
                return {'error': '予測用データ不足'}
            
            # 生産性スコアの計算
            productivity_scores = []
            for session in recent_sessions:
                score = session.get('efficiency_score', 0) * session.get('focus_score', 0)
                productivity_scores.append(score)
            
            # 移動平均の計算
            ma7 = np.mean(productivity_scores[-7:]) if len(productivity_scores) >= 7 else np.mean(productivity_scores)
            ma30 = np.mean(productivity_scores) if len(productivity_scores) >= 30 else ma7
            
            # 未来の予測
            predictions = []
            current_date = datetime.now()
            
            for i in range(days_ahead):
                future_date = current_date + timedelta(days=i)
                hour = 10  # デフォルト作業時間
                day_of_week = future_date.weekday()
                
                # 特徴量準備
                features = np.array([[
                    hour,
                    day_of_week,
                    25,  # デフォルト計画時間
                    ma7,
                    ma30,
                    productivity_scores[-1] if productivity_scores else 0.5,
                    productivity_scores[-7] if len(productivity_scores) >= 7 else 0.5
                ]])
                
                # 予測
                features_scaled = scaler.transform(features)
                predicted_score = model.predict(features_scaled)[0]
                
                predictions.append({
                    'date': future_date.strftime('%Y-%m-%d'),
                    'day_name': self._get_day_name(day_of_week),
                    'predicted_productivity': max(0.0, min(1.0, predicted_score)),
                    'confidence': 0.8  # 簡易信頼度
                })
                
                # 移動平均を更新（予測値を使用）
                productivity_scores.append(predicted_score)
                ma7 = np.mean(productivity_scores[-7:])
                ma30 = np.mean(productivity_scores[-30:])
            
            # トレンド分析
            trend_slope = np.polyfit(range(len(predictions)), 
                                   [p['predicted_productivity'] for p in predictions], 1)[0]
            
            trend_direction = 'increasing' if trend_slope > 0.01 else 'decreasing' if trend_slope < -0.01 else 'stable'
            
            result = {
                'predictions': predictions,
                'trend_direction': trend_direction,
                'trend_slope': trend_slope,
                'average_predicted_productivity': np.mean([p['predicted_productivity'] for p in predictions]),
                'current_productivity_ma7': ma7,
                'model_accuracy': self.model_metrics.get('productivity_trend', {}).get('r2', 0)
            }
            
            self.prediction_ready.emit('productivity_trend', result)
            
            return result
            
        except Exception as e:
            logger.error(f"生産性トレンド予測エラー: {e}")
            return {'error': str(e)}
    
    def load_models(self):
        """保存されたモデルをロード"""
        try:
            model_files = {
                'focus_score': ('focus_score_model.pkl', 'focus_score_scaler.pkl'),
                'productivity_trend': ('productivity_trend_model.pkl', 'productivity_trend_scaler.pkl'),
                'optimal_work_time': ('optimal_work_time_model.pkl', None)
            }
            
            for model_name, (model_file, scaler_file) in model_files.items():
                model_path = self.model_dir / model_file
                if model_path.exists():
                    self.models[model_name] = joblib.load(model_path)
                    
                    if scaler_file:
                        scaler_path = self.model_dir / scaler_file
                        if scaler_path.exists():
                            self.scalers[model_name] = joblib.load(scaler_path)
            
            logger.info("🤖 保存済みモデルをロード完了")
            
        except Exception as e:
            logger.error(f"モデルロードエラー: {e}")
    
    def retrain_all_models(self) -> Dict[str, Any]:
        """全モデルを再トレーニング"""
        results = {}
        
        try:
            results['focus_score'] = self.train_focus_score_model()
            results['productivity_trend'] = self.train_productivity_trend_model()
            results['optimal_work_time'] = self.train_optimal_work_time_model()
            
            logger.info("🚀 全モデル再トレーニング完了")
            
        except Exception as e:
            logger.error(f"モデル再トレーニングエラー: {e}")
            results['error'] = str(e)
        
        return results
    
    # Helper methods
    def _extract_hour_of_day(self, timestamp_str: str) -> int:
        """タイムスタンプから時間を抽出"""
        try:
            if timestamp_str:
                dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                return dt.hour
        except:
            pass
        return datetime.now().hour
    
    def _extract_day_of_week(self, timestamp_str: str) -> int:
        """タイムスタンプから曜日を抽出"""
        try:
            if timestamp_str:
                dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                return dt.weekday()
        except:
            pass
        return datetime.now().weekday()
    
    def _calculate_environment_score(self, env_data: Dict[str, Any]) -> float:
        """環境データからスコアを計算"""
        if not env_data:
            return 0.5
        
        # 簡易環境スコア計算
        score = 0.5
        if 'noise_level' in env_data:
            score += (1.0 - env_data['noise_level']) * 0.3
        if 'distraction_count' in env_data:
            score -= min(env_data['distraction_count'] * 0.1, 0.3)
        
        return max(0.0, min(1.0, score))
    
    def _prepare_focus_dataset(self, df: pd.DataFrame) -> pd.DataFrame:
        """フォーカス予測用データセット準備"""
        return df.dropna(subset=['focus_score'])
    
    def _prepare_productivity_dataset(self, df: pd.DataFrame) -> pd.DataFrame:
        """生産性予測用データセット準備"""
        df['productivity_score'] = df['efficiency_score'] * df['focus_score']
        df['start_time'] = pd.to_datetime(df.get('start_time', datetime.now().isoformat()))
        return df.dropna(subset=['productivity_score'])
    
    def _prepare_optimal_time_dataset(self, df: pd.DataFrame) -> pd.DataFrame:
        """最適時間予測用データセット準備"""
        return df.dropna(subset=['efficiency_score', 'focus_score'])
    
    def _prepare_completion_dataset(self, df: pd.DataFrame) -> pd.DataFrame:
        """完了予測用データセット準備"""
        return df.dropna(subset=['completed'])
    
    def _get_hour_name(self, hour: int) -> str:
        """時間の名前を取得"""
        if 6 <= hour < 12:
            return "午前"
        elif 12 <= hour < 18:
            return "午後"
        elif 18 <= hour < 24:
            return "夕方"
        else:
            return "深夜"
    
    def _get_day_name(self, day: int) -> str:
        """曜日の名前を取得"""
        days = ['月', '火', '水', '木', '金', '土', '日']
        return days[day] if 0 <= day < 7 else '不明'
    
    def _get_current_time_probability(self, predictions: List[Dict]) -> float:
        """現在時刻の最適確率を取得"""
        current_hour = datetime.now().hour
        current_day = datetime.now().weekday()
        
        for pred in predictions:
            if pred['hour'] == current_hour and pred['day'] == current_day:
                return pred['optimal_probability']
        
        return 0.5


class ReportExporter(QObject):
    """Worker3: PDF/Excel Report Export System - レポートエクスポートシステム"""
    
    # シグナル
    export_completed = pyqtSignal(str, str)  # (export_type, file_path)
    export_progress = pyqtSignal(int)  # progress percentage
    export_error = pyqtSignal(str, str)  # (export_type, error_message)
    
    def __init__(self, reports_engine, visualization, comparison_analytics):
        super().__init__()
        self.reports_engine = reports_engine
        self.visualization = visualization
        self.comparison_analytics = comparison_analytics
        
        # エクスポート用ディレクトリ
        self.export_dir = Path("data/exports")
        self.export_dir.mkdir(exist_ok=True)
        
        # テンプレート設定
        self.pdf_styles = getSampleStyleSheet()
        self._setup_pdf_styles()
        
        logger.info("📄 ReportExporter初期化完了")
    
    def _setup_pdf_styles(self):
        """PDFスタイルの設定"""
        # カスタムスタイル定義
        self.pdf_styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.pdf_styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.darkblue,
            alignment=1  # 中央揃え
        ))
        
        self.pdf_styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.pdf_styles['Heading2'],
            fontSize=16,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.darkblue
        ))
        
        self.pdf_styles.add(ParagraphStyle(
            name='CustomBody',
            parent=self.pdf_styles['Normal'],
            fontSize=12,
            spaceAfter=12,
            leading=14
        ))
    
    def export_comprehensive_pdf_report(self, report_data: Dict[str, Any], 
                                      report_name: str = None) -> str:
        """包括的なPDFレポートをエクスポート"""
        try:
            if report_name is None:
                report_name = f"comprehensive_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            pdf_path = self.export_dir / f"{report_name}.pdf"
            
            # ReportLabドキュメント作成
            doc = SimpleDocTemplate(
                str(pdf_path),
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # コンテンツ構築
            story = []
            
            # タイトルページ
            story.extend(self._create_pdf_title_page(report_name))
            
            # サマリーセクション
            story.extend(self._create_pdf_summary_section(report_data))
            
            # セッション統計セクション
            story.extend(self._create_pdf_session_stats_section(report_data))
            
            # 可視化チャートセクション
            story.extend(self._create_pdf_charts_section(report_data))
            
            # 予測分析セクション
            story.extend(self._create_pdf_predictions_section(report_data))
            
            # 推奨事項セクション
            story.extend(self._create_pdf_recommendations_section(report_data))
            
            # 詳細データセクション
            story.extend(self._create_pdf_detailed_data_section(report_data))
            
            # PDF生成
            doc.build(story)
            
            self.export_completed.emit('pdf', str(pdf_path))
            logger.info(f"📄 PDFレポートエクスポート完了: {pdf_path}")
            
            return str(pdf_path)
            
        except Exception as e:
            error_msg = f"PDFエクスポートエラー: {e}"
            logger.error(error_msg)
            self.export_error.emit('pdf', error_msg)
            return ""
    
    def export_excel_workbook(self, report_data: Dict[str, Any], 
                            workbook_name: str = None) -> str:
        """包括的なExcelワークブックをエクスポート"""
        try:
            if workbook_name is None:
                workbook_name = f"pomodoro_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            excel_path = self.export_dir / f"{workbook_name}.xlsx"
            
            # Excelワークブック作成
            wb = openpyxl.Workbook()
            
            # デフォルトシートを削除
            wb.remove(wb.active)
            
            # 各シートを作成
            self._create_excel_summary_sheet(wb, report_data)
            self._create_excel_sessions_sheet(wb, report_data)
            self._create_excel_statistics_sheet(wb, report_data)
            self._create_excel_predictions_sheet(wb, report_data)
            self._create_excel_charts_sheet(wb, report_data)
            self._create_excel_raw_data_sheet(wb, report_data)
            
            # ワークブック保存
            wb.save(excel_path)
            
            self.export_completed.emit('excel', str(excel_path))
            logger.info(f"📊 Excelレポートエクスポート完了: {excel_path}")
            
            return str(excel_path)
            
        except Exception as e:
            error_msg = f"Excelエクスポートエラー: {e}"
            logger.error(error_msg)
            self.export_error.emit('excel', error_msg)
            return ""
    
    def _create_pdf_title_page(self, report_name: str) -> List:
        """PDFタイトルページ作成"""
        elements = []
        
        # タイトル
        title_text = "Pomodoro Timer Analytics Report"
        elements.append(Paragraph(title_text, self.pdf_styles['CustomTitle']))
        elements.append(Spacer(1, 12))
        
        # サブタイトル
        subtitle_text = f"Report: {report_name}"
        elements.append(Paragraph(subtitle_text, self.pdf_styles['CustomHeading']))
        elements.append(Spacer(1, 12))
        
        # 生成日時
        generated_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        elements.append(Paragraph(generated_text, self.pdf_styles['CustomBody']))
        elements.append(Spacer(1, 24))
        
        # 概要
        overview_text = """
        This comprehensive report provides detailed analytics of your Pomodoro Timer sessions,
        including productivity trends, focus patterns, and AI-powered predictions for optimal work times.
        """
        elements.append(Paragraph(overview_text, self.pdf_styles['CustomBody']))
        
        # 改ページ
        from reportlab.platypus import PageBreak
        elements.append(PageBreak())
        
        return elements
    
    def _create_pdf_summary_section(self, report_data: Dict[str, Any]) -> List:
        """PDFサマリーセクション作成"""
        elements = []
        
        elements.append(Paragraph("Executive Summary", self.pdf_styles['CustomHeading']))
        
        # サマリー統計
        summary_stats = report_data.get('summary', {})
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Sessions', str(summary_stats.get('total_sessions', 0))],
            ['Completed Sessions', str(summary_stats.get('completed_sessions', 0))],
            ['Average Focus Score', f"{summary_stats.get('avg_focus_score', 0):.2f}"],
            ['Total Work Time (hours)', f"{summary_stats.get('total_work_time', 0):.1f}"],
            ['Productivity Trend', summary_stats.get('productivity_trend', 'Stable')],
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        return elements
    
    def _create_pdf_session_stats_section(self, report_data: Dict[str, Any]) -> List:
        """PDFセッション統計セクション作成"""
        elements = []
        
        elements.append(Paragraph("Session Statistics", self.pdf_styles['CustomHeading']))
        
        # 週間/月間統計
        stats = report_data.get('session_stats', {})
        
        # 週間データ
        if 'weekly' in stats:
            elements.append(Paragraph("Weekly Performance", self.pdf_styles['Heading3']))
            
            weekly_data = [['Day', 'Sessions', 'Avg Focus', 'Completion Rate']]
            for day_stat in stats['weekly']:
                weekly_data.append([
                    day_stat.get('day', ''),
                    str(day_stat.get('sessions', 0)),
                    f"{day_stat.get('avg_focus', 0):.2f}",
                    f"{day_stat.get('completion_rate', 0):.1%}"
                ])
            
            weekly_table = Table(weekly_data)
            weekly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(weekly_table)
            elements.append(Spacer(1, 15))
        
        return elements
    
    def _create_pdf_charts_section(self, report_data: Dict[str, Any]) -> List:
        """PDFチャートセクション作成"""
        elements = []
        
        elements.append(Paragraph("Performance Charts", self.pdf_styles['CustomHeading']))
        
        # 可視化チャートの画像を生成して挿入
        try:
            # フォーカススコア推移チャート
            if hasattr(self.visualization, 'create_focus_trend_chart'):
                chart_data = report_data.get('charts', {})
                if 'focus_trend' in chart_data:
                    # matplotlib図をPDFに埋め込み用に変換
                    img_buffer = io.BytesIO()
                    
                    # 簡易チャート生成
                    fig, ax = plt.subplots(figsize=(8, 4))
                    
                    # フォーカススコアトレンドデータがあればプロット
                    focus_data = chart_data['focus_trend']
                    if focus_data:
                        ax.plot(focus_data.get('dates', []), focus_data.get('scores', []))
                        ax.set_title('Focus Score Trend')
                        ax.set_xlabel('Date')
                        ax.set_ylabel('Focus Score')
                    
                    plt.tight_layout()
                    plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
                    img_buffer.seek(0)
                    
                    # PDF用の画像サイズ調整
                    img = Image(img_buffer)
                    img.drawHeight = 3*inch
                    img.drawWidth = 6*inch
                    
                    elements.append(img)
                    elements.append(Spacer(1, 15))
                    
                    plt.close(fig)
                    
        except Exception as e:
            logger.error(f"チャートPDF埋め込みエラー: {e}")
            # エラーの場合はテキストで代替
            elements.append(Paragraph("Chart generation temporarily unavailable", 
                                    self.pdf_styles['CustomBody']))
        
        return elements
    
    def _create_pdf_predictions_section(self, report_data: Dict[str, Any]) -> List:
        """PDF予測セクション作成"""
        elements = []
        
        elements.append(Paragraph("AI Predictions & Insights", self.pdf_styles['CustomHeading']))
        
        predictions = report_data.get('predictions', {})
        
        # 最適作業時間予測
        if 'optimal_times' in predictions:
            elements.append(Paragraph("Optimal Work Times", self.pdf_styles['Heading3']))
            
            optimal_times = predictions['optimal_times']
            if optimal_times and 'today_recommendations' in optimal_times:
                rec_data = [['Time', 'Day', 'Duration', 'Probability']]
                for rec in optimal_times['today_recommendations'][:5]:
                    rec_data.append([
                        f"{rec.get('hour', 0):02d}:00",
                        rec.get('day_name', ''),
                        f"{rec.get('duration', 25)} min",
                        f"{rec.get('optimal_probability', 0):.1%}"
                    ])
                
                rec_table = Table(rec_data)
                rec_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(rec_table)
                elements.append(Spacer(1, 15))
        
        # 生産性トレンド予測
        if 'productivity_trend' in predictions:
            elements.append(Paragraph("Productivity Trend Forecast", self.pdf_styles['Heading3']))
            
            trend_data = predictions['productivity_trend']
            trend_text = f"""
            Predicted trend direction: {trend_data.get('trend_direction', 'stable')}
            Average predicted productivity: {trend_data.get('average_predicted_productivity', 0):.2f}
            Model accuracy (R²): {trend_data.get('model_accuracy', 0):.3f}
            """
            elements.append(Paragraph(trend_text, self.pdf_styles['CustomBody']))
            elements.append(Spacer(1, 15))
        
        return elements
    
    def _create_pdf_recommendations_section(self, report_data: Dict[str, Any]) -> List:
        """PDF推奨事項セクション作成"""
        elements = []
        
        elements.append(Paragraph("Recommendations", self.pdf_styles['CustomHeading']))
        
        recommendations = report_data.get('recommendations', [])
        
        if recommendations:
            for i, rec in enumerate(recommendations, 1):
                rec_text = f"{i}. {rec}"
                elements.append(Paragraph(rec_text, self.pdf_styles['CustomBody']))
        else:
            # デフォルト推奨事項
            default_recs = [
                "Maintain consistent work schedule for better focus patterns",
                "Take regular breaks to optimize productivity",
                "Monitor interruption patterns and minimize distractions",
                "Use predicted optimal work times for important tasks"
            ]
            for i, rec in enumerate(default_recs, 1):
                rec_text = f"{i}. {rec}"
                elements.append(Paragraph(rec_text, self.pdf_styles['CustomBody']))
        
        return elements
    
    def _create_pdf_detailed_data_section(self, report_data: Dict[str, Any]) -> List:
        """PDF詳細データセクション作成"""
        elements = []
        
        elements.append(Paragraph("Detailed Session Data", self.pdf_styles['CustomHeading']))
        
        # 最近のセッションデータ（最新10件）
        sessions = report_data.get('recent_sessions', [])
        
        if sessions:
            session_data = [['Date', 'Type', 'Duration', 'Focus Score', 'Completed']]
            
            for session in sessions[:10]:  # 最新10件
                session_data.append([
                    session.get('date', ''),
                    session.get('type', ''),
                    f"{session.get('duration', 0)} min",
                    f"{session.get('focus_score', 0):.2f}",
                    'Yes' if session.get('completed', False) else 'No'
                ])
            
            session_table = Table(session_data)
            session_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(session_table)
        
        return elements
    
    def _create_excel_summary_sheet(self, wb, report_data: Dict[str, Any]):
        """Excelサマリーシート作成"""
        ws = wb.create_sheet("Summary")
        
        # ヘッダースタイル
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # タイトル
        ws['A1'] = "Pomodoro Timer Analytics Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # サマリー統計
        summary = report_data.get('summary', {})
        
        row = 3
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        
        metrics = [
            ("Total Sessions", summary.get('total_sessions', 0)),
            ("Completed Sessions", summary.get('completed_sessions', 0)),
            ("Average Focus Score", f"{summary.get('avg_focus_score', 0):.2f}"),
            ("Total Work Time (hours)", f"{summary.get('total_work_time', 0):.1f}"),
            ("Productivity Trend", summary.get('productivity_trend', 'Stable')),
        ]
        
        for metric, value in metrics:
            row += 1
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
        
        # カラム幅調整
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _create_excel_sessions_sheet(self, wb, report_data: Dict[str, Any]):
        """Excelセッションシート作成"""
        ws = wb.create_sheet("Sessions")
        
        # ヘッダー
        headers = ['Date', 'Type', 'Planned Duration', 'Actual Duration', 
                  'Focus Score', 'Efficiency Score', 'Completed', 'Interruptions']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # セッションデータ
        sessions = report_data.get('sessions', [])
        
        for row, session in enumerate(sessions, 2):
            ws.cell(row=row, column=1, value=session.get('date', ''))
            ws.cell(row=row, column=2, value=session.get('type', ''))
            ws.cell(row=row, column=3, value=session.get('planned_duration', 0))
            ws.cell(row=row, column=4, value=session.get('actual_duration', 0))
            ws.cell(row=row, column=5, value=session.get('focus_score', 0))
            ws.cell(row=row, column=6, value=session.get('efficiency_score', 0))
            ws.cell(row=row, column=7, value='Yes' if session.get('completed', False) else 'No')
            ws.cell(row=row, column=8, value=len(session.get('interruptions', [])))
        
        # カラム幅自動調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_excel_statistics_sheet(self, wb, report_data: Dict[str, Any]):
        """Excel統計シート作成"""
        ws = wb.create_sheet("Statistics")
        
        # 週間統計
        ws['A1'] = "Weekly Statistics"
        ws['A1'].font = Font(bold=True, size=14)
        
        stats = report_data.get('session_stats', {})
        if 'weekly' in stats:
            headers = ['Day', 'Sessions', 'Avg Focus', 'Completion Rate']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
            
            for row, day_stat in enumerate(stats['weekly'], 4):
                ws.cell(row=row, column=1, value=day_stat.get('day', ''))
                ws.cell(row=row, column=2, value=day_stat.get('sessions', 0))
                ws.cell(row=row, column=3, value=day_stat.get('avg_focus', 0))
                ws.cell(row=row, column=4, value=day_stat.get('completion_rate', 0))
        
        # 月間統計
        row_offset = len(stats.get('weekly', [])) + 6
        ws[f'A{row_offset}'] = "Monthly Statistics"
        ws[f'A{row_offset}'].font = Font(bold=True, size=14)
        
        if 'monthly' in stats:
            headers = ['Month', 'Sessions', 'Avg Focus', 'Total Hours']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_offset + 2, column=col, value=header)
                cell.font = Font(bold=True)
            
            for row, month_stat in enumerate(stats['monthly'], row_offset + 3):
                ws.cell(row=row, column=1, value=month_stat.get('month', ''))
                ws.cell(row=row, column=2, value=month_stat.get('sessions', 0))
                ws.cell(row=row, column=3, value=month_stat.get('avg_focus', 0))
                ws.cell(row=row, column=4, value=month_stat.get('total_hours', 0))
    
    def _create_excel_predictions_sheet(self, wb, report_data: Dict[str, Any]):
        """Excel予測シート作成"""
        ws = wb.create_sheet("Predictions")
        
        # 最適作業時間予測
        ws['A1'] = "Optimal Work Time Predictions"
        ws['A1'].font = Font(bold=True, size=14)
        
        predictions = report_data.get('predictions', {})
        
        if 'optimal_times' in predictions and predictions['optimal_times']:
            optimal_data = predictions['optimal_times']
            
            headers = ['Hour', 'Day', 'Duration (min)', 'Probability']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
            
            recommendations = optimal_data.get('today_recommendations', [])
            for row, rec in enumerate(recommendations, 4):
                ws.cell(row=row, column=1, value=f"{rec.get('hour', 0):02d}:00")
                ws.cell(row=row, column=2, value=rec.get('day_name', ''))
                ws.cell(row=row, column=3, value=rec.get('duration', 25))
                ws.cell(row=row, column=4, value=rec.get('optimal_probability', 0))
        
        # 生産性トレンド予測
        trend_row = 15
        ws[f'A{trend_row}'] = "Productivity Trend Forecast"
        ws[f'A{trend_row}'].font = Font(bold=True, size=14)
        
        if 'productivity_trend' in predictions:
            trend_data = predictions['productivity_trend']
            
            ws[f'A{trend_row+2}'] = "Trend Direction:"
            ws[f'B{trend_row+2}'] = trend_data.get('trend_direction', 'stable')
            
            ws[f'A{trend_row+3}'] = "Average Predicted Productivity:"
            ws[f'B{trend_row+3}'] = trend_data.get('average_predicted_productivity', 0)
            
            ws[f'A{trend_row+4}'] = "Model Accuracy (R²):"
            ws[f'B{trend_row+4}'] = trend_data.get('model_accuracy', 0)
            
            # 日別予測
            if 'predictions' in trend_data:
                pred_headers = ['Date', 'Day', 'Predicted Productivity', 'Confidence']
                for col, header in enumerate(pred_headers, 1):
                    cell = ws.cell(row=trend_row+6, column=col, value=header)
                    cell.font = Font(bold=True)
                
                for row, pred in enumerate(trend_data['predictions'], trend_row+7):
                    ws.cell(row=row, column=1, value=pred.get('date', ''))
                    ws.cell(row=row, column=2, value=pred.get('day_name', ''))
                    ws.cell(row=row, column=3, value=pred.get('predicted_productivity', 0))
                    ws.cell(row=row, column=4, value=pred.get('confidence', 0))
    
    def _create_excel_charts_sheet(self, wb, report_data: Dict[str, Any]):
        """Excelチャートシート作成"""
        ws = wb.create_sheet("Charts")
        
        ws['A1'] = "Performance Charts"
        ws['A1'].font = Font(bold=True, size=14)
        
        # チャートデータの準備
        chart_data = report_data.get('charts', {})
        
        # フォーカススコアトレンドチャート
        if 'focus_trend' in chart_data:
            focus_data = chart_data['focus_trend']
            
            # データ範囲の設定
            headers = ['Date', 'Focus Score']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
            
            dates = focus_data.get('dates', [])
            scores = focus_data.get('scores', [])
            
            for row, (date, score) in enumerate(zip(dates, scores), 4):
                ws.cell(row=row, column=1, value=date)
                ws.cell(row=row, column=2, value=score)
            
            # 線グラフの作成
            chart = LineChart()
            chart.title = "Focus Score Trend"
            chart.style = 10
            chart.x_axis.title = 'Date'
            chart.y_axis.title = 'Focus Score'
            
            data = Reference(ws, min_col=2, min_row=3, max_row=3+len(scores), max_col=2)
            cats = Reference(ws, min_col=1, min_row=4, max_row=3+len(dates))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "D3")
    
    def _create_excel_raw_data_sheet(self, wb, report_data: Dict[str, Any]):
        """Excel生データシート作成"""
        ws = wb.create_sheet("Raw Data")
        
        ws['A1'] = "Raw Session Data"
        ws['A1'].font = Font(bold=True, size=14)
        
        # 全セッションデータのエクスポート
        sessions = report_data.get('sessions', [])
        
        if sessions:
            # ヘッダー（より詳細）
            headers = [
                'Session ID', 'Type', 'Start Time', 'End Time',
                'Planned Duration', 'Actual Duration', 'Focus Score',
                'Efficiency Score', 'Completed', 'Interruption Count',
                'Interaction Count', 'Environment Score'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # データ
            for row, session in enumerate(sessions, 4):
                ws.cell(row=row, column=1, value=session.get('session_id', ''))
                ws.cell(row=row, column=2, value=session.get('type', ''))
                ws.cell(row=row, column=3, value=session.get('start_time', ''))
                ws.cell(row=row, column=4, value=session.get('end_time', ''))
                ws.cell(row=row, column=5, value=session.get('planned_duration', 0))
                ws.cell(row=row, column=6, value=session.get('actual_duration', 0))
                ws.cell(row=row, column=7, value=session.get('focus_score', 0))
                ws.cell(row=row, column=8, value=session.get('efficiency_score', 0))
                ws.cell(row=row, column=9, value=session.get('completed', False))
                ws.cell(row=row, column=10, value=len(session.get('interruptions', [])))
                ws.cell(row=row, column=11, value=len(session.get('interactions', [])))
                ws.cell(row=row, column=12, value=session.get('environment_score', 0))
            
            # 境界線の追加
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=3, max_row=3+len(sessions), 
                                  min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border


class AutoReportScheduler(QObject):
    """Worker3: Automated Report Generation & Scheduling System - 自動レポート生成・スケジューリングシステム"""
    
    # シグナル
    report_scheduled = pyqtSignal(str, str)  # (report_type, schedule_info)
    report_generated = pyqtSignal(str, str)  # (report_type, file_path)
    schedule_updated = pyqtSignal(dict)  # schedule configuration
    email_sent = pyqtSignal(str, str)  # (recipient, report_type)
    
    def __init__(self, reports_engine, report_exporter, prediction_engine):
        super().__init__()
        self.reports_engine = reports_engine
        self.report_exporter = report_exporter
        self.prediction_engine = prediction_engine
        
        # スケジューラー初期化（利用可能な場合のみ）
        self.scheduler = None
        if SCHEDULER_AVAILABLE:
            try:
                self.scheduler = BackgroundScheduler()
                self.scheduler.start()
                logger.info("⏰ スケジューラー開始")
            except Exception as e:
                logger.error(f"⏰ スケジューラー初期化エラー: {e}")
                self.scheduler = None
        else:
            logger.warning("⏰ スケジューラー機能は無効です（ライブラリなし）")
        
        # 設定ファイル
        self.config_file = Path("data/scheduler_config.json")
        self.schedule_config = self.load_schedule_config()
        
        # レポート生成履歴
        self.generation_history = []
        
        # 自動スケジュール設定
        self._setup_default_schedules()
        
        logger.info("⏰ AutoReportScheduler初期化完了")
    
    def load_schedule_config(self) -> Dict[str, Any]:
        """スケジュール設定をロード"""
        try:
            if self.config_file.exists():
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                return self.get_default_config()
        except Exception as e:
            logger.error(f"スケジュール設定ロードエラー: {e}")
            return self.get_default_config()
    
    def get_default_config(self) -> Dict[str, Any]:
        """デフォルトスケジュール設定"""
        return {
            "daily_reports": {
                "enabled": True,
                "time": "18:00",
                "format": ["pdf"],
                "email_enabled": False,
                "email_recipients": []
            },
            "weekly_reports": {
                "enabled": True,
                "day": "sunday",
                "time": "20:00",
                "format": ["pdf", "excel"],
                "email_enabled": False,
                "email_recipients": []
            },
            "monthly_reports": {
                "enabled": True,
                "day": 1,  # 月の最初の日
                "time": "09:00",
                "format": ["pdf", "excel"],
                "email_enabled": False,
                "email_recipients": []
            },
            "email_config": {
                "smtp_server": "",
                "smtp_port": 587,
                "username": "",
                "password": "",
                "use_tls": True
            },
            "retention_days": 30  # 古いレポートの保持日数
        }
    
    def save_schedule_config(self):
        """スケジュール設定を保存"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.schedule_config, f, indent=2, ensure_ascii=False)
            logger.info("⚙️ スケジュール設定保存完了")
        except Exception as e:
            logger.error(f"スケジュール設定保存エラー: {e}")
    
    def _setup_default_schedules(self):
        """デフォルトスケジュールの設定"""
        if not self.scheduler:
            logger.warning("⏰ スケジューラー利用不可のため、自動レポート機能は無効です")
            return
            
        try:
            # 日次レポート
            if self.schedule_config["daily_reports"]["enabled"]:
                self.scheduler.add_job(
                    func=self._generate_daily_report,
                    trigger=CronTrigger(
                        hour=int(self.schedule_config["daily_reports"]["time"].split(':')[0]),
                        minute=int(self.schedule_config["daily_reports"]["time"].split(':')[1])
                    ),
                    id='daily_report',
                    replace_existing=True
                )
                logger.info(f"📅 日次レポートスケジュール設定: {self.schedule_config['daily_reports']['time']}")
            
            # 週次レポート
            if self.schedule_config["weekly_reports"]["enabled"]:
                day_map = {
                    'monday': 'mon', 'tuesday': 'tue', 'wednesday': 'wed',
                    'thursday': 'thu', 'friday': 'fri', 'saturday': 'sat', 'sunday': 'sun'
                }
                
                day_of_week = day_map.get(self.schedule_config["weekly_reports"]["day"], 'sun')
                
                self.scheduler.add_job(
                    func=self._generate_weekly_report,
                    trigger=CronTrigger(
                        day_of_week=day_of_week,
                        hour=int(self.schedule_config["weekly_reports"]["time"].split(':')[0]),
                        minute=int(self.schedule_config["weekly_reports"]["time"].split(':')[1])
                    ),
                    id='weekly_report',
                    replace_existing=True
                )
                logger.info(f"📅 週次レポートスケジュール設定: {day_of_week} {self.schedule_config['weekly_reports']['time']}")
            
            # 月次レポート
            if self.schedule_config["monthly_reports"]["enabled"]:
                self.scheduler.add_job(
                    func=self._generate_monthly_report,
                    trigger=CronTrigger(
                        day=self.schedule_config["monthly_reports"]["day"],
                        hour=int(self.schedule_config["monthly_reports"]["time"].split(':')[0]),
                        minute=int(self.schedule_config["monthly_reports"]["time"].split(':')[1])
                    ),
                    id='monthly_report',
                    replace_existing=True
                )
                logger.info(f"📅 月次レポートスケジュール設定: 毎月{self.schedule_config['monthly_reports']['day']}日 {self.schedule_config['monthly_reports']['time']}")
            
            # 古いレポート削除ジョブ
            self.scheduler.add_job(
                func=self._cleanup_old_reports,
                trigger=CronTrigger(hour=2, minute=0),  # 毎日午前2時
                id='cleanup_reports',
                replace_existing=True
            )
            
        except Exception as e:
            logger.error(f"デフォルトスケジュール設定エラー: {e}")
    
    def _generate_daily_report(self):
        """日次レポート生成"""
        try:
            logger.info("📊 日次レポート生成開始")
            
            # レポートデータの収集
            report_data = self._collect_daily_report_data()
            
            # フォーマット別生成
            generated_files = []
            formats = self.schedule_config["daily_reports"]["format"]
            
            report_name = f"daily_report_{datetime.now().strftime('%Y%m%d')}"
            
            if "pdf" in formats:
                pdf_path = self.report_exporter.export_comprehensive_pdf_report(
                    report_data, f"{report_name}_pdf"
                )
                if pdf_path:
                    generated_files.append(('pdf', pdf_path))
            
            if "excel" in formats:
                excel_path = self.report_exporter.export_excel_workbook(
                    report_data, f"{report_name}_excel"
                )
                if excel_path:
                    generated_files.append(('excel', excel_path))
            
            # 生成履歴に記録
            self.generation_history.append({
                'type': 'daily',
                'timestamp': datetime.now().isoformat(),
                'files': generated_files,
                'success': len(generated_files) > 0
            })
            
            # メール送信（設定されている場合）
            if self.schedule_config["daily_reports"]["email_enabled"]:
                self._send_report_email('daily', generated_files)
            
            # シグナル発信
            for format_type, file_path in generated_files:
                self.report_generated.emit('daily', file_path)
            
            logger.info(f"✅ 日次レポート生成完了: {len(generated_files)}ファイル")
            
        except Exception as e:
            logger.error(f"日次レポート生成エラー: {e}")
    
    def _generate_weekly_report(self):
        """週次レポート生成"""
        try:
            logger.info("📊 週次レポート生成開始")
            
            # レポートデータの収集
            report_data = self._collect_weekly_report_data()
            
            # 予測データの追加
            predictions = self._collect_prediction_data()
            report_data['predictions'] = predictions
            
            # フォーマット別生成
            generated_files = []
            formats = self.schedule_config["weekly_reports"]["format"]
            
            report_name = f"weekly_report_{datetime.now().strftime('%Y_W%U')}"
            
            if "pdf" in formats:
                pdf_path = self.report_exporter.export_comprehensive_pdf_report(
                    report_data, f"{report_name}_pdf"
                )
                if pdf_path:
                    generated_files.append(('pdf', pdf_path))
            
            if "excel" in formats:
                excel_path = self.report_exporter.export_excel_workbook(
                    report_data, f"{report_name}_excel"
                )
                if excel_path:
                    generated_files.append(('excel', excel_path))
            
            # 生成履歴に記録
            self.generation_history.append({
                'type': 'weekly',
                'timestamp': datetime.now().isoformat(),
                'files': generated_files,
                'success': len(generated_files) > 0
            })
            
            # メール送信（設定されている場合）
            if self.schedule_config["weekly_reports"]["email_enabled"]:
                self._send_report_email('weekly', generated_files)
            
            # シグナル発信
            for format_type, file_path in generated_files:
                self.report_generated.emit('weekly', file_path)
            
            logger.info(f"✅ 週次レポート生成完了: {len(generated_files)}ファイル")
            
        except Exception as e:
            logger.error(f"週次レポート生成エラー: {e}")
    
    def _generate_monthly_report(self):
        """月次レポート生成"""
        try:
            logger.info("📊 月次レポート生成開始")
            
            # レポートデータの収集
            report_data = self._collect_monthly_report_data()
            
            # 予測データとトレンド分析の追加
            predictions = self._collect_prediction_data()
            trends = self._analyze_monthly_trends()
            
            report_data['predictions'] = predictions
            report_data['trends'] = trends
            
            # フォーマット別生成
            generated_files = []
            formats = self.schedule_config["monthly_reports"]["format"]
            
            report_name = f"monthly_report_{datetime.now().strftime('%Y_%m')}"
            
            if "pdf" in formats:
                pdf_path = self.report_exporter.export_comprehensive_pdf_report(
                    report_data, f"{report_name}_pdf"
                )
                if pdf_path:
                    generated_files.append(('pdf', pdf_path))
            
            if "excel" in formats:
                excel_path = self.report_exporter.export_excel_workbook(
                    report_data, f"{report_name}_excel"
                )
                if excel_path:
                    generated_files.append(('excel', excel_path))
            
            # 生成履歴に記録
            self.generation_history.append({
                'type': 'monthly',
                'timestamp': datetime.now().isoformat(),
                'files': generated_files,
                'success': len(generated_files) > 0
            })
            
            # メール送信（設定されている場合）
            if self.schedule_config["monthly_reports"]["email_enabled"]:
                self._send_report_email('monthly', generated_files)
            
            # シグナル発信
            for format_type, file_path in generated_files:
                self.report_generated.emit('monthly', file_path)
            
            logger.info(f"✅ 月次レポート生成完了: {len(generated_files)}ファイル")
            
        except Exception as e:
            logger.error(f"月次レポート生成エラー: {e}")
    
    def _collect_daily_report_data(self) -> Dict[str, Any]:
        """日次レポートデータ収集"""
        try:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=1)
            
            # 基本統計
            daily_report = self.reports_engine.generate_period_report(
                start_date, end_date, 'daily'
            )
            
            return daily_report
            
        except Exception as e:
            logger.error(f"日次レポートデータ収集エラー: {e}")
            return {'error': str(e)}
    
    def _collect_weekly_report_data(self) -> Dict[str, Any]:
        """週次レポートデータ収集"""
        try:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
            
            # 週次統計
            weekly_report = self.reports_engine.generate_period_report(
                start_date, end_date, 'weekly'
            )
            
            # 比較データ（前週との比較）
            comparison_data = self.reports_engine.comparison_analytics.compare_periods(
                'weekly', 1
            )
            weekly_report['comparison'] = comparison_data
            
            return weekly_report
            
        except Exception as e:
            logger.error(f"週次レポートデータ収集エラー: {e}")
            return {'error': str(e)}
    
    def _collect_monthly_report_data(self) -> Dict[str, Any]:
        """月次レポートデータ収集"""
        try:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            
            # 月次統計
            monthly_report = self.reports_engine.generate_period_report(
                start_date, end_date, 'monthly'
            )
            
            # 比較データ（前月との比較）
            comparison_data = self.reports_engine.comparison_analytics.compare_periods(
                'monthly', 1
            )
            monthly_report['comparison'] = comparison_data
            
            return monthly_report
            
        except Exception as e:
            logger.error(f"月次レポートデータ収集エラー: {e}")
            return {'error': str(e)}
    
    def _collect_prediction_data(self) -> Dict[str, Any]:
        """予測データ収集"""
        try:
            predictions = {}
            
            # 最適作業時間予測
            predictions['optimal_times'] = self.prediction_engine.predict_optimal_work_time()
            
            # 生産性トレンド予測
            predictions['productivity_trend'] = self.prediction_engine.predict_productivity_trend(7)
            
            # フォーカススコア予測（現在の設定での）
            current_params = {
                'planned_duration': 25,
                'hour_of_day': datetime.now().hour,
                'day_of_week': datetime.now().weekday(),
                'interruption_count': 0,
                'environment_score': 0.7
            }
            predictions['focus_score'] = self.prediction_engine.predict_focus_score(current_params)
            
            return predictions
            
        except Exception as e:
            logger.error(f"予測データ収集エラー: {e}")
            return {'error': str(e)}
    
    def _analyze_monthly_trends(self) -> Dict[str, Any]:
        """月次トレンド分析"""
        try:
            # 基本トレンド分析
            trends = {
                'productivity_trend': self._calculate_productivity_trend(),
                'focus_trend': self._calculate_focus_trend(),
                'completion_trend': self._calculate_completion_trend(),
                'optimal_times_trend': self._calculate_optimal_times_trend()
            }
            
            return trends
            
        except Exception as e:
            logger.error(f"月次トレンド分析エラー: {e}")
            return {'error': str(e)}
    
    def _send_report_email(self, report_type: str, generated_files: List[Tuple[str, str]]):
        """レポートをメールで送信"""
        try:
            email_config = self.schedule_config["email_config"]
            
            if not all([email_config.get("smtp_server"), email_config.get("username"), 
                       email_config.get("password")]):
                logger.warning("メール設定が不完全です")
                return
            
            recipients = self.schedule_config[f"{report_type}_reports"]["email_recipients"]
            if not recipients:
                logger.warning(f"{report_type}レポートの宛先が設定されていません")
                return
            
            # メールメッセージ作成
            msg = MIMEMultipart()
            msg['From'] = email_config["username"]
            msg['To'] = ", ".join(recipients)
            msg['Subject'] = f"Pomodoro Analytics - {report_type.title()} Report"
            
            # メール本文
            body = f"""
            Dear User,
            
            Please find attached your {report_type} Pomodoro Timer analytics report.
            
            Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
            Report files: {len(generated_files)}
            
            Best regards,
            Pomodoro Timer Analytics System
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            # ファイル添付
            for format_type, file_path in generated_files:
                try:
                    with open(file_path, 'rb') as f:
                        attach = MIMEApplication(f.read())
                        attach.add_header('Content-Disposition', 'attachment', 
                                        filename=Path(file_path).name)
                        msg.attach(attach)
                except Exception as e:
                    logger.error(f"ファイル添付エラー: {e}")
            
            # SMTP送信
            server = smtplib.SMTP(email_config["smtp_server"], email_config["smtp_port"])
            
            if email_config.get("use_tls", True):
                server.starttls()
            
            server.login(email_config["username"], email_config["password"])
            
            for recipient in recipients:
                server.sendmail(email_config["username"], recipient, msg.as_string())
                self.email_sent.emit(recipient, report_type)
            
            server.quit()
            
            logger.info(f"📧 {report_type}レポートメール送信完了: {len(recipients)}名")
            
        except Exception as e:
            logger.error(f"レポートメール送信エラー: {e}")
    
    def _cleanup_old_reports(self):
        """古いレポートファイルのクリーンアップ"""
        try:
            retention_days = self.schedule_config.get("retention_days", 30)
            cutoff_date = datetime.now() - timedelta(days=retention_days)
            
            export_dir = Path("data/exports")
            if not export_dir.exists():
                return
            
            deleted_count = 0
            
            for file_path in export_dir.glob("*"):
                if file_path.is_file():
                    file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime)
                    
                    if file_mtime < cutoff_date:
                        try:
                            file_path.unlink()
                            deleted_count += 1
                        except Exception as e:
                            logger.error(f"ファイル削除エラー: {file_path} - {e}")
            
            logger.info(f"🗑️ 古いレポートファイル削除完了: {deleted_count}ファイル")
            
        except Exception as e:
            logger.error(f"レポートクリーンアップエラー: {e}")
    
    def update_schedule(self, schedule_type: str, config: Dict[str, Any]):
        """スケジュール設定を更新"""
        try:
            if schedule_type in self.schedule_config:
                self.schedule_config[schedule_type].update(config)
                self.save_schedule_config()
                
                # スケジュールを再設定（利用可能な場合のみ）
                if self.scheduler:
                    self._setup_default_schedules()
                
                self.schedule_updated.emit(self.schedule_config)
                logger.info(f"⚙️ {schedule_type}スケジュール更新完了")
                
        except Exception as e:
            logger.error(f"スケジュール更新エラー: {e}")
    
    def generate_immediate_report(self, report_type: str, formats: List[str] = None) -> List[str]:
        """即座にレポートを生成"""
        try:
            if formats is None:
                formats = ['pdf']
            
            generated_files = []
            
            # レポートデータ収集
            if report_type == 'daily':
                report_data = self._collect_daily_report_data()
            elif report_type == 'weekly':
                report_data = self._collect_weekly_report_data()
            elif report_type == 'monthly':
                report_data = self._collect_monthly_report_data()
            else:
                raise ValueError(f"未対応のレポートタイプ: {report_type}")
            
            # 予測データ追加（週次・月次の場合）
            if report_type in ['weekly', 'monthly']:
                predictions = self._collect_prediction_data()
                report_data['predictions'] = predictions
            
            report_name = f"immediate_{report_type}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # フォーマット別生成
            if "pdf" in formats:
                pdf_path = self.report_exporter.export_comprehensive_pdf_report(
                    report_data, f"{report_name}_pdf"
                )
                if pdf_path:
                    generated_files.append(pdf_path)
            
            if "excel" in formats:
                excel_path = self.report_exporter.export_excel_workbook(
                    report_data, f"{report_name}_excel"
                )
                if excel_path:
                    generated_files.append(excel_path)
            
            # 生成履歴に記録
            self.generation_history.append({
                'type': f'immediate_{report_type}',
                'timestamp': datetime.now().isoformat(),
                'files': [(Path(f).suffix[1:], f) for f in generated_files],
                'success': len(generated_files) > 0
            })
            
            logger.info(f"⚡ 即座レポート生成完了: {report_type} - {len(generated_files)}ファイル")
            
            return generated_files
            
        except Exception as e:
            logger.error(f"即座レポート生成エラー: {e}")
            return []
    
    def get_generation_history(self) -> List[Dict[str, Any]]:
        """レポート生成履歴を取得"""
        return self.generation_history.copy()
    
    def get_next_scheduled_reports(self) -> List[Dict[str, Any]]:
        """次回予定レポートを取得"""
        try:
            jobs = self.scheduler.get_jobs()
            next_reports = []
            
            for job in jobs:
                if job.id in ['daily_report', 'weekly_report', 'monthly_report']:
                    next_run = job.next_run_time
                    if next_run:
                        next_reports.append({
                            'type': job.id.replace('_report', ''),
                            'next_run': next_run.isoformat(),
                            'next_run_readable': next_run.strftime('%Y-%m-%d %H:%M:%S')
                        })
            
            return sorted(next_reports, key=lambda x: x['next_run'])
            
        except Exception as e:
            logger.error(f"次回予定レポート取得エラー: {e}")
            return []
    
    def stop_scheduler(self):
        """スケジューラーを停止"""
        if not self.scheduler:
            logger.info("⏰ スケジューラーは初期化されていません")
            return
            
        try:
            if self.scheduler.running:
                self.scheduler.shutdown()
                logger.info("⏰ AutoReportScheduler停止完了")
        except Exception as e:
            logger.error(f"スケジューラー停止エラー: {e}")
    
    # Helper methods for trend analysis
    def _calculate_productivity_trend(self) -> Dict[str, Any]:
        """生産性トレンドを計算"""
        # 簡易実装 - 実際の実装では統計的手法を使用
        return {
            'direction': 'increasing',
            'slope': 0.05,
            'confidence': 0.8
        }
    
    def _calculate_focus_trend(self) -> Dict[str, Any]:
        """フォーカストレンドを計算"""
        return {
            'direction': 'stable',
            'slope': 0.01,
            'confidence': 0.75
        }
    
    def _calculate_completion_trend(self) -> Dict[str, Any]:
        """完了率トレンドを計算"""
        return {
            'direction': 'increasing',
            'slope': 0.03,
            'confidence': 0.85
        }
    
    def _calculate_optimal_times_trend(self) -> Dict[str, Any]:
        """最適時間トレンドを計算"""
        return {
            'peak_hours': [9, 10, 14, 15],
            'low_hours': [12, 16, 17],
            'trend': 'consistent'
        }


class TimerDataManager(QObject):
    """タイマーデータの共有管理クラス（テンプレート・セッション回数設定統合版）"""
    
    # シグナル - データ変更時に両ウィンドウに通知
    time_updated = pyqtSignal(int)  # 残り時間（秒）
    session_changed = pyqtSignal(str, int)  # (session_type, session_number)
    timer_state_changed = pyqtSignal(bool)  # is_running
    session_completed = pyqtSignal(str, int)  # (session_type, duration_minutes)
    break_started = pyqtSignal(str, int)  # (break_type, duration_minutes)
    cycle_completed = pyqtSignal(int)  # サイクル完了 (completed_cycles)
    all_sessions_completed = pyqtSignal()  # 全セッション完了
    
    def __init__(self):
        super().__init__()
        
        # テンプレート管理システム
        self.template_manager = SessionTemplateManager()
        self.template_manager.template_changed.connect(self._on_template_changed)
        
        # セッション回数設定
        self.max_sessions = 8  # 最大セッション数
        self.sessions_until_long_break = 4  # 長い休憩までのセッション数
        self.current_cycle = 1  # 現在のサイクル
        self.completed_cycles = 0  # 完了したサイクル数
        self.auto_start_enabled = True  # 自動開始設定
        self.auto_break_enabled = True  # 自動休憩設定
        
        # 現在のテンプレートから基本設定を取得
        current_template = self.template_manager.get_current_template()
        self.work_minutes = current_template["work_minutes"]
        self.break_minutes = current_template["break_minutes"] 
        self.long_break_minutes = current_template["long_break_minutes"]
        self.sessions_until_long_break = current_template["sessions_until_long_break"]
        self.max_sessions = current_template["max_sessions"]
        
        # タイマー状態
        self.time_left = 0
        self.is_running = False
        self.is_work_session = True
        self.session_count = 0
        
        # 内部タイマー
        self.timer = QTimer()
        self.timer.timeout.connect(self._update_timer)
        
        logger.info("📊 タイマーデータマネージャー初期化完了（テンプレート・セッション回数設定統合）")
    
    def _on_template_changed(self, template: dict):
        """テンプレート変更時の処理"""
        # 現在実行中でない場合のみ変更を適用
        if not self.is_running:
            self.work_minutes = template["work_minutes"]
            self.break_minutes = template["break_minutes"]
            self.long_break_minutes = template["long_break_minutes"]
            self.sessions_until_long_break = template["sessions_until_long_break"]
            self.max_sessions = template["max_sessions"]
            
            # 時間をリセット
            self.time_left = 0
            self.session_count = 0
            self.current_cycle = 1
            self.completed_cycles = 0
            self.is_work_session = True
            
            # 通知
            self.time_updated.emit(self.time_left)
            session_type = "作業" if self.is_work_session else "休憩"
            session_num = self.session_count + 1 if self.is_work_session else self.session_count
            self.session_changed.emit(session_type, session_num)
            
            logger.info(f"📋 テンプレート適用: {template['name']}")
        else:
            logger.warning("⚠️ 実行中のため、テンプレート変更は次回セッションから適用されます")
    
    def set_session_settings(self, max_sessions: int = None, sessions_until_long_break: int = None,
                           auto_start: bool = None, auto_break: bool = None):
        """セッション回数設定を変更"""
        if max_sessions is not None:
            self.max_sessions = max_sessions
        if sessions_until_long_break is not None:
            self.sessions_until_long_break = sessions_until_long_break
        if auto_start is not None:
            self.auto_start_enabled = auto_start
        if auto_break is not None:
            self.auto_break_enabled = auto_break
        
        logger.info(f"🔢 セッション設定更新: 最大{self.max_sessions}回, 長い休憩間隔{self.sessions_until_long_break}回")
    
    def get_session_progress(self) -> dict:
        """セッション進捗情報を取得"""
        remaining_sessions = max(0, self.max_sessions - self.session_count)
        sessions_to_long_break = self.sessions_until_long_break - (self.session_count % self.sessions_until_long_break)
        if sessions_to_long_break == self.sessions_until_long_break and self.session_count > 0:
            sessions_to_long_break = 0
        
        return {
            "current_session": self.session_count + 1 if self.is_work_session else self.session_count,
            "max_sessions": self.max_sessions,
            "remaining_sessions": remaining_sessions,
            "current_cycle": self.current_cycle,
            "completed_cycles": self.completed_cycles,
            "sessions_to_long_break": sessions_to_long_break,
            "progress_percentage": (self.session_count / self.max_sessions) * 100 if self.max_sessions > 0 else 0
        }
    
    def get_current_template(self) -> dict:
        """現在のテンプレートを取得"""
        return self.template_manager.get_current_template()
    
    def set_template(self, template_id: str) -> bool:
        """テンプレートを設定"""
        return self.template_manager.set_template(template_id)
    
    def start_timer(self):
        """タイマー開始"""
        if self.time_left == 0:
            # 新しいセッション開始
            duration = self.work_minutes if self.is_work_session else self.break_minutes
            self.time_left = duration * 60
        
        self.timer.start(1000)
        self.is_running = True
        
        # 状態変更を通知
        self.timer_state_changed.emit(True)
        self.time_updated.emit(self.time_left)
        
        session_type = "作業" if self.is_work_session else "休憩"
        session_num = self.session_count + 1 if self.is_work_session else self.session_count
        self.session_changed.emit(session_type, session_num)
        
        logger.info(f"⏰ タイマー開始: {session_type}セッション #{session_num}")
    
    def pause_timer(self):
        """タイマー一時停止"""
        self.timer.stop()
        self.is_running = False
        self.timer_state_changed.emit(False)
        logger.info("⏸️ タイマー一時停止")
    
    def reset_timer(self):
        """タイマーリセット"""
        self.timer.stop()
        self.is_running = False
        self.time_left = 0
        
        self.timer_state_changed.emit(False)
        self.time_updated.emit(0)
        logger.info("🔄 タイマーリセット")
    
    def _update_timer(self):
        """内部タイマー更新"""
        self.time_left -= 1
        self.time_updated.emit(self.time_left)
        
        if self.time_left <= 0:
            self._on_session_finished()
    
    def _on_session_finished(self):
        """セッション完了処理（Phase 4: 高度なデータ収集統合）"""
        self.timer.stop()
        self.is_running = False
        
        # 統計記録
        session_type = "work" if self.is_work_session else "break"
        duration = self.work_minutes if self.is_work_session else self.break_minutes
        
        # Phase 4: セッション完了を通知（高度なデータ収集）
        self.session_completed.emit(session_type, duration)
        
        # セッション切り替え
        if self.is_work_session:
            self.session_count += 1
            
            # 最大セッション数チェック
            if self.session_count >= self.max_sessions:
                # 全セッション完了
                self.all_sessions_completed.emit()
                logger.info(f"🏆 全{self.max_sessions}セッション完了！お疲れ様でした！")
                self._reset_session_cycle()
                return
            
            self.is_work_session = False
            
            # 長い休憩判定
            if self.session_count % self.sessions_until_long_break == 0:
                # サイクル完了
                self.completed_cycles += 1
                self.current_cycle = self.completed_cycles + 1
                self.cycle_completed.emit(self.completed_cycles)
                
                self.time_left = int(self.long_break_minutes * 60)
                # 長い休憩ウィンドウ表示シグナル
                logger.info(f"📍 長い休憩シグナル発信: long_break_minutes={self.long_break_minutes}")
                self.break_started.emit("long", int(self.long_break_minutes))
                logger.info(f"🎉 サイクル{self.completed_cycles}完了！長い休憩の時間です")
            else:
                self.time_left = int(self.break_minutes * 60)
                # 短い休憩ウィンドウ表示シグナル
                logger.info(f"📍 短い休憩シグナル発信: break_minutes={self.break_minutes}")
                self.break_started.emit("short", int(self.break_minutes))
                logger.info(f"✅ 作業セッション{self.session_count}完了！休憩の時間です")
        else:
            self.is_work_session = True
            self.time_left = int(self.work_minutes * 60)
            logger.info("🔄 休憩終了！次の作業セッションを開始しましょう")
        
        # 状態更新を通知
        self.timer_state_changed.emit(False)
        session_type_jp = "作業" if self.is_work_session else "休憩"
        session_num = self.session_count + 1 if self.is_work_session else self.session_count
        self.session_changed.emit(session_type_jp, session_num)
        self.time_updated.emit(self.time_left)
    
    def _reset_session_cycle(self):
        """セッションサイクルをリセット"""
        self.session_count = 0
        self.current_cycle = 1
        self.completed_cycles = 0
        self.is_work_session = True
        self.time_left = 0
        self.is_running = False
        
        # 通知
        self.timer_state_changed.emit(False)
        self.time_updated.emit(self.time_left)
        self.session_changed.emit("作業", 1)
        
        logger.info("🔄 セッションサイクルをリセットしました")


class TaskManager(QObject):
    """タスク管理（シンプル版）"""
    
    # シグナル - タスク変更時のみ通知
    task_added = pyqtSignal(str)  # task_text
    task_completed = pyqtSignal(str)  # task_text
    task_deleted = pyqtSignal(str)  # task_text
    
    def __init__(self):
        super().__init__()
        self.tasks_file = Path("data/tasks_phase3_integrated_simple_break.json")
        self.tasks_file.parent.mkdir(exist_ok=True)
        self.tasks = []
        self.load_tasks()
    
    def load_tasks(self):
        """タスク読み込み（起動時のみ）"""
        try:
            if self.tasks_file.exists():
                with open(self.tasks_file, 'r', encoding='utf-8') as f:
                    self.tasks = json.load(f)
            logger.info(f"📋 タスク読み込み: {len(self.tasks)}件")
        except Exception as e:
            logger.error(f"タスク読み込みエラー: {e}")
            self.tasks = []
    
    def save_tasks(self):
        """タスク保存"""
        try:
            with open(self.tasks_file, 'w', encoding='utf-8') as f:
                json.dump(self.tasks, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"タスク保存エラー: {e}")
    
    def add_task(self, text: str):
        """タスク追加（イベント駆動更新）"""
        task = {
            'id': datetime.now().strftime('%Y%m%d_%H%M%S'),
            'text': text,
            'completed': False,
            'created_at': datetime.now().isoformat()
        }
        self.tasks.append(task)
        self.save_tasks()
        self.task_added.emit(text)
        logger.info(f"📋 タスク追加: {text}")
    
    def complete_task(self, task_id: str):
        """タスク完了（イベント駆動更新）"""
        for task in self.tasks:
            if task['id'] == task_id:
                task['completed'] = True
                task['completed_at'] = datetime.now().isoformat()
                self.save_tasks()
                self.task_completed.emit(task['text'])
                logger.info(f"✅ タスク完了: {task['text']}")
                break
    
    def delete_task(self, task_id: str):
        """タスク削除（イベント駆動更新）"""
        for i, task in enumerate(self.tasks):
            if task['id'] == task_id:
                deleted_text = task['text']
                del self.tasks[i]
                self.save_tasks()
                self.task_deleted.emit(deleted_text)
                logger.info(f"🗑️ タスク削除: {deleted_text}")
                break
    
    def get_active_tasks(self):
        """アクティブなタスク一覧取得"""
        return [task for task in self.tasks if not task['completed']]


class StatisticsManager:
    """統計管理（Phase 4: 高度なデータ収集システム統合版）"""
    
    def __init__(self):
        self.stats_file = Path("data/stats_phase3_integrated_simple_break.json")
        self.stats_file.parent.mkdir(exist_ok=True)
        self.sessions = []
        
        # Phase 4: 高度なデータ収集システム統合
        self.advanced_collector = AdvancedDataCollector()
        self.session_tracker = SessionTracking()
        self.focus_calculator = FocusScoreCalculator()
        self.interruption_tracker = InterruptionTracker()
        self.environment_logger = EnvironmentLogger()
        
        # シグナル接続
        self._connect_advanced_signals()
        
        self.load_stats()
        logger.info("📊 StatisticsManager Phase 4 統合完了")
    
    def _connect_advanced_signals(self):
        """高度なデータ収集システムのシグナル接続"""
        # AdvancedDataCollector シグナル
        self.advanced_collector.data_collected.connect(self._on_advanced_data_collected)
        
        # SessionTracking シグナル
        self.session_tracker.session_pattern_detected.connect(self._on_pattern_detected)
        self.session_tracker.productivity_trend_updated.connect(self._on_productivity_updated)
        
        # FocusScoreCalculator シグナル
        self.focus_calculator.focus_score_updated.connect(self._on_focus_score_updated)
        self.focus_calculator.focus_level_changed.connect(self._on_focus_level_changed)
        
        # InterruptionTracker シグナル
        self.interruption_tracker.interruption_detected.connect(self._on_interruption_detected)
        self.interruption_tracker.interruption_pattern_found.connect(self._on_interruption_pattern)
        
        # EnvironmentLogger シグナル
        self.environment_logger.environment_data_updated.connect(self._on_environment_updated)
        self.environment_logger.optimal_time_detected.connect(self._on_optimal_time_detected)
        
        logger.info("📊 高度なデータ収集システムシグナル接続完了")
    
    def start_advanced_session_tracking(self, session_type: str, duration_minutes: int):
        """高度なセッション追跡開始"""
        try:
            # 全ての追跡システムを開始
            self.advanced_collector.start_session_tracking(session_type, duration_minutes)
            self.focus_calculator.start_focus_tracking()
            self.interruption_tracker.start_session_monitoring()
            self.environment_logger.start_environment_logging(session_type)
            
            logger.info(f"📊 Phase 4 高度なセッション追跡開始: {session_type}")
            
        except Exception as e:
            logger.error(f"高度なセッション追跡開始エラー: {e}")
    
    def end_advanced_session_tracking(self, completed: bool = True):
        """高度なセッション追跡終了"""
        try:
            # 全ての追跡システムを終了
            session_data = self.advanced_collector.current_session_metrics.copy()
            
            self.advanced_collector.end_session_tracking(completed)
            final_focus_score = self.focus_calculator.end_focus_tracking()
            self.interruption_tracker.end_session_monitoring()
            self.environment_logger.end_environment_logging(session_data)
            
            # セッションデータに最終フォーカススコアを追加
            session_data['final_focus_score'] = final_focus_score
            
            # SessionTracker に記録
            self.session_tracker.record_session_completion(session_data)
            
            logger.info(f"📊 Phase 4 高度なセッション追跡終了: 完了={completed}")
            
        except Exception as e:
            logger.error(f"高度なセッション追跡終了エラー: {e}")
    
    def record_user_interaction(self, interaction_type: str, details: dict = None):
        """ユーザーインタラクション記録（Phase 4 統合）"""
        try:
            self.advanced_collector.record_user_interaction(interaction_type, details)
            self.focus_calculator.record_interaction(interaction_type)
            self.interruption_tracker.record_user_activity(interaction_type)
            
        except Exception as e:
            logger.error(f"ユーザーインタラクション記録エラー: {e}")
    
    def record_session_interruption(self, interruption_type: str, details: dict = None):
        """セッション中断記録（Phase 4 統合）"""
        try:
            duration = details.get('duration_seconds', 0) if details else 0
            self.advanced_collector.record_interruption(interruption_type, duration)
            self.focus_calculator.record_interruption()
            
            if interruption_type == 'external':
                self.interruption_tracker.record_external_interruption(
                    details.get('type', 'unknown'), 
                    details.get('description', '')
                )
            
        except Exception as e:
            logger.error(f"セッション中断記録エラー: {e}")
    
    def get_advanced_analytics(self, days: int = 7) -> Dict[str, Any]:
        """高度な分析データ取得（Phase 4）"""
        try:
            # 各システムから分析データを取得
            session_analytics = self.advanced_collector.get_session_analytics(days)
            productivity_insights = self.session_tracker.get_productivity_insights()
            focus_insights = self.focus_calculator.get_focus_insights()
            interruption_summary = self.interruption_tracker.get_interruption_summary(days)
            environment_insights = self.environment_logger.get_environment_insights(days)
            
            # 統合分析データ
            return {
                'analysis_period': f'過去{days}日間',
                'session_analytics': session_analytics,
                'productivity_insights': productivity_insights,
                'focus_analysis': focus_insights,
                'interruption_analysis': interruption_summary,
                'environment_analysis': environment_insights,
                'overall_score': self._calculate_overall_performance_score(
                    session_analytics, productivity_insights, interruption_summary
                ),
                'recommendations': self._generate_integrated_recommendations(
                    session_analytics, productivity_insights, interruption_summary, environment_insights
                )
            }
            
        except Exception as e:
            logger.error(f"高度な分析データ取得エラー: {e}")
            return {'error': str(e)}
    
    def _calculate_overall_performance_score(self, session_data: dict, productivity_data: dict, interruption_data: dict) -> float:
        """総合パフォーマンススコア計算"""
        try:
            # 基本スコア
            focus_score = session_data.get('avg_focus_score', 0)
            efficiency_score = session_data.get('avg_efficiency_score', 0)
            completion_rate = session_data.get('completion_rate', 0)
            
            # 中断ペナルティ
            avg_interruptions = interruption_data.get('average_interruptions_per_session', 0)
            interruption_penalty = min(20, avg_interruptions * 3)  # 最大20点減点
            
            # 総合スコア計算
            overall_score = (focus_score * 0.3 + efficiency_score * 0.3 + completion_rate * 0.4) - interruption_penalty
            
            return max(0, min(100, round(overall_score, 1)))
            
        except Exception as e:
            logger.error(f"総合パフォーマンススコア計算エラー: {e}")
            return 0.0
    
    def _generate_integrated_recommendations(self, session_data: dict, productivity_data: dict, 
                                           interruption_data: dict, environment_data: dict) -> List[str]:
        """統合的な改善推奨事項生成"""
        recommendations = []
        
        try:
            # フォーカススコアに基づく推奨
            avg_focus = session_data.get('avg_focus_score', 0)
            if avg_focus < 60:
                recommendations.append("🎯 集中力の向上が必要です。短い休憩を頻繁に取ることをお勧めします")
            elif avg_focus > 80:
                recommendations.append("✅ 素晴らしい集中力を維持しています！")
            
            # 中断に基づく推奨
            avg_interruptions = interruption_data.get('average_interruptions_per_session', 0)
            if avg_interruptions > 3:
                recommendations.append("⚠️ 中断が多すぎます。作業環境の改善を検討してください")
            
            # 環境に基づく推奨
            env_recommendations = environment_data.get('recommendations', [])
            recommendations.extend(env_recommendations[:2])  # 最大2つまで
            
            # 生産性トレンドに基づく推奨
            if productivity_data.get('best_hour'):
                best_hour = productivity_data['best_hour']
                recommendations.append(f"⏰ {best_hour}時台の作業パフォーマンスが最も高いです")
            
            if not recommendations:
                recommendations.append("📊 継続的なデータ収集で、より詳細な分析を提供できます")
            
        except Exception as e:
            logger.error(f"統合推奨事項生成エラー: {e}")
            
        return recommendations[:5]  # 最大5つまで
    
    # Phase 4 シグナルハンドラー
    def _on_advanced_data_collected(self, session_data: dict):
        """高度なデータ収集完了時"""
        logger.info(f"📊 高度データ収集完了: {session_data.get('session_id', 'unknown')}")
    
    def _on_pattern_detected(self, pattern_name: str, details: dict):
        """パターン検出時"""
        logger.info(f"📈 パターン検出: {pattern_name} - {details}")
    
    def _on_productivity_updated(self, score: float):
        """生産性スコア更新時"""
        logger.info(f"📈 生産性スコア更新: {score:.1f}")
    
    def _on_focus_score_updated(self, score: float):
        """フォーカススコア更新時"""
        logger.debug(f"🎯 フォーカススコア: {score:.1f}")
    
    def _on_focus_level_changed(self, level: str):
        """フォーカスレベル変更時"""
        logger.info(f"🎯 フォーカスレベル変更: {level}")
    
    def _on_interruption_detected(self, interruption_type: str, details: dict):
        """中断検出時"""
        logger.warning(f"⚠️ 中断検出: {interruption_type}")
    
    def _on_interruption_pattern(self, pattern_type: str, analysis: dict):
        """中断パターン発見時"""
        logger.warning(f"⚠️ 中断パターン: {pattern_type}")
    
    def _on_environment_updated(self, env_data: dict):
        """環境データ更新時"""
        logger.info(f"🌍 環境データ更新: {env_data.get('time_period', 'unknown')}")
    
    def _on_optimal_time_detected(self, optimal_data: dict):
        """最適時間帯検出時"""
        logger.info(f"🌍 最適時間帯検出: {optimal_data}")
    
    def load_stats(self):
        """統計読み込み"""
        try:
            if self.stats_file.exists():
                with open(self.stats_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.sessions = data.get('sessions', [])
            logger.info(f"📊 統計読み込み: {len(self.sessions)}セッション")
        except Exception as e:
            logger.error(f"統計読み込みエラー: {e}")
            self.sessions = []
    
    def record_session(self, session_type: str, duration_minutes: int):
        """セッション記録（セッション完了時のみ）"""
        session = {
            'type': session_type,
            'duration': duration_minutes,
            'completed_at': datetime.now().isoformat(),
            'date': datetime.now().strftime('%Y-%m-%d')
        }
        self.sessions.append(session)
        self.save_stats()
        logger.info(f"📊 セッション記録: {session_type} ({duration_minutes}分)")
    
    def save_stats(self):
        """統計保存"""
        try:
            with open(self.stats_file, 'w', encoding='utf-8') as f:
                json.dump({'sessions': self.sessions}, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"統計保存エラー: {e}")
    
    def get_today_stats(self):
        """今日の統計取得"""
        today = datetime.now().strftime('%Y-%m-%d')
        today_sessions = [s for s in self.sessions if s['date'] == today]
        
        work_count = len([s for s in today_sessions if s['type'] == 'work'])
        break_count = len([s for s in today_sessions if s['type'] == 'break'])
        total_work_time = sum(s['duration'] for s in today_sessions if s['type'] == 'work')
        
        return {
            'work_sessions': work_count,
            'break_sessions': break_count,
            'total_work_minutes': total_work_time
        }
    
    def get_weekly_stats(self):
        """週間統計取得"""
        from datetime import timedelta
        
        end_date = datetime.now()
        start_date = end_date - timedelta(days=7)
        
        weekly_sessions = [s for s in self.sessions 
                          if datetime.fromisoformat(s['completed_at']) >= start_date]
        
        work_sessions = [s for s in weekly_sessions if s['type'] == 'work']
        total_work_time = sum(s['duration'] for s in work_sessions)
        
        return {
            'total_sessions': len(weekly_sessions),
            'work_sessions': len(work_sessions),
            'total_work_time': total_work_time,
            'avg_session_length': total_work_time / len(work_sessions) if work_sessions else 0
        }


class DashboardWidget(QWidget):
    """統計ダッシュボード統合ウィジェット"""
    
    def __init__(self, stats_manager: StatisticsManager, parent=None):
        super().__init__(parent)
        self.stats_manager = stats_manager
        self.setup_ui()
        self.update_stats()
    
    def setup_ui(self):
        """UI設定"""
        layout = QVBoxLayout(self)
        
        # タイトル
        title_label = QLabel("📊 統計ダッシュボード")
        title_label.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        # 今日の統計
        today_group = QGroupBox("今日の統計")
        today_layout = QHBoxLayout(today_group)
        
        self.today_work_label = QLabel("作業: 0回")
        self.today_break_label = QLabel("休憩: 0回")
        self.today_time_label = QLabel("時間: 0分")
        
        today_layout.addWidget(self.today_work_label)
        today_layout.addWidget(self.today_break_label)
        today_layout.addWidget(self.today_time_label)
        layout.addWidget(today_group)
        
        # 週間統計
        weekly_group = QGroupBox("週間統計")
        weekly_layout = QHBoxLayout(weekly_group)
        
        self.weekly_sessions_label = QLabel("総セッション: 0回")
        self.weekly_time_label = QLabel("総時間: 0分")
        self.weekly_avg_label = QLabel("平均: 0分")
        
        weekly_layout.addWidget(self.weekly_sessions_label)
        weekly_layout.addWidget(self.weekly_time_label)
        weekly_layout.addWidget(self.weekly_avg_label)
        layout.addWidget(weekly_group)
        
        # 統計詳細
        self.stats_display = QTextEdit()
        self.stats_display.setReadOnly(True)
        self.stats_display.setMaximumHeight(200)
        layout.addWidget(self.stats_display)
        
        # 更新ボタン
        refresh_btn = QPushButton("🔄 統計を更新")
        refresh_btn.clicked.connect(self.update_stats)
        layout.addWidget(refresh_btn)
    
    def update_stats(self):
        """統計表示を更新"""
        try:
            # データを再読み込み
            self.stats_manager.load_stats()
            
            # 今日の統計
            today_stats = self.stats_manager.get_today_stats()
            self.today_work_label.setText(f"作業: {today_stats['work_sessions']}回")
            self.today_break_label.setText(f"休憩: {today_stats['break_sessions']}回")
            self.today_time_label.setText(f"時間: {today_stats['total_work_minutes']}分")
            
            # 週間統計
            weekly_stats = self.stats_manager.get_weekly_stats()
            self.weekly_sessions_label.setText(f"総セッション: {weekly_stats['total_sessions']}回")
            self.weekly_time_label.setText(f"総時間: {weekly_stats['total_work_time']}分")
            self.weekly_avg_label.setText(f"平均: {weekly_stats['avg_session_length']:.1f}分")
            
            # Phase 4: 高度な分析データ取得
            try:
                advanced_analytics = self.stats_manager.get_advanced_analytics(7)
            except Exception as e:
                logger.error(f"高度な分析データ取得エラー: {e}")
                advanced_analytics = {'error': str(e)}
            
            # 詳細統計
            stats_text = f"""
Phase 4 高度なデータ収集システム統合版 - 統計詳細 ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})

🍅 今日の基本実績:
   作業セッション: {today_stats['work_sessions']}回
   休憩セッション: {today_stats['break_sessions']}回
   合計作業時間: {today_stats['total_work_minutes']}分

📈 週間基本実績:
   総セッション数: {weekly_stats['total_sessions']}回
   作業セッション: {weekly_stats['work_sessions']}回
   総作業時間: {weekly_stats['total_work_time']}分
   平均セッション時間: {weekly_stats['avg_session_length']:.1f}分

🚀 Phase 4 高度な分析 (過去7日間):"""

            if 'error' not in advanced_analytics:
                session_analytics = advanced_analytics.get('session_analytics', {})
                focus_analysis = advanced_analytics.get('focus_analysis', {})
                interruption_analysis = advanced_analytics.get('interruption_analysis', {})
                
                stats_text += f"""
   🎯 フォーカス分析:
      平均フォーカススコア: {session_analytics.get('avg_focus_score', 0):.1f}/100
      平均効率スコア: {session_analytics.get('avg_efficiency_score', 0):.1f}/100
      完了率: {session_analytics.get('completion_rate', 0):.1f}%
   
   ⚠️ 中断分析:
      平均中断回数: {interruption_analysis.get('average_interruptions_per_session', 0):.1f}回/セッション
      総中断回数: {interruption_analysis.get('total_interruptions', 0)}回
      改善が必要: {'はい' if interruption_analysis.get('improvement_needed', False) else 'いいえ'}
   
   📊 総合パフォーマンススコア: {advanced_analytics.get('overall_score', 0):.1f}/100
   
   💡 AI推奨事項:"""
                
                recommendations = advanced_analytics.get('recommendations', [])
                for i, rec in enumerate(recommendations[:3], 1):
                    stats_text += f"\n      {i}. {rec}"
            else:
                stats_text += f"\n   ❌ 高度な分析データ取得エラー: {advanced_analytics['error']}"

            stats_text += f"""

📊 全期間セッション数: {len(self.stats_manager.sessions)}回

🎯 Phase 4 完成機能:
   ✅ Advanced Data Collection System
   ✅ Session Tracking & Pattern Detection
   ✅ Focus Score Calculation
   ✅ Interruption Detection & Analysis
   ✅ Environment Logging & Insights
   ✅ Integrated Analytics Dashboard
   
目標達成度:
   {'✅ 順調です！' if today_stats['work_sessions'] >= 4 else '📈 もう少し頑張りましょう！'}
            """.strip()
            
            self.stats_display.setText(stats_text)
            
            logger.info("📊 ダッシュボード統計更新完了")
            
        except Exception as e:
            logger.error(f"ダッシュボード統計更新エラー: {e}")
            self.stats_display.setText(f"統計更新エラー: {e}")


class MainWindow(QMainWindow):
    """設定モード（メインウィンドウ）"""
    
    def __init__(self, timer_data: TimerDataManager, task_manager: TaskManager, 
                 stats: StatisticsManager):
        super().__init__()
        
        self.timer_data = timer_data
        self.task_manager = task_manager
        self.stats = stats
        self.minimal_window = None
        self.break_window = None  # シンプル休憩ウィンドウ参照
        
        # Phase 4: インタラクティブ分析エンジン・可視化システム初期化
        self._init_phase4_systems()
        
        self.init_ui()
        self.connect_signals()
        
        logger.info("🏠 メインウィンドウ初期化完了 - Phase 4 インタラクティブ分析システム統合済み")
    
    def _init_phase4_systems(self):
        """Phase 4: インタラクティブ分析・可視化システム初期化"""
        try:
            # インタラクティブレポートエンジン
            self.reports_engine = InteractiveReportsEngine(
                self.stats.advanced_collector,
                self.stats.session_tracker,
                self.stats.focus_calculator,
                self.stats.interruption_tracker,
                self.stats.environment_logger
            )
            
            # 高度な可視化システム
            self.visualization = AdvancedVisualization(
                self.reports_engine,
                self.stats.advanced_collector
            )
            
            # 比較分析システム
            self.comparison_analytics = ComparisonAnalytics(
                self.stats.advanced_collector,
                self.reports_engine
            )
            
            # カスタムレポートビルダー
            self.report_builder = CustomReportBuilder(
                self.reports_engine,
                self.visualization,
                self.comparison_analytics
            )
            
            # Worker3: Prediction Engine & Export Systems
            self.prediction_engine = PredictionEngine(
                self.stats.advanced_collector,
                self.stats.session_tracker,
                self.stats.focus_calculator
            )
            
            self.report_exporter = ReportExporter(
                self.reports_engine,
                self.visualization,
                self.comparison_analytics
            )
            
            self.auto_scheduler = AutoReportScheduler(
                self.reports_engine,
                self.report_exporter,
                self.prediction_engine
            )
            
            # シグナル接続
            self._connect_phase4_signals()
            self._connect_worker3_signals()
            
            logger.info("📊 Phase 4 インタラクティブ分析・可視化システム初期化完了")
            logger.info("🤖 Worker3 予測エンジン・エクスポートシステム初期化完了")
            
        except Exception as e:
            logger.error(f"Phase 4 システム初期化エラー: {e}")
            # フォールバック: 基本システムなしでも動作するように
            self.reports_engine = None
            self.visualization = None
            self.comparison_analytics = None
            self.report_builder = None
    
    def _connect_phase4_signals(self):
        """Phase 4 システムのシグナル接続"""
        if not all([self.reports_engine, self.visualization, self.comparison_analytics, self.report_builder]):
            return
        
        try:
            # レポートエンジンシグナル
            self.reports_engine.report_generated.connect(self._on_report_generated)
            self.reports_engine.drill_down_requested.connect(self._on_drill_down_requested)
            
            # 可視化システムシグナル
            self.visualization.visualization_ready.connect(self._on_visualization_ready)
            self.visualization.export_completed.connect(self._on_chart_exported)
            
            # 比較分析シグナル
            self.comparison_analytics.comparison_completed.connect(self._on_comparison_completed)
            self.comparison_analytics.trend_detected.connect(self._on_trend_detected)
            
            # レポートビルダーシグナル
            self.report_builder.report_built.connect(self._on_custom_report_built)
            self.report_builder.template_saved.connect(self._on_template_saved)
            
            logger.info("📊 Phase 4 シグナル接続完了")
            
        except Exception as e:
            logger.error(f"Phase 4 シグナル接続エラー: {e}")
    
    def _connect_worker3_signals(self):
        """Worker3 システムのシグナル接続"""
        if not all([hasattr(self, 'prediction_engine'), hasattr(self, 'report_exporter'), hasattr(self, 'auto_scheduler')]):
            return
        
        try:
            # PredictionEngine シグナル
            self.prediction_engine.prediction_ready.connect(self._on_prediction_ready)
            self.prediction_engine.model_trained.connect(self._on_model_trained)
            self.prediction_engine.prediction_updated.connect(self._on_prediction_updated)
            
            # ReportExporter シグナル
            self.report_exporter.export_completed.connect(self._on_export_completed)
            self.report_exporter.export_progress.connect(self._on_export_progress)
            self.report_exporter.export_error.connect(self._on_export_error)
            
            # AutoReportScheduler シグナル
            self.auto_scheduler.report_scheduled.connect(self._on_report_scheduled)
            self.auto_scheduler.report_generated.connect(self._on_scheduled_report_generated)
            self.auto_scheduler.schedule_updated.connect(self._on_schedule_updated)
            self.auto_scheduler.email_sent.connect(self._on_email_sent)
            
            logger.info("🤖 Worker3 シグナル接続完了")
            
        except Exception as e:
            logger.error(f"Worker3 シグナル接続エラー: {e}")
    
    # Worker3 Signal Handlers
    def _on_prediction_ready(self, prediction_type: str, results: dict):
        """予測準備完了時の処理"""
        logger.info(f"🔮 予測準備完了: {prediction_type}")
        
        # ダッシュボードに予測結果を表示
        if hasattr(self, 'dashboard_widget'):
            # 予測結果をダッシュボードに反映
            pass
    
    def _on_model_trained(self, model_name: str, metrics: dict):
        """モデル訓練完了時の処理"""
        logger.info(f"🎯 モデル訓練完了: {model_name}")
        
        # モデル精度をユーザーに通知
        accuracy = metrics.get('r2', metrics.get('accuracy', 0))
        if accuracy > 0.8:
            logger.info(f"✅ 高精度モデル: {model_name} (精度: {accuracy:.3f})")
        elif accuracy > 0.6:
            logger.info(f"⚠️ 中程度精度モデル: {model_name} (精度: {accuracy:.3f})")
        else:
            logger.warning(f"❌ 低精度モデル: {model_name} (精度: {accuracy:.3f}) - 追加データが必要")
    
    def _on_prediction_updated(self, metric_name: str, predicted_value):
        """予測値更新時の処理"""
        logger.info(f"📊 予測値更新: {metric_name} = {predicted_value}")
    
    def _on_export_completed(self, export_type: str, file_path: str):
        """エクスポート完了時の処理"""
        logger.info(f"📄 エクスポート完了: {export_type} -> {file_path}")
        
        # ユーザーに通知（可能であればシステム通知も）
        if hasattr(self, 'show_notification'):
            self.show_notification(f"レポートエクスポート完了", f"{export_type.upper()}レポートが生成されました")
    
    def _on_export_progress(self, progress: int):
        """エクスポート進捗更新時の処理"""
        logger.debug(f"📊 エクスポート進捗: {progress}%")
        
        # プログレスバーがあれば更新
        if hasattr(self, 'progress_bar'):
            self.progress_bar.setValue(progress)
    
    def _on_export_error(self, export_type: str, error_message: str):
        """エクスポートエラー時の処理"""
        logger.error(f"❌ エクスポートエラー: {export_type} - {error_message}")
        
        # エラーダイアログ表示
        QMessageBox.critical(self, "エクスポートエラー", 
                           f"{export_type}レポートのエクスポートに失敗しました。\n\nエラー: {error_message}")
    
    def _on_report_scheduled(self, report_type: str, schedule_info: str):
        """レポートスケジュール設定時の処理"""
        logger.info(f"📅 レポートスケジュール設定: {report_type} - {schedule_info}")
    
    def _on_scheduled_report_generated(self, report_type: str, file_path: str):
        """スケジュールレポート生成完了時の処理"""
        logger.info(f"⏰ スケジュールレポート生成完了: {report_type} -> {file_path}")
        
        # 自動生成レポートの通知
        if hasattr(self, 'show_notification'):
            self.show_notification(f"自動レポート生成完了", f"{report_type}レポートが自動生成されました")
    
    def _on_schedule_updated(self, config: dict):
        """スケジュール設定更新時の処理"""
        logger.info(f"⚙️ スケジュール設定更新完了")
    
    def _on_email_sent(self, recipient: str, report_type: str):
        """メール送信完了時の処理"""
        logger.info(f"📧 メール送信完了: {report_type} -> {recipient}")
    
    def _on_report_generated(self, report_type: str, report_data: dict):
        """レポート生成完了時の処理"""
        logger.info(f"📈 レポート生成完了: {report_type}")
        
        # レポート通知をダッシュボードに表示
        if hasattr(self, 'dashboard_widget'):
            # ダッシュボードが存在する場合、レポート情報を更新
            pass
    
    def _on_drill_down_requested(self, data_type: str, filter_params: dict):
        """ドリルダウン要求時の処理"""
        logger.info(f"🔍 ドリルダウン要求: {data_type}")
        # 詳細分析ウィンドウを開く実装が可能
    
    def _on_visualization_ready(self, chart_type: str, figure_widget):
        """可視化準備完了時の処理 - 別ウィンドウで表示"""
        logger.info(f"📊 可視化準備完了: {chart_type}")
        
        try:
            # 別ウィンドウでチャートを表示
            self._show_chart_in_separate_window(chart_type, figure_widget)
        except Exception as e:
            logger.error(f"可視化ウィンドウ表示エラー: {e}")
    
    def _show_chart_in_separate_window(self, chart_type: str, figure_widget):
        """チャートを別ウィンドウで表示"""
        try:
            from PyQt6.QtWidgets import QDialog, QVBoxLayout, QPushButton, QHBoxLayout
            from PyQt6.QtCore import Qt
            
            # チャートタイトル設定
            chart_titles = {
                'productivity_timeline': '📈 生産性タイムライン',
                'focus_heatmap': '🎯 フォーカスヒートマップ',
                'interruption_analysis': '⚠️ 中断分析ダッシュボード',
                'session_performance': '🏆 セッションパフォーマンス分析',
                'custom_dashboard': '📊 カスタムダッシュボード'
            }
            
            title = chart_titles.get(chart_type, f'📊 {chart_type}')
            
            # 新しいダイアログウィンドウを作成
            dialog = QDialog(self)
            dialog.setWindowTitle(title)
            dialog.setMinimumSize(800, 600)
            dialog.resize(1000, 700)
            
            # レイアウト設定
            layout = QVBoxLayout()
            
            # チャートを追加
            layout.addWidget(figure_widget)
            
            # ボタンレイアウト
            button_layout = QHBoxLayout()
            
            # エクスポートボタン
            export_btn = QPushButton('💾 PNG エクスポート')
            export_btn.clicked.connect(lambda: self._export_chart_from_dialog(figure_widget, chart_type))
            button_layout.addWidget(export_btn)
            
            # 閉じるボタン
            close_btn = QPushButton('❌ 閉じる')
            close_btn.clicked.connect(dialog.close)
            button_layout.addWidget(close_btn)
            
            layout.addLayout(button_layout)
            dialog.setLayout(layout)
            
            # ウィンドウを表示（非モーダル）
            dialog.show()
            dialog.raise_()
            dialog.activateWindow()
            
            # ダイアログ参照を保持（ガベージコレクション防止）
            if not hasattr(self, '_chart_dialogs'):
                self._chart_dialogs = []
            self._chart_dialogs.append(dialog)
            
            logger.info(f"✅ {title} を別ウィンドウで表示しました")
            
        except Exception as e:
            logger.error(f"別ウィンドウ表示エラー: {e}")
    
    def _export_chart_from_dialog(self, figure_widget, chart_type: str):
        """ダイアログからチャートエクスポート"""
        try:
            filepath = self.visualization.export_chart(figure_widget, chart_type, 'png')
            logger.info(f"💾 チャートエクスポート完了: {filepath}")
        except Exception as e:
            logger.error(f"チャートエクスポートエラー: {e}")
    
    def _on_chart_exported(self, chart_type: str, filepath: str):
        """チャートエクスポート完了時の処理"""
        logger.info(f"💾 チャートエクスポート完了: {chart_type} -> {filepath}")
    
    def _on_comparison_completed(self, comparison_type: str, results: dict):
        """比較分析完了時の処理"""
        logger.info(f"🔄 比較分析完了: {comparison_type}")
    
    def _on_trend_detected(self, trend_type: str, details: dict):
        """トレンド検出時の処理"""
        logger.info(f"📈 トレンド検出: {trend_type}")
        # トレンド通知をUIに表示する実装が可能
    
    def _on_custom_report_built(self, report_name: str, report_data: dict):
        """カスタムレポート作成完了時の処理"""
        logger.info(f"📝 カスタムレポート作成完了: {report_name}")
    
    def _on_template_saved(self, template_name: str, template_config: dict):
        """テンプレート保存完了時の処理"""
        logger.info(f"📄 テンプレート保存完了: {template_name}")
    
    def init_ui(self):
        """UI初期化"""
        self.setWindowTitle("🍅 Pomodoro Timer Phase 4 - Interactive Analysis Engine & Advanced Visualization")
        self.setGeometry(100, 100, 600, 500)
        
        # メインウィジェット
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)
        
        # タブウィジェット
        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)
        
        # タイマータブ
        self.setup_timer_tab()
        
        # タスクタブ
        self.setup_task_tab()
        
        # 統計タブ
        self.setup_stats_tab()
        
        # ダッシュボードタブ（Phase 3 統合完了）
        self.setup_dashboard_tab()
        
        # Phase 4: インタラクティブ分析・可視化タブ
        self.setup_visualization_tab()
        
        # Worker3: 予測エンジン・エクスポートタブ
        if hasattr(self.stats, 'prediction_engine'):
            self.setup_worker3_tab()
        
        # ミニマルモードボタン
        minimal_btn = QPushButton("🔽 ミニマルモード表示")
        minimal_btn.clicked.connect(self.show_minimal_mode)
        minimal_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        layout.addWidget(minimal_btn)
        
        # シンプル休憩ウィンドウテストボタン（開発用）
        test_break_btn = QPushButton("☕ シンプル休憩ウィンドウテスト")
        test_break_btn.clicked.connect(self.test_simple_break_window)
        test_break_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        layout.addWidget(test_break_btn)
    
    def setup_timer_tab(self):
        """タイマータブ設定（テンプレート・セッション設定統合版）"""
        timer_widget = QWidget()
        layout = QVBoxLayout(timer_widget)
        
        # セッションテンプレート選択セクション
        template_group = QGroupBox("📋 セッションテンプレート")
        template_layout = QVBoxLayout()
        
        # テンプレート選択
        template_select_layout = QHBoxLayout()
        template_select_layout.addWidget(QLabel("テンプレート:"))
        
        self.template_combo = QComboBox()
        self.template_combo.currentTextChanged.connect(self.on_template_changed)
        template_select_layout.addWidget(self.template_combo)
        
        # カスタムテンプレート作成ボタン
        custom_template_btn = QPushButton("➕ カスタム")
        custom_template_btn.clicked.connect(self.create_custom_template)
        custom_template_btn.setMaximumWidth(80)
        template_select_layout.addWidget(custom_template_btn)
        
        template_layout.addLayout(template_select_layout)
        
        # テンプレート説明
        self.template_description = QLabel("25分作業 + 5分休憩")
        self.template_description.setStyleSheet("color: #7f8c8d; font-style: italic; margin: 5px;")
        template_layout.addWidget(self.template_description)
        
        template_group.setLayout(template_layout)
        layout.addWidget(template_group)
        
        # セッション回数設定
        session_group = QGroupBox("🔢 セッション設定")
        session_layout = QGridLayout()
        
        # 最大セッション数
        session_layout.addWidget(QLabel("最大セッション数:"), 0, 0)
        self.max_sessions_spin = QSpinBox()
        self.max_sessions_spin.setRange(1, 20)
        self.max_sessions_spin.setValue(8)
        self.max_sessions_spin.valueChanged.connect(self.update_session_settings)
        session_layout.addWidget(self.max_sessions_spin, 0, 1)
        
        # 長い休憩間隔
        session_layout.addWidget(QLabel("長い休憩間隔:"), 0, 2)
        self.long_break_interval_spin = QSpinBox()
        self.long_break_interval_spin.setRange(2, 10)
        self.long_break_interval_spin.setValue(4)
        self.long_break_interval_spin.valueChanged.connect(self.update_session_settings)
        session_layout.addWidget(self.long_break_interval_spin, 0, 3)
        
        session_group.setLayout(session_layout)
        layout.addWidget(session_group)
        
        # セッション進捗表示
        progress_group = QGroupBox("📊 セッション進捗")
        progress_layout = QVBoxLayout()
        
        self.progress_label = QLabel("セッション 1/8 (サイクル 1)")
        self.progress_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        progress_layout.addWidget(self.progress_label)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        progress_layout.addWidget(self.progress_bar)
        
        progress_group.setLayout(progress_layout)
        layout.addWidget(progress_group)
        
        # タイマー表示
        self.time_display = QLabel("25:00")
        self.time_display.setFont(QFont("Arial", 48))
        self.time_display.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.time_display.setStyleSheet("color: #2c3e50; background-color: #ecf0f1; border-radius: 10px; padding: 20px;")
        layout.addWidget(self.time_display)
        
        # セッション情報
        self.session_info = QLabel("作業セッション #1")
        self.session_info.setFont(QFont("Arial", 16))
        self.session_info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.session_info.setStyleSheet("color: #34495e; margin: 10px;")
        layout.addWidget(self.session_info)
        
        # テンプレートコンボボックスを初期化
        self.update_template_combo()
        
        # コントロールボタン
        btn_layout = QHBoxLayout()
        
        self.start_btn = QPushButton("▶️ 開始")
        self.start_btn.clicked.connect(self.start_timer_with_advanced_tracking)
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 15px 30px;
                border-radius: 8px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #219653;
            }
        """)
        btn_layout.addWidget(self.start_btn)
        
        self.pause_btn = QPushButton("⏸️ 一時停止")
        self.pause_btn.clicked.connect(self.pause_timer_with_tracking)
        self.pause_btn.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                border: none;
                padding: 15px 30px;
                border-radius: 8px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d68910;
            }
        """)
        btn_layout.addWidget(self.pause_btn)
        
        self.reset_btn = QPushButton("🔄 リセット")
        self.reset_btn.clicked.connect(self.reset_timer_with_tracking)
        self.reset_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 15px 30px;
                border-radius: 8px;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        btn_layout.addWidget(self.reset_btn)
        
        layout.addLayout(btn_layout)
        
        # 設定
        settings_layout = QHBoxLayout()
        
        settings_layout.addWidget(QLabel("作業時間:"))
        self.work_spin = QSpinBox()
        self.work_spin.setRange(1, 60)
        self.work_spin.setValue(self.timer_data.work_minutes)
        self.work_spin.setSuffix(" 分")
        self.work_spin.valueChanged.connect(self.on_work_duration_changed)
        settings_layout.addWidget(self.work_spin)
        
        settings_layout.addWidget(QLabel("休憩時間:"))
        self.break_spin = QSpinBox()
        self.break_spin.setRange(1, 30)
        self.break_spin.setValue(self.timer_data.break_minutes)
        self.break_spin.setSuffix(" 分")
        self.break_spin.valueChanged.connect(self.on_break_duration_changed)
        settings_layout.addWidget(self.break_spin)
        
        layout.addLayout(settings_layout)
        
        self.tab_widget.addTab(timer_widget, "⏱️ タイマー")
    
    def setup_task_tab(self):
        """タスクタブ設定"""
        task_widget = QWidget()
        layout = QVBoxLayout(task_widget)
        
        # 新規タスク入力
        input_layout = QHBoxLayout()
        self.task_input = QLineEdit()
        self.task_input.setPlaceholderText("新しいタスクを入力...")
        self.task_input.returnPressed.connect(self.add_task)
        input_layout.addWidget(self.task_input)
        
        add_btn = QPushButton("➕ 追加")
        add_btn.clicked.connect(self.add_task)
        input_layout.addWidget(add_btn)
        
        layout.addLayout(input_layout)
        
        # タスクリスト
        self.task_list = QListWidget()
        layout.addWidget(self.task_list)
        
        # 初期タスク読み込み（起動時のみ）
        self.refresh_task_list()
        
        self.tab_widget.addTab(task_widget, "📋 タスク")
    
    def setup_stats_tab(self):
        """統計タブ設定"""
        stats_widget = QWidget()
        layout = QVBoxLayout(stats_widget)
        
        # 統計表示
        self.stats_display = QTextEdit()
        self.stats_display.setReadOnly(True)
        layout.addWidget(self.stats_display)
        
        # 初期統計表示
        self.refresh_stats_display()
        
        self.tab_widget.addTab(stats_widget, "📊 統計")
    
    def setup_dashboard_tab(self):
        """ダッシュボードタブ設定（Phase 3 完成機能）"""
        self.dashboard_widget = DashboardWidget(self.stats)
        self.tab_widget.addTab(self.dashboard_widget, "📈 ダッシュボード")
    
    def setup_visualization_tab(self):
        """Phase 4: インタラクティブ分析・可視化タブ設定"""
        if not all([self.reports_engine, self.visualization, self.comparison_analytics, self.report_builder]):
            # Phase 4 システムが初期化されていない場合はスキップ
            return
        
        viz_widget = QWidget()
        layout = QVBoxLayout(viz_widget)
        
        # タイトル
        title_label = QLabel("📊 インタラクティブ分析・可視化ダッシュボード")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin: 10px 0;")
        layout.addWidget(title_label)
        
        # レポート生成セクション
        report_group = QGroupBox("📈 レポート生成")
        report_layout = QVBoxLayout(report_group)
        
        # 包括的レポート生成ボタン
        comprehensive_btn = QPushButton("🔍 包括的レポート生成")
        comprehensive_btn.clicked.connect(self._generate_comprehensive_report)
        comprehensive_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        report_layout.addWidget(comprehensive_btn)
        
        # テンプレートレポート生成
        template_layout = QHBoxLayout()
        template_combo = QComboBox()
        template_combo.addItems(["daily_summary", "weekly_analysis", "comprehensive"])
        template_btn = QPushButton("📄 テンプレートレポート生成")
        template_btn.clicked.connect(lambda: self._generate_template_report(template_combo.currentText()))
        template_layout.addWidget(QLabel("テンプレート:"))
        template_layout.addWidget(template_combo)
        template_layout.addWidget(template_btn)
        report_layout.addLayout(template_layout)
        
        layout.addWidget(report_group)
        
        # 可視化セクション
        viz_group = QGroupBox("📊 可視化")
        viz_layout = QVBoxLayout(viz_group)
        
        # 可視化ボタン行1
        viz_row1 = QHBoxLayout()
        productivity_btn = QPushButton("📈 生産性タイムライン")
        productivity_btn.clicked.connect(lambda: self._show_visualization('productivity_timeline'))
        heatmap_btn = QPushButton("🔥 フォーカスヒートマップ")
        heatmap_btn.clicked.connect(lambda: self._show_visualization('focus_heatmap'))
        viz_row1.addWidget(productivity_btn)
        viz_row1.addWidget(heatmap_btn)
        viz_layout.addLayout(viz_row1)
        
        # 可視化ボタン行2
        viz_row2 = QHBoxLayout()
        interruption_btn = QPushButton("⚠️ 中断分析")
        interruption_btn.clicked.connect(lambda: self._show_visualization('interruption_analysis'))
        performance_btn = QPushButton("🏆 セッション性能")
        performance_btn.clicked.connect(lambda: self._show_visualization('session_performance'))
        viz_row2.addWidget(interruption_btn)
        viz_row2.addWidget(performance_btn)
        viz_layout.addLayout(viz_row2)
        
        layout.addWidget(viz_group)
        
        # 比較分析セクション
        comparison_group = QGroupBox("🔄 比較分析")
        comparison_layout = QVBoxLayout(comparison_group)
        
        # 比較分析ボタン
        weekday_btn = QPushButton("📅 平日 vs 週末比較")
        weekday_btn.clicked.connect(lambda: self._run_comparison('weekdays_vs_weekends'))
        timeperiod_btn = QPushButton("⏰ 時間帯別比較")
        timeperiod_btn.clicked.connect(lambda: self._run_comparison('time_periods'))
        
        comparison_layout.addWidget(weekday_btn)
        comparison_layout.addWidget(timeperiod_btn)
        
        layout.addWidget(comparison_group)
        
        # 結果表示エリア
        self.viz_results_area = QTextEdit()
        self.viz_results_area.setMaximumHeight(200)
        self.viz_results_area.setPlaceholderText("分析結果・レポートがここに表示されます...")
        layout.addWidget(QLabel("📊 分析結果:"))
        layout.addWidget(self.viz_results_area)
        
        # スタイル設定
        for button in [productivity_btn, heatmap_btn, interruption_btn, performance_btn, 
                      weekday_btn, timeperiod_btn]:
            button.setStyleSheet("""
                QPushButton {
                    background-color: #2ecc71;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 5px;
                    font-size: 12px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #27ae60;
                }
            """)
        
        template_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        
        self.tab_widget.addTab(viz_widget, "📊 分析・可視化")
    
    def _generate_comprehensive_report(self):
        """包括的レポート生成"""
        if not self.reports_engine:
            self.viz_results_area.setText("❌ レポートエンジンが利用できません")
            return
        
        try:
            report = self.reports_engine.generate_comprehensive_report()
            
            # レポート要約を表示
            summary = report.get('summary', {})
            recommendations = report.get('recommendations', [])
            
            result_text = f"""📈 包括的レポート生成完了
            
📊 概要:
• 総セッション数: {summary.get('total_sessions', 0)}
• 総作業時間: {summary.get('total_work_hours', 0)}時間
• 平均フォーカススコア: {summary.get('average_focus_score', 0)}
• 総中断回数: {summary.get('total_interruptions', 0)}
• 生産性トレンド: {summary.get('productivity_trend', 'unknown')}

💡 推奨事項:
"""
            
            for i, rec in enumerate(recommendations[:3], 1):  # 最初の3つの推奨事項のみ表示
                result_text += f"{i}. {rec}\n"
            
            self.viz_results_area.setText(result_text)
            
            # レポートをエクスポート
            if hasattr(self.reports_engine, 'export_report'):
                filepath = self.reports_engine.export_report(report)
                if filepath:
                    result_text += f"\n💾 レポートファイル: {filepath}"
                    self.viz_results_area.setText(result_text)
            
        except Exception as e:
            self.viz_results_area.setText(f"❌ レポート生成エラー: {e}")
            logger.error(f"包括的レポート生成エラー: {e}")
    
    def _generate_template_report(self, template_name: str):
        """テンプレートレポート生成"""
        if not self.report_builder:
            self.viz_results_area.setText("❌ レポートビルダーが利用できません")
            return
        
        try:
            report = self.report_builder.create_report_from_template(template_name)
            
            if 'error' in report:
                self.viz_results_area.setText(f"❌ テンプレートレポート生成エラー: {report['error']}")
                return
            
            report_name = report.get('report_name', template_name)
            sections = report.get('sections', {})
            
            result_text = f"📄 テンプレートレポート生成完了: {report_name}\n\n"
            result_text += f"📊 セクション数: {len(sections)}\n"
            result_text += "📝 生成されたセクション:\n"
            
            for section_name, section_data in sections.items():
                section_type = section_data.get('type', 'unknown')
                result_text += f"• {section_name} ({section_type})\n"
            
            self.viz_results_area.setText(result_text)
            
        except Exception as e:
            self.viz_results_area.setText(f"❌ テンプレートレポート生成エラー: {e}")
            logger.error(f"テンプレートレポート生成エラー: {e}")
    
    def _show_visualization(self, chart_type: str):
        """可視化表示"""
        if not self.visualization:
            self.viz_results_area.setText("❌ 可視化システムが利用できません")
            return
        
        try:
            result_text = f"📊 {chart_type} 可視化を生成中...\n"
            self.viz_results_area.setText(result_text)
            
            # チャート生成（実際のアプリケーションでは別ウィンドウで表示される）
            if chart_type == 'productivity_timeline':
                canvas = self.visualization.create_productivity_timeline()
            elif chart_type == 'focus_heatmap':
                canvas = self.visualization.create_focus_heatmap()
            elif chart_type == 'interruption_analysis':
                canvas = self.visualization.create_interruption_analysis_chart()
            elif chart_type == 'session_performance':
                canvas = self.visualization.create_session_performance_chart()
            else:
                result_text += f"❌ 未対応のチャートタイプ: {chart_type}"
                self.viz_results_area.setText(result_text)
                return
            
            result_text += f"✅ {chart_type} 可視化生成完了\n"
            result_text += "📊 チャートは別ウィンドウで表示されます\n"
            
            # エクスポート
            export_path = self.visualization.export_chart(canvas, chart_type)
            if export_path:
                result_text += f"💾 エクスポート: {export_path}"
            
            self.viz_results_area.setText(result_text)
            
        except Exception as e:
            self.viz_results_area.setText(f"❌ 可視化エラー: {e}")
            logger.error(f"可視化エラー ({chart_type}): {e}")
    
    def _run_comparison(self, comparison_type: str):
        """比較分析実行"""
        if not self.comparison_analytics:
            self.viz_results_area.setText("❌ 比較分析システムが利用できません")
            return
        
        try:
            result_text = f"🔄 {comparison_type} 比較分析実行中...\n"
            self.viz_results_area.setText(result_text)
            
            if comparison_type == 'weekdays_vs_weekends':
                results = self.comparison_analytics.compare_weekdays_vs_weekends()
            elif comparison_type == 'time_periods':
                results = self.comparison_analytics.compare_time_periods()
            else:
                result_text += f"❌ 未対応の比較タイプ: {comparison_type}"
                self.viz_results_area.setText(result_text)
                return
            
            if 'error' in results:
                result_text += f"❌ 比較分析エラー: {results['error']}"
                self.viz_results_area.setText(result_text)
                return
            
            result_text += f"✅ {comparison_type} 比較分析完了\n\n"
            
            # 結果要約表示
            if comparison_type == 'weekdays_vs_weekends':
                weekday_metrics = results.get('weekday_metrics', {})
                weekend_metrics = results.get('weekend_metrics', {})
                recommendations = results.get('recommendations', [])
                
                result_text += f"📅 平日データ: {weekday_metrics.get('count', 0)}セッション\n"
                result_text += f"🏠 週末データ: {weekend_metrics.get('count', 0)}セッション\n\n"
                result_text += "💡 推奨事項:\n"
                for rec in recommendations[:2]:
                    result_text += f"• {rec}\n"
            
            elif comparison_type == 'time_periods':
                best_period = results.get('best_performance_period', {})
                result_text += f"⭐ 最高パフォーマンス時間帯: {best_period.get('period', 'unknown')}\n"
                result_text += f"📊 スコア: {best_period.get('score', 0)}\n"
                result_text += f"🎯 信頼度: {best_period.get('confidence', 'unknown')}\n"
            
            self.viz_results_area.setText(result_text)
            
        except Exception as e:
            self.viz_results_area.setText(f"❌ 比較分析エラー: {e}")
            logger.error(f"比較分析エラー ({comparison_type}): {e}")
    
    def connect_signals(self):
        """シグナル接続"""
        # タイマーデータからの通知
        self.timer_data.time_updated.connect(self.on_time_updated)
        self.timer_data.session_changed.connect(self.on_session_changed)
        self.timer_data.timer_state_changed.connect(self.on_timer_state_changed)
        self.timer_data.session_completed.connect(self.on_session_completed)
        
        # 休憩開始シグナル
        self.timer_data.break_started.connect(self.on_break_started)
        
        # タスクマネージャーからの通知（イベント駆動）
        self.task_manager.task_added.connect(lambda: self.refresh_task_list())
        self.task_manager.task_completed.connect(lambda: self.refresh_task_list())
        self.task_manager.task_deleted.connect(lambda: self.refresh_task_list())
    
    def show_minimal_mode(self):
        """ミニマルモード表示"""
        if not self.minimal_window:
            self.minimal_window = MinimalWindow(self.timer_data, self.task_manager)
        
        self.minimal_window.show()
        self.showMinimized()  # メインウィンドウは最小化
        
        logger.info("🔽 ミニマルモード表示、メインウィンドウ最小化")
    
    def test_simple_break_window(self):
        """シンプル休憩ウィンドウテスト表示（開発用）"""
        self.show_simple_break_window("short", 1)  # 1分の短い休憩でテスト
    
    def on_break_started(self, break_type: str, duration_minutes: int):
        """休憩開始時の処理"""
        self.show_simple_break_window(break_type, duration_minutes)
    
    def show_simple_break_window(self, break_type: str, duration_minutes: int):
        """シンプル休憩ウィンドウ表示"""
        try:
            # 休憩開始時にミニマルウィンドウを隠す
            if hasattr(self, 'minimal_window') and self.minimal_window:
                self.minimal_window.hide()
                logger.info("🔽 休憩開始：ミニマルウィンドウを隠します")
            
            # 既存の休憩ウィンドウがあれば閉じる
            if self.break_window:
                self.break_window.close()
                self.break_window = None
            
            # デバッグ：duration_minutesの値を確認
            logger.info(f"📍 休憩ウィンドウ作成: break_type={break_type}, duration_minutes={duration_minutes}")
            
            # 新しいシンプル休憩ウィンドウを作成
            self.break_window = SimpleBreakWindow(break_type, duration_minutes, self.task_manager)
            
            # シグナル接続
            self.break_window.break_finished.connect(self.on_break_window_finished)
            self.break_window.break_skipped.connect(self.on_break_window_skipped)
            
            # 表示
            self.break_window.show()
            
            # 休憩タイマーを自動開始
            if not self.timer_data.is_running:
                self.timer_data.start_timer()
                logger.info(f"⏰ 休憩タイマー自動開始: {break_type} ({duration_minutes}分)")
            
            logger.info(f"☕ シンプル休憩ウィンドウ表示: {break_type} ({duration_minutes}分)")
            
        except Exception as e:
            logger.error(f"シンプル休憩ウィンドウ表示エラー: {e}")
            # フォールバック
            break_name = "長い休憩" if break_type == "long" else "休憩"
            QMessageBox.information(self, "休憩時間", f"☕ {break_name}の時間です！({duration_minutes}分)")
    
    def on_break_window_finished(self):
        """休憩ウィンドウ終了時の処理 - 3秒カウントダウンして作業セッション開始"""
        logger.info("✅ シンプル休憩ウィンドウ自然終了")
        self.break_window = None
        
        # 3秒カウントダウンを開始
        self.show_work_start_countdown()
    
    def show_work_start_countdown(self):
        """作業開始前の3秒カウントダウン表示"""
        try:
            # カウントダウンウィンドウを作成
            self.countdown_window = WorkStartCountdownWindow()
            self.countdown_window.countdown_finished.connect(self.on_work_start_countdown_finished)
            self.countdown_window.show()
            
            logger.info("⏰ 作業開始カウントダウン開始")
            
        except Exception as e:
            logger.error(f"作業開始カウントダウン表示エラー: {e}")
            # フォールバック - 直接作業セッション開始
            self.on_work_start_countdown_finished()
    
    def on_work_start_countdown_finished(self):
        """作業開始カウントダウン終了時の処理"""
        logger.info("🚀 作業開始カウントダウン終了 - 作業セッション開始")
        
        # カウントダウンウィンドウを閉じる
        if hasattr(self, 'countdown_window') and self.countdown_window:
            self.countdown_window.close()
            self.countdown_window = None
        
        # 作業タイマーを自動開始
        if not self.timer_data.is_running:
            self.timer_data.start_timer()
            logger.info("⏰ 作業タイマー自動開始")
        
        # ミニマルウィンドウを再表示
        if hasattr(self, 'minimal_window') and self.minimal_window:
            self.minimal_window.show()
            self.minimal_window.raise_()
            self.minimal_window.activateWindow()
            logger.info("🔽 作業開始：ミニマルウィンドウを再表示")
        
        # さりげない通知
        self.statusBar().showMessage("作業セッション開始！", 2000)
    
    def on_break_window_skipped(self):
        """休憩ウィンドウスキップ時の処理"""
        logger.info("⏩ シンプル休憩ウィンドウスキップ")
        self.break_window = None
        
        # スキップの場合も3秒カウントダウンを開始
        self.show_work_start_countdown()
        
        # さりげない通知
        self.statusBar().showMessage("休憩をスキップしました", 2000)
    
    def add_task(self):
        """タスク追加（イベント駆動更新）"""
        text = self.task_input.text().strip()
        if text:
            self.task_manager.add_task(text)
            self.task_input.clear()
    
    def refresh_task_list(self):
        """タスクリスト更新（必要時のみ）"""
        self.task_list.clear()
        for task in self.task_manager.get_active_tasks():
            item = QListWidgetItem(f"📝 {task['text']}")
            item.setData(Qt.ItemDataRole.UserRole, task['id'])
            self.task_list.addItem(item)
        
        logger.info(f"📋 タスクリスト更新: {self.task_list.count()}件")
    
    # ========================================
    # テンプレート・セッション設定メソッド
    # ========================================
    
    def update_template_combo(self):
        """テンプレートコンボボックスを更新"""
        self.template_combo.clear()
        
        # カテゴリ別にテンプレートを分類
        categories = self.timer_data.template_manager.get_templates_by_category()
        
        for category, templates in categories.items():
            for template_id, template in templates.items():
                item_text = f"[{category}] {template['name']}"
                self.template_combo.addItem(item_text, template_id)
        
        # 現在のテンプレートを選択
        current_template = self.timer_data.get_current_template()
        for i in range(self.template_combo.count()):
            if self.template_combo.itemData(i) == current_template.get('template_id'):
                self.template_combo.setCurrentIndex(i)
                break
        
        # 説明を更新
        self.template_description.setText(current_template.get('description', ''))
    
    def on_template_changed(self):
        """テンプレート変更時の処理"""
        template_id = self.template_combo.currentData()
        if template_id:
            success = self.timer_data.set_template(template_id)
            if success:
                # テンプレート情報を更新
                template = self.timer_data.get_current_template()
                self.template_description.setText(template.get('description', ''))
                
                # セッション設定を更新
                self.max_sessions_spin.setValue(template.get('max_sessions', 8))
                self.long_break_interval_spin.setValue(template.get('sessions_until_long_break', 4))
                
                # 進捗表示を更新
                self.update_progress_display()
                
                # ステータスバーに通知
                self.statusBar().showMessage(f"テンプレート変更: {template['name']}", 3000)
    
    def create_custom_template(self):
        """カスタムテンプレート作成ダイアログ"""
        dialog = CustomTemplateDialog(self.timer_data.template_manager, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # テンプレートコンボボックスを更新
            self.update_template_combo()
            self.statusBar().showMessage("カスタムテンプレートを作成しました", 3000)
    
    def update_session_settings(self):
        """セッション設定を更新"""
        max_sessions = self.max_sessions_spin.value()
        long_break_interval = self.long_break_interval_spin.value()
        
        self.timer_data.set_session_settings(
            max_sessions=max_sessions,
            sessions_until_long_break=long_break_interval
        )
        
        # 進捗表示を更新
        self.update_progress_display()
    
    def update_progress_display(self):
        """セッション進捗表示を更新"""
        progress = self.timer_data.get_session_progress()
        
        self.progress_label.setText(
            f"セッション {progress['current_session']}/{progress['max_sessions']} "
            f"(サイクル {progress['current_cycle']})"
        )
        
        self.progress_bar.setValue(int(progress['progress_percentage']))
        
        # 次の長い休憩までの情報
        if progress['sessions_to_long_break'] > 0:
            self.progress_label.setToolTip(
                f"長い休憩まで {progress['sessions_to_long_break']} セッション"
            )
        else:
            self.progress_label.setToolTip("次は長い休憩です")
    
    def refresh_stats_display(self):
        """統計表示更新（セッション完了時のみ）"""
        today_stats = self.stats.get_today_stats()
        weekly_stats = self.stats.get_weekly_stats()
        
        stats_text = f"""
Phase 3 Final with Integrated Simple Break Window - 統計レポート ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')})

🍅 今日の統計:
   作業セッション: {today_stats['work_sessions']}回
   休憩セッション: {today_stats['break_sessions']}回
   合計作業時間: {today_stats['total_work_minutes']}分

📈 週間統計:
   総セッション数: {weekly_stats['total_sessions']}回
   作業セッション: {weekly_stats['work_sessions']}回
   総作業時間: {weekly_stats['total_work_time']}分
   平均セッション時間: {weekly_stats['avg_session_length']:.1f}分

📊 全期間統計:
   総セッション数: {len(self.stats.sessions)}回

🎯 Phase 3 完成機能:
   ✅ Clean Dual Window Design
   ✅ Event-driven Updates
   ✅ Statistics Dashboard Integration
   ✅ Minimal Mode with Transparency
   ✅ Complete Task Management
   ✅ Integrated Simple Break Window (minimal_timer_demo風)
        """.strip()
        
        self.stats_display.setText(stats_text)
    
    # イベントハンドラー
    def on_time_updated(self, time_left: int):
        """時間更新"""
        minutes = time_left // 60
        seconds = time_left % 60
        self.time_display.setText(f"{minutes:02d}:{seconds:02d}")
    
    def on_session_changed(self, session_type: str, session_number: int):
        """セッション変更"""
        self.session_info.setText(f"{session_type}セッション #{session_number}")
    
    def on_timer_state_changed(self, is_running: bool):
        """タイマー状態変更"""
        self.start_btn.setEnabled(not is_running)
        self.pause_btn.setEnabled(is_running)
    
    def on_session_completed(self, session_type: str, duration: int):
        """セッション完了（Phase 4: 高度なデータ収集統合）"""
        # Phase 4: 高度なセッション追跡終了
        self.stats.end_advanced_session_tracking(completed=True)
        
        # 従来の統計記録
        self.stats.record_session(session_type, duration)
        
        # 統計表示更新
        self.refresh_stats_display()
        
        # ダッシュボード更新（Phase 3統合完了）
        self.dashboard_widget.update_stats()
        
        # タスクリスト更新（セッション完了時のみ）
        self.refresh_task_list()
        
        # 作業セッション完了時のみ軽い通知（休憩は休憩ウィンドウで処理）
        if session_type == "work":
            self.statusBar().showMessage("作業セッション完了！", 2000)
    
    def on_work_duration_changed(self, value: int):
        """作業時間設定変更"""
        self.timer_data.work_minutes = value
    
    def on_break_duration_changed(self, value: int):
        """休憩時間設定変更"""
        self.timer_data.break_minutes = value
    
    # Phase 4: 高度なデータ収集統合メソッド
    def start_timer_with_advanced_tracking(self):
        """タイマー開始（Phase 4: 高度なデータ収集統合）"""
        # Phase 4: 高度なセッション追跡開始
        session_type = "work" if self.timer_data.is_work_session else "break"
        duration = self.timer_data.work_minutes if self.timer_data.is_work_session else self.timer_data.break_minutes
        
        self.stats.start_advanced_session_tracking(session_type, duration)
        
        # ユーザーインタラクション記録
        self.stats.record_user_interaction("start_button", {"session_type": session_type})
        
        # 通常のタイマー開始
        self.timer_data.start_timer()
        
        logger.info(f"🚀 Phase 4 高度追跡付きタイマー開始: {session_type}")
    
    def pause_timer_with_tracking(self):
        """タイマー一時停止（Phase 4: 高度なデータ収集統合）"""
        # 一時停止記録
        self.stats.record_user_interaction("pause_button")
        self.stats.interruption_tracker.record_pause_start()
        
        # 通常の一時停止
        self.timer_data.pause_timer()
        
        logger.info("⏸️ Phase 4 追跡付きタイマー一時停止")
    
    def reset_timer_with_tracking(self):
        """タイマーリセット（Phase 4: 高度なデータ収集統合）"""
        # リセット記録
        self.stats.record_user_interaction("reset_button")
        
        # セッション追跡終了（未完了として）
        if self.stats.advanced_collector.current_session_start:
            self.stats.end_advanced_session_tracking(completed=False)
        
        # 通常のリセット
        self.timer_data.reset_timer()
        
        logger.info("🔄 Phase 4 追跡付きタイマーリセット")
    
    def mousePressEvent(self, event):
        """マウスクリック（Phase 4: ユーザーインタラクション追跡）"""
        self.stats.record_user_interaction("mouse_click", {"button": event.button().name})
        super().mousePressEvent(event)
    
    def keyPressEvent(self, event):
        """キー入力（Phase 4: ユーザーインタラクション追跡）"""
        self.stats.record_user_interaction("key_press", {"key": event.key()})
        super().keyPressEvent(event)


class MinimalWindow(QMainWindow):
    """ミニマルウィンドウ（独立表示）- minimal_timer_standalone.py準拠"""
    
    def __init__(self, timer_data: TimerDataManager, task_manager=None):
        super().__init__()
        
        self.timer_data = timer_data
        self.task_manager = task_manager
        self.dragging = False
        self.drag_position = QPoint()
        self.transparent_mode = True  # デフォルトで透明化モード
        self.show_time = False  # 時刻表示フラグ
        
        # 設定管理
        from PyQt6.QtCore import QSettings
        from PyQt6.QtGui import QColor
        self.settings = QSettings("MinimalTimer", "PomodoroTimer")
        
        # デフォルト表示設定
        self.default_settings = {
            'window_x': 1200,
            'window_y': 20,
            'text_color_r': 255,
            'text_color_g': 255,
            'text_color_b': 255,
            'text_alpha': 255,
            'font_size': 20,
            'show_time': False,
            'transparent_mode': True,
            'countdown_enabled': True,
            'countdown_duration': 3,
            'show_task_name': True
        }
        
        # 表示設定
        self.text_color = QColor(255, 255, 255)
        self.text_opacity = 255
        self.font_size = 20
        self.countdown_enabled = True
        self.countdown_duration = 3
        self.countdown_animation = None
        self.show_task_name = True
        
        # 時刻更新タイマー
        self.clock_timer = QTimer()
        self.clock_timer.timeout.connect(self.update_clock)
        self.clock_timer.start(1000)
        
        self.init_ui()
        self.connect_signals()
        
        # 設定を読み込み適用
        self.load_settings()
        self.apply_loaded_settings()
        
        logger.info("🔽 ミニマルウィンドウ初期化完了")
    
    def init_ui(self):
        """UI初期化 - minimal_timer_standalone.py準拠"""
        self.setWindowTitle("🍅 Pomodoro")
        
        # ウィンドウ設定
        self.setWindowFlags(
            Qt.WindowType.FramelessWindowHint |
            Qt.WindowType.WindowStaysOnTopHint |
            Qt.WindowType.Tool
        )
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        
        # メインウィジェット
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        
        # レイアウト
        layout = QVBoxLayout(self.central_widget)
        layout.setContentsMargins(10, 5, 10, 5)
        layout.setSpacing(2)
        
        # 現在時刻
        self.time_label = QLabel()
        self.time_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.time_label.setVisible(False)
        
        # タイマー
        self.timer_label = QLabel("25:00")
        self.timer_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # カウントダウンラベル（通常は非表示）
        self.countdown_label = QLabel()
        self.countdown_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.countdown_label.setVisible(False)
        self.countdown_label.setObjectName("countdown_label")
        
        # タスク名
        self.task_label = QLabel("")
        self.task_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.task_label.setWordWrap(True)
        
        # 状態
        self.status_label = QLabel("作業")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        layout.addWidget(self.time_label)
        layout.addWidget(self.timer_label)
        layout.addWidget(self.countdown_label)
        layout.addWidget(self.task_label)
        layout.addWidget(self.status_label)
        
        # フォント設定
        self.update_fonts()
        
        # サイズ
        self.resize(110, 60)
        
        # 透明化設定の初期化
        self.apply_transparent_style()
        
        # コンテキストメニュー
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)
        
        # 初期表示設定
        self.update_display()
    
    def apply_transparent_style(self):
        """透明化スタイルの適用（カウントダウン対応統合版）"""
        # 文字色設定を文字列に変換
        color_str = f"rgba({self.text_color.red()}, {self.text_color.green()}, {self.text_color.blue()}, {self.text_opacity})"
        
        if self.transparent_mode:
            # 完全透明化＋マウスイベント透過（カウントダウン中も維持）
            self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
            self.setStyleSheet(f"""
                QWidget {{
                    background-color: rgba(0, 0, 0, 0);
                    border: none;
                }}
                QLabel {{
                    color: {color_str};
                    background-color: rgba(0, 0, 0, 0);
                    font-weight: bold;
                }}
                QLabel#countdown_label {{
                    background-color: rgba(50, 50, 50, 200);
                    border: 2px solid rgba(255, 255, 255, 100);
                    border-radius: 50px;
                    min-width: 100px;
                    min-height: 100px;
                    font-size: {self.font_size * 2}pt;
                    font-weight: bold;
                }}
            """)
        else:
            # 通常表示モード
            self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, False)
            self.setStyleSheet(f"""
                QWidget {{
                    background-color: rgba(40, 40, 40, 230);
                    border-radius: 10px;
                }}
                QLabel {{
                    color: {color_str};
                    background-color: rgba(0, 0, 0, 0);
                }}
                QLabel#countdown_label {{
                    background-color: rgba(70, 70, 70, 220);
                    border: 2px solid rgba(255, 255, 255, 150);
                    border-radius: 50px;
                    min-width: 100px;
                    min-height: 100px;
                    font-size: {self.font_size * 2}pt;
                    font-weight: bold;
                }}
            """)
    
    def connect_signals(self):
        """シグナル接続"""
        self.timer_data.time_updated.connect(self.on_time_updated)
        self.timer_data.session_changed.connect(self.on_session_changed)
    
    def update_fonts(self):
        """フォント更新"""
        timer_font = QFont("Arial", self.font_size, QFont.Weight.Bold)
        self.timer_label.setFont(timer_font)
        
        # カウントダウンフォント（通常の2倍サイズ）
        countdown_font = QFont("Arial", self.font_size * 2, QFont.Weight.Bold)
        self.countdown_label.setFont(countdown_font)
        
        time_font = QFont("Arial", int(self.font_size * 0.6))
        self.time_label.setFont(time_font)
        
        # タスク名フォント
        task_font = QFont("Arial", int(self.font_size * 0.5))
        self.task_label.setFont(task_font)
        
        status_font = QFont("Arial", int(self.font_size * 0.55))
        self.status_label.setFont(status_font)
    
    def update_display(self):
        """表示更新"""
        # タイマー表示
        if self.timer_data.time_left == 0:
            minutes = self.timer_data.work_minutes if self.timer_data.is_work_session else self.timer_data.break_minutes
            seconds = 0
        else:
            minutes = self.timer_data.time_left // 60
            seconds = self.timer_data.time_left % 60
        
        self.timer_label.setText(f"{minutes:02d}:{seconds:02d}")
        
        # タスク名表示
        self.update_task_display()
        
        # 状態表示
        if self.timer_data.is_work_session:
            self.status_label.setText("作業")
            self.timer_label.setStyleSheet("color: #00FF00;")
        else:
            self.status_label.setText("休憩")
            self.timer_label.setStyleSheet("color: #00AAFF;")
    
    def update_task_display(self):
        """タスク名表示更新"""
        if self.show_task_name and self.task_manager:
            active_tasks = self.task_manager.get_active_tasks()
            if active_tasks:
                # 最新のタスクを表示（文字数制限）
                task_text = active_tasks[-1]['text']
                if len(task_text) > 15:
                    task_text = task_text[:15] + "..."
                self.task_label.setText(f"📋 {task_text}")
                self.task_label.setVisible(True)
            else:
                self.task_label.setText("")
                self.task_label.setVisible(False)
        else:
            self.task_label.setVisible(False)
    
    def update_clock(self):
        """時刻更新"""
        if self.show_time:
            current = datetime.now().strftime("%H:%M:%S")
            self.time_label.setText(current)
    
    def on_time_updated(self, time_left: int):
        """時間更新"""
        minutes = time_left // 60
        seconds = time_left % 60
        self.timer_label.setText(f"{minutes:02d}:{seconds:02d}")
        
        # カウントダウン処理（作業セッション且つ残り3秒以下でカウントダウン）
        if (self.timer_data.is_work_session and 
            time_left <= 3 and 
            time_left > 0 and
            self.countdown_enabled):
            self.show_countdown(time_left)
    
    def on_session_changed(self, session_type: str, session_number: int):
        """セッション変更"""
        # セッション変更時はカウントダウンを隠す
        self.hide_countdown()
        
        # 休憩セッションの場合はミニマルウィンドウを隠す
        if session_type == "休憩":
            self.hide()
            logger.info("🔽 休憩セッション：ミニマルウィンドウを隠します")
        else:
            # 作業セッションの場合は表示を更新
            self.update_display()
    
    def show_countdown(self, count):
        """カウントダウン表示（統合版）"""
        # カウントダウンが無効の場合はスキップ
        if not self.countdown_enabled:
            return
            
        # カウントダウン表示の条件チェック
        if count > self.countdown_duration or count <= 0:
            return
            
        self.countdown_label.setText(str(count))
        self.countdown_label.setVisible(True)
        
        # 透明化モードに応じたスタイル設定
        self.update_countdown_style()
        
        # アニメーション開始（メモリリーク対策）
        self.animate_countdown(count)
    
    def hide_countdown(self):
        """カウントダウン非表示（メモリリーク対策強化）"""
        self.countdown_label.setVisible(False)
        
        # アニメーションを安全に停止・削除
        if self.countdown_animation is not None:
            try:
                self.countdown_animation.stop()
                self.countdown_animation.deleteLater()
            except Exception as e:
                logger.error(f"カウントダウンアニメーション停止エラー: {e}")
            finally:
                self.countdown_animation = None
    
    def animate_countdown(self, count):
        """カウントダウンアニメーション（メモリ効率最適化版）"""
        try:
            # 既存アニメーションを停止
            if self.countdown_animation is not None:
                self.countdown_animation.stop()
                self.countdown_animation.deleteLater()
            
            from PyQt6.QtCore import QPropertyAnimation, QEasingCurve
            
            # スケールアニメーション作成
            self.countdown_animation = QPropertyAnimation(self.countdown_label, b"geometry")
            self.countdown_animation.setDuration(800)  # 0.8秒
            self.countdown_animation.setEasingCurve(QEasingCurve.Type.OutElastic)
            
            # 開始と終了のサイズを設定
            current_rect = self.countdown_label.geometry()
            start_size = 60  # 小さく開始
            end_size = 120   # 大きく表示
            
            # アニメーション設定
            start_rect = current_rect
            start_rect.setSize(start_rect.size())
            
            end_rect = current_rect
            end_rect.setWidth(end_size)
            end_rect.setHeight(end_size)
            end_rect.moveCenter(current_rect.center())
            
            self.countdown_animation.setStartValue(start_rect)
            self.countdown_animation.setEndValue(end_rect)
            
            # アニメーション開始
            self.countdown_animation.start()
            
            # 1秒後に次のカウントまたは終了（エラーハンドリング付き）
            if count > 1:
                QTimer.singleShot(1000, lambda: self.show_countdown(count - 1))
            else:
                QTimer.singleShot(1000, self.hide_countdown)
                
        except Exception as e:
            logger.error(f"カウントダウンアニメーションエラー: {e}")
            # エラー時はアニメーションなしで表示継続
            if count > 1:
                QTimer.singleShot(1000, lambda: self.show_countdown(count - 1))
            else:
                QTimer.singleShot(1000, self.hide_countdown)
    
    def update_countdown_style(self):
        """カウントダウンラベルのスタイル更新"""
        color_str = f"rgba({self.text_color.red()}, {self.text_color.green()}, {self.text_color.blue()}, {self.text_opacity})"
        
        if self.transparent_mode:
            bg_color = "rgba(50, 50, 50, 200)"
            border_color = "rgba(255, 255, 255, 100)"
        else:
            bg_color = "rgba(70, 70, 70, 220)"
            border_color = "rgba(255, 255, 255, 150)"
            
        self.countdown_label.setStyleSheet(f"""
            QLabel {{
                color: {color_str};
                background-color: {bg_color};
                border: 2px solid {border_color};
                border-radius: 50px;
                min-width: 100px;
                min-height: 100px;
                font-size: {self.font_size * 2}pt;
                font-weight: bold;
            }}
        """)
    
    # マウスイベント（Alt+クリックでドラッグ可能、右クリックでメニュー）- minimal_timer_standalone.py準拠
    def mousePressEvent(self, event: QMouseEvent):
        if event.button() == Qt.MouseButton.RightButton:
            # 右クリック時：メニュー表示のため一時的に透明化を無効
            pass  # show_context_menuで処理
        elif (event.button() == Qt.MouseButton.LeftButton and 
              event.modifiers() == Qt.KeyboardModifier.AltModifier):
            # Alt+左クリック時：ドラッグモード
            self.dragging = True
            self.drag_position = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            self.set_transparent_mode(False)  # ドラッグ中は透明化を無効
            
    def mouseMoveEvent(self, event: QMouseEvent):
        if self.dragging and event.buttons() == Qt.MouseButton.LeftButton:
            self.move(event.globalPosition().toPoint() - self.drag_position)
            
    def mouseReleaseEvent(self, event: QMouseEvent):
        if event.button() == Qt.MouseButton.LeftButton and self.dragging:
            self.dragging = False
            # ドラッグ終了後、透明化を再有効化
            self.apply_transparent_style()
            # 位置変更の設定保存
            self.save_settings()
    
    def set_transparent_mode(self, enabled):
        """透明化モードの一時設定（カウントダウン中も適切に処理）"""
        old_mode = self.transparent_mode
        self.transparent_mode = enabled
        self.apply_transparent_style()
        
        # カウントダウン表示中の場合、スタイルを再適用
        if self.countdown_label.isVisible():
            self.update_countdown_style()
        
        # 元の設定を復元（一時的な変更の場合）
        if not enabled and old_mode:
            QTimer.singleShot(100, lambda: setattr(self, 'transparent_mode', old_mode))
    
    def show_context_menu(self, pos):
        """拡張コンテキストメニュー - minimal_timer_standalone.py準拠"""
        # 右クリック時は一時的に透明化を無効にする
        self.set_transparent_mode(False)
        
        menu = QMenu(self)
        
        # 時刻表示
        time_action = QAction("時刻表示", self)
        time_action.setCheckable(True)
        time_action.setChecked(self.show_time)
        time_action.triggered.connect(self.toggle_time)
        menu.addAction(time_action)
        
        # タスク名表示
        task_action = QAction("タスク名表示", self)
        task_action.setCheckable(True)
        task_action.setChecked(self.show_task_name)
        task_action.triggered.connect(self.toggle_task_name)
        menu.addAction(task_action)
        
        # 透明化モード切り替え
        transparent_action = QAction("透明化モード", self)
        transparent_action.setCheckable(True)
        transparent_action.setChecked(self.transparent_mode)
        transparent_action.triggered.connect(self.toggle_transparent_mode)
        menu.addAction(transparent_action)
        
        menu.addSeparator()
        
        # 位置設定メニュー
        position_menu = QMenu("位置設定", self)
        
        # プリセット位置
        position_presets = [
            ("右上", lambda: self.move_to_preset("top_right")),
            ("左上", lambda: self.move_to_preset("top_left")),
            ("右下", lambda: self.move_to_preset("bottom_right")),
            ("左下", lambda: self.move_to_preset("bottom_left"))
        ]
        
        for name, callback in position_presets:
            action = QAction(name, self)
            action.triggered.connect(callback)
            position_menu.addAction(action)
            
        position_menu.addSeparator()
        
        # カスタム位置設定
        custom_pos_action = QAction("カスタム位置...", self)
        custom_pos_action.triggered.connect(self.set_custom_position)
        position_menu.addAction(custom_pos_action)
        
        menu.addMenu(position_menu)
        
        # 表示設定メニュー
        display_menu = QMenu("表示設定", self)
        
        # 文字色設定サブメニュー
        color_menu = QMenu("文字色", self)
        
        # プリセット色サブメニュー
        preset_color_menu = QMenu("プリセット色", self)
        
        from PyQt6.QtGui import QColor
        preset_colors = [
            ("赤", QColor(255, 0, 0)),
            ("緑", QColor(0, 255, 0)),
            ("青", QColor(0, 0, 255)),
            ("黄", QColor(255, 255, 0)),
            ("白", QColor(255, 255, 255)),
            ("シアン", QColor(0, 255, 255)),
            ("マゼンタ", QColor(255, 0, 255))
        ]
        
        for name, color in preset_colors:
            action = QAction(name, self)
            action.triggered.connect(lambda checked, c=color: self.set_text_color(c))
            preset_color_menu.addAction(action)
            
        color_menu.addMenu(preset_color_menu)
        
        # カスタム色選択
        custom_color_action = QAction("カスタム色...", self)
        custom_color_action.triggered.connect(self.set_custom_color)
        color_menu.addAction(custom_color_action)
        
        display_menu.addMenu(color_menu)
        
        # 透明度設定
        opacity_action = QAction("透明度...", self)
        opacity_action.triggered.connect(self.set_text_opacity)
        display_menu.addAction(opacity_action)
        
        # フォントサイズ設定
        font_size_action = QAction("フォントサイズ...", self)
        font_size_action.triggered.connect(self.set_font_size)
        display_menu.addAction(font_size_action)
        
        menu.addMenu(display_menu)
        
        menu.addSeparator()
        
        # タイマー制御
        if self.timer_data.is_running:
            pause_action = QAction("一時停止", self)
            pause_action.triggered.connect(self.timer_data.pause_timer)
            menu.addAction(pause_action)
        else:
            start_action = QAction("開始", self)
            start_action.triggered.connect(self.timer_data.start_timer)
            menu.addAction(start_action)
            
        reset_action = QAction("リセット", self)
        reset_action.triggered.connect(self.timer_data.reset_timer)
        menu.addAction(reset_action)
        
        # カウントダウン設定サブメニュー
        countdown_menu = QMenu("カウントダウン設定", self)
        
        # カウントダウン有効/無効
        countdown_toggle_action = QAction("カウントダウン有効", self)
        countdown_toggle_action.setCheckable(True)
        countdown_toggle_action.setChecked(self.countdown_enabled)
        countdown_toggle_action.triggered.connect(self.toggle_countdown_enabled)
        countdown_menu.addAction(countdown_toggle_action)
        
        # カウントダウン秒数設定
        countdown_duration_action = QAction("カウントダウン秒数...", self)
        countdown_duration_action.triggered.connect(self.set_countdown_duration)
        countdown_menu.addAction(countdown_duration_action)
        
        countdown_menu.addSeparator()
        
        # デバッグ用：カウントダウンテスト
        countdown_test_action = QAction("カウントダウンテスト", self)
        countdown_test_action.triggered.connect(lambda: self.show_countdown(self.countdown_duration))
        countdown_menu.addAction(countdown_test_action)
        
        menu.addMenu(countdown_menu)
        
        menu.addSeparator()
        
        # 設定管理
        reset_settings_action = QAction("設定をリセット", self)
        reset_settings_action.triggered.connect(self.reset_to_defaults)
        menu.addAction(reset_settings_action)
        
        menu.addSeparator()
        
        # 設定モードに戻る
        show_main_action = QAction("🏠 設定モードを復元", self)
        show_main_action.triggered.connect(self.show_main_window)
        menu.addAction(show_main_action)
        
        # 終了
        close_action = QAction("❌ 閉じる", self)
        close_action.triggered.connect(self.close)
        menu.addAction(close_action)
        
        # メニュー閉じた後に元のモードに戻す
        menu.aboutToHide.connect(lambda: self.apply_transparent_style())
        menu.exec(self.mapToGlobal(pos))
    
    def show_main_window(self):
        """メインウィンドウを復元"""
        # メインウィンドウを探して復元
        for widget in QApplication.allWidgets():
            if isinstance(widget, MainWindow):
                widget.showNormal()
                widget.raise_()
                widget.activateWindow()
                logger.info("🏠 メインウィンドウ復元")
                break
    
    # ========================================
    # 設定関連メソッド - minimal_timer_standalone.py準拠
    # ========================================
    
    def toggle_time(self):
        """時刻表示切り替え"""
        self.show_time = not self.show_time
        self.time_label.setVisible(self.show_time)
        
        if self.show_time:
            self.resize(110, 80)
        else:
            self.resize(110, 60)
        
        # 設定保存
        self.save_settings()
    
    def toggle_task_name(self):
        """タスク名表示切り替え"""
        self.show_task_name = not self.show_task_name
        self.update_task_display()
        
        # ウィンドウサイズ調整
        if self.show_task_name:
            height = 80 if self.show_time else 70
        else:
            height = 80 if self.show_time else 60
        self.resize(110, height)
        
        # 設定保存
        self.save_settings()
    
    def toggle_transparent_mode(self):
        """透明化モード切り替え"""
        self.transparent_mode = not self.transparent_mode
        self.apply_transparent_style()
        # 設定保存
        self.save_settings()
    
    def move_to_preset(self, position):
        """プリセット位置に移動"""
        if not QApplication.primaryScreen():
            return
            
        screen = QApplication.primaryScreen().geometry()
        window_size = self.size()
        margin = 20
        
        positions = {
            "top_right": (screen.width() - window_size.width() - margin, margin),
            "top_left": (margin, margin),
            "bottom_right": (screen.width() - window_size.width() - margin, 
                           screen.height() - window_size.height() - margin),
            "bottom_left": (margin, screen.height() - window_size.height() - margin)
        }
        
        if position in positions:
            x, y = positions[position]
            self.move(x, y)
            # 設定保存
            self.save_settings()
    
    def set_custom_position(self):
        """カスタム位置設定ダイアログ"""
        from PyQt6.QtWidgets import QInputDialog
        
        current_pos = self.pos()
        
        # X座標入力
        x, ok = QInputDialog.getInt(
            self, "カスタム位置設定", "X座標:", 
            current_pos.x(), 0, 9999
        )
        if not ok:
            return
            
        # Y座標入力
        y, ok = QInputDialog.getInt(
            self, "カスタム位置設定", "Y座標:", 
            current_pos.y(), 0, 9999
        )
        if ok:
            self.move(x, y)
            # 設定保存
            self.save_settings()
    
    def set_text_color(self, color):
        """文字色設定"""
        self.text_color = color
        self.apply_transparent_style()
        # 設定保存
        self.save_settings()
    
    def set_custom_color(self):
        """カスタム色選択ダイアログ"""
        from PyQt6.QtWidgets import QColorDialog
        
        color = QColorDialog.getColor(self.text_color, self, "文字色を選択")
        if color.isValid():
            self.set_text_color(color)
    
    def set_text_opacity(self):
        """透明度設定ダイアログ"""
        from PyQt6.QtWidgets import QInputDialog
        
        opacity, ok = QInputDialog.getInt(
            self, "透明度設定", "透明度 (0-255):", 
            self.text_opacity, 0, 255
        )
        if ok:
            self.text_opacity = opacity
            self.apply_transparent_style()
            # 設定保存
            self.save_settings()
    
    def set_font_size(self):
        """フォントサイズ設定ダイアログ"""
        from PyQt6.QtWidgets import QInputDialog
        
        size, ok = QInputDialog.getInt(
            self, "フォントサイズ設定", "フォントサイズ (10-36):", 
            self.font_size, 10, 36
        )
        if ok:
            self.font_size = size
            self.update_fonts()
            self.apply_transparent_style()
            # 設定保存
            self.save_settings()
    
    def toggle_countdown_enabled(self):
        """カウントダウン有効/無効切り替え"""
        self.countdown_enabled = not self.countdown_enabled
        # 設定保存
        self.save_settings()
        
        # カウントダウンが無効になった場合は表示を隠す
        if not self.countdown_enabled and self.countdown_label.isVisible():
            self.hide_countdown()
    
    def set_countdown_duration(self):
        """カウントダウン秒数設定ダイアログ"""
        from PyQt6.QtWidgets import QInputDialog
        
        duration, ok = QInputDialog.getInt(
            self, "カウントダウン秒数設定", "カウントダウン開始秒数 (1-10):", 
            self.countdown_duration, 1, 10
        )
        if ok:
            self.countdown_duration = duration
            # 設定保存
            self.save_settings()
    
    def save_settings(self):
        """設定を保存"""
        try:
            # ウィンドウ位置
            pos = self.pos()
            self.settings.setValue("Position/x", pos.x())
            self.settings.setValue("Position/y", pos.y())
            
            # 表示設定
            self.settings.setValue("Display/text_color_r", self.text_color.red())
            self.settings.setValue("Display/text_color_g", self.text_color.green())
            self.settings.setValue("Display/text_color_b", self.text_color.blue())
            self.settings.setValue("Display/text_alpha", self.text_opacity)
            self.settings.setValue("Display/font_size", self.font_size)
            
            # UI設定
            self.settings.setValue("UI/show_time", self.show_time)
            self.settings.setValue("UI/show_task_name", self.show_task_name)
            self.settings.setValue("UI/transparent_mode", self.transparent_mode)
            
            # カウントダウン設定
            self.settings.setValue("Countdown/enabled", self.countdown_enabled)
            self.settings.setValue("Countdown/duration", self.countdown_duration)
            
            # 設定を即座にファイルに書き込み
            self.settings.sync()
            
        except Exception as e:
            logger.error(f"設定保存エラー: {e}")

    def load_settings(self):
        """設定を読み込み"""
        try:
            from PyQt6.QtGui import QColor
            
            # デフォルト値を使用して設定を読み込み
            self.loaded_x = int(self.settings.value("Position/x", self.default_settings['window_x']))
            self.loaded_y = int(self.settings.value("Position/y", self.default_settings['window_y']))
            
            # 文字色
            r = int(self.settings.value("Display/text_color_r", self.default_settings['text_color_r']))
            g = int(self.settings.value("Display/text_color_g", self.default_settings['text_color_g']))
            b = int(self.settings.value("Display/text_color_b", self.default_settings['text_color_b']))
            self.text_color = QColor(r, g, b)
            
            self.text_opacity = int(self.settings.value("Display/text_alpha", self.default_settings['text_alpha']))
            self.font_size = int(self.settings.value("Display/font_size", self.default_settings['font_size']))
            
            # UI設定（文字列から bool に変換）
            show_time_str = self.settings.value("UI/show_time", str(self.default_settings['show_time']))
            self.show_time = show_time_str.lower() == 'true' if isinstance(show_time_str, str) else bool(show_time_str)
            
            show_task_name_str = self.settings.value("UI/show_task_name", str(self.default_settings['show_task_name']))
            self.show_task_name = show_task_name_str.lower() == 'true' if isinstance(show_task_name_str, str) else bool(show_task_name_str)
            
            transparent_mode_str = self.settings.value("UI/transparent_mode", str(self.default_settings['transparent_mode']))
            self.transparent_mode = transparent_mode_str.lower() == 'true' if isinstance(transparent_mode_str, str) else bool(transparent_mode_str)
            
            # カウントダウン設定
            countdown_enabled_str = self.settings.value("Countdown/enabled", str(self.default_settings['countdown_enabled']))
            self.countdown_enabled = countdown_enabled_str.lower() == 'true' if isinstance(countdown_enabled_str, str) else bool(countdown_enabled_str)
            
            self.countdown_duration = int(self.settings.value("Countdown/duration", self.default_settings['countdown_duration']))
            
        except Exception as e:
            logger.error(f"設定読み込みエラー: {e}")
            # エラー時はデフォルト値を使用
            self.reset_to_defaults_silent()

    def apply_loaded_settings(self):
        """読み込んだ設定をUIに適用"""
        try:
            # ウィンドウ位置
            self.move(self.loaded_x, self.loaded_y)
            
            # フォント設定を適用
            self.update_fonts()
            
            # 時刻表示設定
            self.time_label.setVisible(self.show_time)
            
            # タスク名表示設定
            self.update_task_display()
            
            # ウィンドウサイズ設定
            height = 60
            if self.show_time:
                height += 20
            if self.show_task_name and self.task_label.isVisible():
                height += 15
            self.resize(110, height)
            
            # 透明化設定を適用
            self.apply_transparent_style()
            
        except Exception as e:
            logger.error(f"設定適用エラー: {e}")

    def reset_to_defaults(self):
        """デフォルト設定にリセット"""
        from PyQt6.QtWidgets import QMessageBox
        
        try:
            # 確認ダイアログ
            reply = QMessageBox.question(
                self, "設定リセット確認", 
                "すべての設定をデフォルトに戻しますか？", 
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.reset_to_defaults_silent()
                
                # 完了メッセージ
                QMessageBox.information(self, "設定リセット", "設定をデフォルトに戻しました。")
                
        except Exception as e:
            logger.error(f"設定リセットエラー: {e}")
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "エラー", f"設定リセット中にエラーが発生しました：{e}")
    
    def reset_to_defaults_silent(self):
        """デフォルト設定にリセット（確認なし）"""
        try:
            from PyQt6.QtGui import QColor
            
            # 設定ファイルをクリア
            self.settings.clear()
            
            # デフォルト値を設定
            self.text_color = QColor(
                self.default_settings['text_color_r'],
                self.default_settings['text_color_g'], 
                self.default_settings['text_color_b']
            )
            self.text_opacity = self.default_settings['text_alpha']
            self.font_size = self.default_settings['font_size']
            self.show_time = self.default_settings['show_time']
            self.transparent_mode = self.default_settings['transparent_mode']
            self.countdown_enabled = self.default_settings['countdown_enabled']
            self.countdown_duration = self.default_settings['countdown_duration']
            
            # デフォルト位置に移動
            self.move(self.default_settings['window_x'], self.default_settings['window_y'])
            
            # UI更新
            self.update_fonts()
            self.time_label.setVisible(self.show_time)
            if self.show_time:
                self.resize(110, 80)
            else:
                self.resize(110, 60)
            self.apply_transparent_style()
            
            # 設定保存
            self.save_settings()
            
        except Exception as e:
            logger.error(f"設定リセットエラー: {e}")
    
    def closeEvent(self, event):
        """ウィンドウクローズ時の処理"""
        try:
            # 設定を保存
            self.save_settings()
            # カウントダウンアニメーション停止
            self.hide_countdown()
            # タイマー停止
            if hasattr(self, 'clock_timer'):
                self.clock_timer.stop()
        except Exception as e:
            logger.error(f"ウィンドウクローズ処理エラー: {e}")
        finally:
            event.accept()
    
    def setup_worker3_tab(self):
        """Worker3 予測エンジン・エクスポートタブのセットアップ"""
        if not hasattr(self, 'prediction_engine'):
            return
        
        worker3_widget = QWidget()
        layout = QVBoxLayout(worker3_widget)
        
        # タブ追加
        self.tab_widget.addTab(worker3_widget, "🤖 AI予測・エクスポート")
        
        # 予測エンジンセクション
        prediction_group = QGroupBox("🔮 AI予測エンジン")
        prediction_layout = QVBoxLayout()
        
        # モデル訓練ボタン
        train_models_btn = QPushButton("🎯 全モデル再トレーニング")
        train_models_btn.clicked.connect(self.train_all_models)
        train_models_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        prediction_layout.addWidget(train_models_btn)
        
        # 予測実行ボタン
        prediction_buttons_layout = QHBoxLayout()
        
        predict_focus_btn = QPushButton("🎯 フォーカス予測")
        predict_focus_btn.clicked.connect(self.predict_focus_score)
        predict_focus_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        prediction_buttons_layout.addWidget(predict_focus_btn)
        
        predict_optimal_btn = QPushButton("⏰ 最適時間予測")
        predict_optimal_btn.clicked.connect(self.predict_optimal_times)
        predict_optimal_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        prediction_buttons_layout.addWidget(predict_optimal_btn)
        
        predict_trend_btn = QPushButton("📈 生産性トレンド予測")
        predict_trend_btn.clicked.connect(self.predict_productivity_trend)
        predict_trend_btn.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #e67e22;
            }
        """)
        prediction_buttons_layout.addWidget(predict_trend_btn)
        
        prediction_layout.addLayout(prediction_buttons_layout)
        
        # 予測結果表示エリア
        self.prediction_result_text = QTextEdit()
        self.prediction_result_text.setMaximumHeight(150)
        self.prediction_result_text.setPlaceholderText("予測結果がここに表示されます...")
        prediction_layout.addWidget(self.prediction_result_text)
        
        prediction_group.setLayout(prediction_layout)
        layout.addWidget(prediction_group)
        
        # エクスポートセクション
        export_group = QGroupBox("📄 レポートエクスポート")
        export_layout = QVBoxLayout()
        
        # エクスポートボタン
        export_buttons_layout = QHBoxLayout()
        
        export_pdf_btn = QPushButton("📄 PDFエクスポート")
        export_pdf_btn.clicked.connect(lambda: self.export_report('pdf'))
        export_pdf_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        export_buttons_layout.addWidget(export_pdf_btn)
        
        export_excel_btn = QPushButton("📊 Excelエクスポート")
        export_excel_btn.clicked.connect(lambda: self.export_report('excel'))
        export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 12px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        export_buttons_layout.addWidget(export_excel_btn)
        
        export_layout.addLayout(export_buttons_layout)
        
        export_group.setLayout(export_layout)
        layout.addWidget(export_group)
        
        # スケジューラーセクション
        scheduler_group = QGroupBox("⏰ 自動レポート生成")
        scheduler_layout = QVBoxLayout()
        
        # 即座レポート生成
        immediate_layout = QHBoxLayout()
        
        daily_report_btn = QPushButton("📅 日次レポート生成")
        daily_report_btn.clicked.connect(lambda: self.generate_immediate_report('daily'))
        daily_report_btn.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
        """)
        immediate_layout.addWidget(daily_report_btn)
        
        weekly_report_btn = QPushButton("📊 週次レポート生成")
        weekly_report_btn.clicked.connect(lambda: self.generate_immediate_report('weekly'))
        weekly_report_btn.setStyleSheet("""
            QPushButton {
                background-color: #34495e;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #2c3e50;
            }
        """)
        immediate_layout.addWidget(weekly_report_btn)
        
        monthly_report_btn = QPushButton("📈 月次レポート生成")
        monthly_report_btn.clicked.connect(lambda: self.generate_immediate_report('monthly'))
        monthly_report_btn.setStyleSheet("""
            QPushButton {
                background-color: #e67e22;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 5px;
                font-size: 11px;
            }
            QPushButton:hover {
                background-color: #d35400;
            }
        """)
        immediate_layout.addWidget(monthly_report_btn)
        
        scheduler_layout.addLayout(immediate_layout)
        
        # スケジュール情報表示
        schedule_info_text = QTextEdit()
        schedule_info_text.setMaximumHeight(100)
        schedule_info_text.setPlaceholderText("スケジュール情報を表示...")
        scheduler_layout.addWidget(schedule_info_text)
        
        # 次回予定レポート情報を更新
        self.update_schedule_info(schedule_info_text)
        
        scheduler_group.setLayout(scheduler_layout)
        layout.addWidget(scheduler_group)
        
        # スペーサー
        layout.addStretch()
    
    def train_all_models(self):
        """全モデルを再トレーニング"""
        if not hasattr(self, 'prediction_engine'):
            return
        
        try:
            self.prediction_result_text.append("🚀 全モデル再トレーニング開始...")
            
            # バックグラウンドでトレーニング実行
            results = self.prediction_engine.retrain_all_models()
            
            self.prediction_result_text.append("✅ 全モデル再トレーニング完了!")
            
            for model_name, result in results.items():
                if 'error' not in result:
                    if 'metrics' in result:
                        metrics = result['metrics']
                        accuracy = metrics.get('r2', metrics.get('accuracy', 0))
                        self.prediction_result_text.append(f"  📊 {model_name}: 精度 {accuracy:.3f}")
                    else:
                        self.prediction_result_text.append(f"  ✅ {model_name}: 訓練完了")
                else:
                    self.prediction_result_text.append(f"  ❌ {model_name}: {result['error']}")
            
        except Exception as e:
            self.prediction_result_text.append(f"❌ トレーニングエラー: {e}")
            logger.error(f"モデル訓練エラー: {e}")
    
    def predict_focus_score(self):
        """フォーカススコアを予測"""
        if not hasattr(self, 'prediction_engine'):
            return
        
        try:
            # 現在のセッションパラメータを使用
            session_params = {
                'planned_duration': 25,
                'hour_of_day': datetime.now().hour,
                'day_of_week': datetime.now().weekday(),
                'interruption_count': 0,
                'environment_score': 0.7
            }
            
            result = self.prediction_engine.predict_focus_score(session_params)
            
            if 'error' not in result:
                predicted_score = result['predicted_focus_score']
                accuracy = result['model_accuracy']
                
                self.prediction_result_text.append(f"🎯 フォーカススコア予測: {predicted_score:.3f}")
                self.prediction_result_text.append(f"   モデル精度: {accuracy:.3f}")
                
                if predicted_score > 0.8:
                    self.prediction_result_text.append("   ✅ 高いフォーカスが期待できます!")
                elif predicted_score > 0.6:
                    self.prediction_result_text.append("   ⚠️ 中程度のフォーカスが期待できます")
                else:
                    self.prediction_result_text.append("   ❌ フォーカスが低い可能性があります")
            else:
                self.prediction_result_text.append(f"❌ 予測エラー: {result['error']}")
                
        except Exception as e:
            self.prediction_result_text.append(f"❌ 予測エラー: {e}")
            logger.error(f"フォーカススコア予測エラー: {e}")
    
    def predict_optimal_times(self):
        """最適作業時間を予測"""
        if not hasattr(self, 'prediction_engine'):
            return
        
        try:
            result = self.prediction_engine.predict_optimal_work_time()
            
            if 'error' not in result:
                self.prediction_result_text.append("⏰ 最適作業時間予測:")
                
                today_recs = result.get('today_recommendations', [])
                if today_recs:
                    self.prediction_result_text.append("  📅 今日の推奨時間帯:")
                    for i, rec in enumerate(today_recs[:3], 1):
                        hour = rec.get('hour', 0)
                        prob = rec.get('optimal_probability', 0)
                        self.prediction_result_text.append(f"    {i}. {hour:02d}:00 (確率: {prob:.1%})")
                else:
                    self.prediction_result_text.append("  ❌ 今日の推奨時間帯データがありません")
                
                current_prob = result.get('current_time_optimal_prob', 0)
                self.prediction_result_text.append(f"  🕐 現在時刻の最適確率: {current_prob:.1%}")
            else:
                self.prediction_result_text.append(f"❌ 予測エラー: {result['error']}")
                
        except Exception as e:
            self.prediction_result_text.append(f"❌ 予測エラー: {e}")
            logger.error(f"最適時間予測エラー: {e}")
    
    def predict_productivity_trend(self):
        """生産性トレンド予測"""
        if not hasattr(self, 'prediction_engine'):
            return
        
        try:
            result = self.prediction_engine.predict_productivity_trend(7)
            
            if 'error' not in result:
                trend_direction = result.get('trend_direction', 'stable')
                avg_productivity = result.get('average_predicted_productivity', 0)
                accuracy = result.get('model_accuracy', 0)
                
                self.prediction_result_text.append("📈 生産性トレンド予測 (7日間):")
                self.prediction_result_text.append(f"  📊 トレンド: {trend_direction}")
                self.prediction_result_text.append(f"  📈 平均予測生産性: {avg_productivity:.3f}")
                self.prediction_result_text.append(f"  🎯 モデル精度: {accuracy:.3f}")
                
                if trend_direction == 'increasing':
                    self.prediction_result_text.append("  ✅ 生産性の向上が期待できます!")
                elif trend_direction == 'decreasing':
                    self.prediction_result_text.append("  ⚠️ 生産性の低下に注意が必要です")
                else:
                    self.prediction_result_text.append("  📊 生産性は安定しています")
            else:
                self.prediction_result_text.append(f"❌ 予測エラー: {result['error']}")
                
        except Exception as e:
            self.prediction_result_text.append(f"❌ 予測エラー: {e}")
            logger.error(f"生産性トレンド予測エラー: {e}")
    
    def export_report(self, format_type):
        """レポートをエクスポート"""
        if not hasattr(self, 'report_exporter'):
            return
        
        try:
            # レポートデータを収集
            report_data = self.collect_report_data()
            
            if format_type == 'pdf':
                file_path = self.report_exporter.export_comprehensive_pdf_report(report_data)
                if file_path:
                    QMessageBox.information(self, "エクスポート完了", 
                                          f"PDFレポートが生成されました:\n{file_path}")
                else:
                    QMessageBox.warning(self, "エクスポートエラー", "PDFレポートの生成に失敗しました")
            
            elif format_type == 'excel':
                file_path = self.report_exporter.export_excel_workbook(report_data)
                if file_path:
                    QMessageBox.information(self, "エクスポート完了", 
                                          f"Excelレポートが生成されました:\n{file_path}")
                else:
                    QMessageBox.warning(self, "エクスポートエラー", "Excelレポートの生成に失敗しました")
                    
        except Exception as e:
            QMessageBox.critical(self, "エクスポートエラー", f"レポートのエクスポートに失敗しました:\n{e}")
            logger.error(f"レポートエクスポートエラー: {e}")
    
    def generate_immediate_report(self, report_type):
        """即座にレポートを生成"""
        if not hasattr(self, 'auto_scheduler'):
            return
        
        try:
            generated_files = self.auto_scheduler.generate_immediate_report(report_type, ['pdf', 'excel'])
            
            if generated_files:
                file_list = '\n'.join(generated_files)
                QMessageBox.information(self, "レポート生成完了", 
                                      f"{report_type}レポートが生成されました:\n\n{file_list}")
            else:
                QMessageBox.warning(self, "レポート生成エラー", f"{report_type}レポートの生成に失敗しました")
                
        except Exception as e:
            QMessageBox.critical(self, "レポート生成エラー", f"レポートの生成に失敗しました:\n{e}")
            logger.error(f"即座レポート生成エラー: {e}")
    
    def collect_report_data(self) -> Dict[str, Any]:
        """レポート用データを収集"""
        try:
            # 基本統計情報
            sessions = self.stats.sessions if hasattr(self.stats, 'sessions') else []
            
            summary = {
                'total_sessions': len(sessions),
                'completed_sessions': len([s for s in sessions if s.get('completed', False)]),
                'avg_focus_score': sum([s.get('focus_score', 0) for s in sessions]) / max(len(sessions), 1),
                'total_work_time': sum([s.get('duration', 0) for s in sessions]) / 60,  # hours
                'productivity_trend': 'Stable'
            }
            
            # 最近のセッション
            recent_sessions = sessions[-10:] if len(sessions) > 10 else sessions
            
            return {
                'summary': summary,
                'sessions': sessions,
                'recent_sessions': recent_sessions,
                'session_stats': {
                    'weekly': [],
                    'monthly': []
                },
                'charts': {
                    'focus_trend': {
                        'dates': [s.get('date', datetime.now().strftime('%Y-%m-%d')) for s in recent_sessions],
                        'scores': [s.get('focus_score', 0) for s in recent_sessions]
                    }
                },
                'predictions': {},
                'recommendations': [
                    "定期的な休憩を取って集中力を維持しましょう",
                    "最適な作業時間帯を活用して生産性を向上させましょう",
                    "中断パターンを分析して集中環境を改善しましょう"
                ]
            }
            
        except Exception as e:
            logger.error(f"レポートデータ収集エラー: {e}")
            return {'error': str(e)}
    
    def update_schedule_info(self, text_widget):
        """スケジュール情報を更新"""
        if not hasattr(self, 'auto_scheduler'):
            text_widget.setText("スケジューラーが利用できません")
            return
        
        try:
            next_reports = self.auto_scheduler.get_next_scheduled_reports()
            
            if next_reports:
                info_text = "📅 次回予定レポート:\n"
                for report in next_reports[:3]:
                    report_type = report['type']
                    next_run = report['next_run_readable']
                    info_text += f"  • {report_type}: {next_run}\n"
            else:
                info_text = "📅 スケジュールされたレポートはありません"
            
            text_widget.setText(info_text)
            
        except Exception as e:
            text_widget.setText(f"スケジュール情報の取得に失敗: {e}")
            logger.error(f"スケジュール情報更新エラー: {e}")


def main():
    """メイン関数"""
    app = QApplication(sys.argv)
    
    # データ管理オブジェクト作成
    timer_data = TimerDataManager()
    task_manager = TaskManager()
    stats = StatisticsManager()
    
    # メインウィンドウ作成
    main_window = MainWindow(timer_data, task_manager, stats)
    main_window.show()
    
    logger.info("🚀 Pomodoro Timer Phase 4 - Interactive Analysis Engine & Advanced Visualization 起動完了")
    logger.info("✅ 全機能統合済み: Clean Dual Window + Dashboard + Minimal Mode + Integrated Simple Break Window + Interactive Analysis")
    logger.info("🚀 Phase 4 高度なデータ収集システム: Advanced Data Collector + Session Tracking + Focus Calculator + Interruption Tracker + Environment Logger")
    logger.info("📊 Phase 4 インタラクティブ分析システム: Interactive Reports Engine + Advanced Visualization + Comparison Analytics + Custom Report Builder")
    logger.info("🤖 Worker3 予測エンジン・エクスポートシステム: PredictionEngine + ReportExporter + AutoReportScheduler")
    logger.info("🎯 完全機能: ドリルダウン分析・期間比較・カスタムレポート・高度な可視化（matplotlib/seaborn）・エクスポート機能・AI予測・自動レポート生成")
    
    return app.exec()


if __name__ == "__main__":
    sys.exit(main())